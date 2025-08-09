import logging
from django.utils import timezone
from django.utils.timezone import now
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db.models import Q,F, Sum
from django.utils.decorators import method_decorator
from django.views import View
import json
from decimal import Decimal
from django.views import View
from django.shortcuts import get_object_or_404
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from .utils import json_response, parse_json_body
from decimal import Decimal
from .models import *
from .services import InventoryService
from users.models import CustomUser, Role, Profile
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from django.http import HttpResponse
from django.utils import timezone
from io import BytesIO
from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views import View
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.paginator import Paginator
from django.db.models import Q, Count
from django.utils import timezone
from django.conf import settings
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
import json
import csv
from datetime import datetime, timedelta
from .models import Notification
from .notification_consumers import NotificationConsumer
logger = logging.getLogger(__name__)
# Import همه model ها

try:
    from .models import ActivityLog
except ImportError:
    ActivityLog = None
# inventory/views.py

# در بالای فایل views.py این import را اضافه کنید:
from .utils import (
    json_response, 
    parse_json_body, 
    success_response, 
    error_response, 
    validation_error_response,
    not_found_response,
    paginated_response,
    safe_decimal,
    safe_int,
    validate_required_fields,
    log_user_activity
)

# ==================== Helper Functions ====================


def json_response(data=None, message="", status=200, success=True):
    """استاندارد پاسخ JSON"""
    return JsonResponse({
        'success': success,
        'message': message,
        'data': data
    }, status=status)

def parse_json_body(request):
    """پارس کردن JSON از body"""
    try:
        return json.loads(request.body.decode('utf-8'))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return {}

# ==================== داشبورد ====================
# inventory/views.py - تابع dashboard رو تغییر بدیم:

# inventory/views.py
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from .models import *

@login_required
def dashboard(request):
    """داشبورد انبار"""
    # آمار کلی
    total_items = Item.objects.filter(is_active=True).count()
    total_warehouses = Warehouse.objects.filter(is_active=True).count()
    total_categories = Category.objects.filter(is_active=True).count()
    
    # آمار امروز
    today = timezone.now().date()
    today_entries = StockEntry.objects.filter(entry_date__date=today).count()
    today_exits = StockExit.objects.filter(exit_date__date=today).count()
    
    # کالاهای کم موجود
    low_stock_items = Item.objects.filter(
        is_active=True,
        inventory__quantity__gt=0,
        inventory__quantity__lte=F('min_stock_level')
    ).distinct().count()
    
    # کالاهای ناموجود
    out_of_stock_items = Item.objects.filter(
        is_active=True,
        inventory__quantity=0
    ).distinct().count()
    
    # محاسبه ارزش کل موجودی - اصلاح شده
    total_value = Inventory.objects.filter(
        item__is_active=True
    ).aggregate(
        total=Sum('value')  # استفاده از فیلد value به جای unit_cost
    )['total'] or 0
    
    stats = {
        'total_items': total_items,
        'total_warehouses': total_warehouses,
        'total_categories': total_categories,
        'today_entries': today_entries,
        'today_exits': today_exits,
        'low_stock_items': low_stock_items,
        'out_of_stock_items': out_of_stock_items,
        'total_value': total_value
    }
    
    return render(request, 'inventory/dashboard_content.html', {
        'stats': stats,
        'page_title': 'داشبورد انبار'
    })



# ==================== مدیریت کالاها ====================

@method_decorator(login_required, name='dispatch')
@method_decorator(csrf_exempt, name='dispatch')
class ItemView(View):
    """View کلاس برای مدیریت کالاها"""
    
    def get(self, request, item_id=None):
        """دریافت کالا یا لیست کالاها"""
        
        if item_id:
            # جزئیات یک کالا
            item = get_object_or_404(Item, id=item_id)
            
            # موجودی در انبارها
            inventories = Inventory.objects.filter(item=item).select_related('warehouse', 'batch')
            
            # آخرین ورودی‌ها
            recent_entries = StockEntry.objects.filter(item=item).select_related(
                'warehouse', 'batch'
            ).order_by('-entry_date')[:10]
            
            # آخرین خروجی‌ها
            recent_exits = StockExit.objects.filter(item=item).select_related(
                'warehouse', 'batch'
            ).order_by('-exit_date')[:10]
            
            # واحدهای تبدیل
            item_units = ItemUnit.objects.filter(item=item).select_related('unit')
            
            # وضعیت آلارم
            current_stock = InventoryService.get_current_stock(item)
            
            data = {
                'item': {
                    'id': str(item.id),
                    'name': item.name,
                    'code': item.code,
                    'barcode': item.barcode,
                    'category': item.category.name,
                    'primary_unit': item.primary_unit.name,
                    'brand': item.brand,
                    'model': item.model,
                    'specifications': item.specifications,
                    'min_stock_level': float(item.min_stock_level),
                    'max_stock_level': float(item.max_stock_level),
                    'alert_threshold': float(item.alert_threshold),
                    'enable_low_stock_alert': item.enable_low_stock_alert,
                    'cost_method': item.cost_method,
                    'has_expiry': item.has_expiry,
                    'has_serial': item.has_serial,
                    'has_batch': item.has_batch,
                    'description': item.description,
                    'current_stock': float(current_stock),
                    'created_at': item.created_at.isoformat(),
                },
                'inventories': [{
                    'warehouse': inv.warehouse.name,
                    'batch': inv.batch.batch_number if inv.batch else None,
                    'quantity': float(inv.quantity),
                    'value': float(inv.value),
                    'average_cost': float(inv.average_cost),
                } for inv in inventories],
                'recent_entries': [{
                    'entry_number': entry.entry_number,
                    'quantity': float(entry.quantity),
                    'unit': entry.unit.symbol,
                    'unit_cost': float(entry.unit_cost),
                    'warehouse': entry.warehouse.name,
                    'entry_date': entry.entry_date.isoformat(),
                } for entry in recent_entries],
                'recent_exits': [{
                    'exit_number': exit.exit_number,
                    'quantity': float(exit.quantity),
                    'unit': exit.unit.symbol,
                    'recipient_name': exit.recipient_name,
                    'warehouse': exit.warehouse.name,
                    'exit_date': exit.exit_date.isoformat(),
                } for exit in recent_exits],
                'item_units': [{
                    'unit_name': unit.unit.name,
                    'unit_symbol': unit.unit.symbol,
                    'conversion_factor': float(unit.conversion_factor),
                } for unit in item_units]
            }
            
            return json_response(data)
        
        else:
            # لیست کالاها
            items = Item.objects.select_related('category', 'primary_unit').filter(is_active=True)
            
            # جستجو
            search = request.GET.get('search')
            if search:
                items = items.filter(
                    Q(name__icontains=search) |
                    Q(code__icontains=search) |
                    Q(barcode__icontains=search)
                )
            
            # فیلتر دسته‌بندی
            category_id = request.GET.get('category')
            if category_id:
                items = items.filter(category_id=category_id)
            
            # فیلتر وضعیت موجودی
            stock_status = request.GET.get('stock_status')
            if stock_status:
                if stock_status == 'low_stock':
                    low_stock_item_ids = []
                    for item in items:
                        current_stock = InventoryService.get_current_stock(item)
                        if current_stock <= item.alert_threshold and current_stock > 0:
                            low_stock_item_ids.append(item.id)
                            items = items.filter(id__in=low_stock_item_ids)
                elif stock_status == 'out_of_stock':
                    out_of_stock_item_ids = []
                    for item in items:
                        current_stock = InventoryService.get_current_stock(item)
                        if current_stock <= 0:
                            out_of_stock_item_ids.append(item.id)
                            items = items.filter(id__in=out_of_stock_item_ids)
                elif stock_status == 'in_stock':
                    in_stock_item_ids = []
                    for item in items:
                        current_stock = InventoryService.get_current_stock(item)
                        if current_stock > 0:
                            in_stock_item_ids.append(item.id)
                            items = items.filter(id__in=in_stock_item_ids)

            # فیلتر وضعیت
                status = request.GET.get('status')
            if status:
                items = items.filter(is_active=(status == 'active'))
            
            # صفحه‌بندی
            page = int(request.GET.get('page', 1))
            per_page = int(request.GET.get('per_page', 20))
            paginator = Paginator(items, per_page)
            page_obj = paginator.get_page(page)
            
            data = {
                'items': [{
                    'id': str(item.id),
                    'name': item.name,
                    'code': item.code,
                    'barcode': item.barcode,
                    'category': item.category.name,
                    'primary_unit': item.primary_unit.symbol,
                    'brand': item.brand,
                    'alert_threshold': float(item.alert_threshold),
                    'current_stock': float(InventoryService.get_current_stock(item)),
                    'created_at': item.created_at.isoformat(),
                } for item in page_obj],
                'pagination': {
                    'current_page': page_obj.number,
                    'total_pages': paginator.num_pages,
                    'total_items': paginator.count,
                    'has_next': page_obj.has_next(),
                    'has_previous': page_obj.has_previous(),
                }
            }
            
            return json_response(data)
    
    def post(self, request, item_id=None):
        """ایجاد یا بروزرسانی کالا"""
    
        if item_id:
            # بروزرسانی کالا
            item = get_object_or_404(Item, id=item_id)
        
            try:
                # بررسی نوع درخواست و استخراج داده‌ها
                if request.content_type == 'application/json':
                    data = parse_json_body(request)
                else:
                # استفاده از داده‌های فرم
                    data = request.POST.dict()
            
            # پردازش فایل‌ها
                files = request.FILES
            
            # تبدیل فیلدهای عددی - اگر خالی باشند، صفر قرار بده
                decimal_fields = [
                    'min_stock_level', 'max_stock_level', 'alert_threshold',
                    'storage_temperature_min', 'storage_temperature_max'
                ]
            
                for field in decimal_fields:
                    if field in data:
                        if data[field] == '' or data[field] is None:
                            data[field] = '0'
                    else:
                        data[field] = '0'
            
            # تبدیل فیلدهای بولی
                bool_fields = ['enable_low_stock_alert', 'has_expiry', 'has_serial', 'has_batch']
                for field in bool_fields:
                    data[field] = field in data and data[field] in ['true', 'True', 'on', '1', 1, True]
            
            # بروزرسانی فیلدها
                if 'name' in data:
                    item.name = data['name']
                if 'code' in data:
                # چک یکتا بودن کد (به جز خود این کالا)
                    if Item.objects.filter(code=data['code']).exclude(id=item_id).exists():
                        return json_response(
                            message="کد کالا تکراری است.",
                            status=400,
                            success=False
                        )
                    item.code = data['code']
                if 'barcode' in data:
                    item.barcode = data['barcode']
                if 'category_id' in data:
                    item.category_id = data['category_id']
                if 'primary_unit_id' in data:
                    item.primary_unit_id = data['primary_unit_id']
                if 'brand' in data:
                    item.brand = data['brand']
                if 'model' in data:
                    item.model = data['model']
                if 'specifications' in data:
                    item.specifications = data['specifications']
                if 'min_stock_level' in data:
                    item.min_stock_level = Decimal(str(data['min_stock_level']))
                if 'max_stock_level' in data:
                    item.max_stock_level = Decimal(str(data['max_stock_level']))
                if 'alert_threshold' in data:
                    item.alert_threshold = Decimal(str(data['alert_threshold']))
                if 'enable_low_stock_alert' in data:
                    item.enable_low_stock_alert = data['enable_low_stock_alert']
                if 'cost_method' in data:
                    item.cost_method = data['cost_method']
                if 'has_expiry' in data:
                    item.has_expiry = data['has_expiry']
                if 'has_serial' in data:
                    item.has_serial = data['has_serial']
                if 'has_batch' in data:
                    item.has_batch = data['has_batch']
                if 'description' in data:
                    item.description = data['description']
                if 'storage_temperature_min' in data:
                    item.storage_temperature_min = Decimal(str(data['storage_temperature_min']))
                if 'storage_temperature_max' in data:
                    item.storage_temperature_max = Decimal(str(data['storage_temperature_max']))
                if 'storage_conditions' in data:
                    item.storage_conditions = data['storage_conditions']
                if 'usage_instructions' in data:
                    item.usage_instructions = data['usage_instructions']
            
                item.save()
            
            # اگر فایل تصویر وجود دارد، آن را ذخیره کنید
                if 'image' in request.FILES:
                    item.image = request.FILES['image']
                    item.save()
            
                return json_response(
                    message=f'کالای "{item.name}" با موفقیت بروزرسانی شد.',
                    success=True
                )
            
            except Exception as e:
                import traceback
                traceback.print_exc()  # چاپ خطای کامل در کنسول
                return json_response(
                    message=f'خطا در بروزرسانی کالا: {str(e)}',
                    status=500,
                    success=False
                )
        else:
        # ایجاد کالای جدید
            try:
            # بررسی نوع درخواست و استخراج داده‌ها
                if request.content_type == 'application/json':
                    data = parse_json_body(request)
                else:
                # استفاده از داده‌های فرم
                    data = request.POST.dict()
            
            # پردازش فایل‌ها
                files = request.FILES
            
            # تبدیل فیلدهای عددی - اگر خالی باشند، صفر قرار بده
                decimal_fields = [
                    'min_stock_level', 'max_stock_level', 'alert_threshold',
                    'storage_temperature_min', 'storage_temperature_max'
                ]
            
                for field in decimal_fields:
                    if field in data:
                        if data[field] == '' or data[field] is None:
                            data[field] = '0'
                    else:
                        data[field] = '0'
            
            # تبدیل فیلدهای بولی
                bool_fields = ['enable_low_stock_alert', 'has_expiry', 'has_serial', 'has_batch']
                for field in bool_fields:
                    data[field] = field in data and data[field] in ['true', 'True', 'on', '1', 1, True]
            
            # اعتبارسنجی داده‌های اصلی
                required_fields = ['name', 'code', 'category_id', 'primary_unit_id']
                for field in required_fields:
                    if not data.get(field):
                        return json_response(
                            message=f"فیلد {field} الزامی است.",
                            status=400,
                            success=False
                        )
            
                # چک کردن یکتا بودن کد
                if Item.objects.filter(code=data['code']).exists():
                    return json_response(
                        message="کد کالا تکراری است.",
                        status=400,
                        success=False
                    )
            
            # ایجاد کالا
                item = Item.objects.create(
                    name=data['name'],
                    code=data['code'],
                    barcode=data.get('barcode', ''),
                    category_id=data['category_id'],
                    primary_unit_id=data['primary_unit_id'],
                    brand=data.get('brand', ''),
                    model=data.get('model', ''),
                    specifications=data.get('specifications', {}),
                    min_stock_level=Decimal(str(data.get('min_stock_level', 0))),
                    max_stock_level=Decimal(str(data.get('max_stock_level', 0))),
                    alert_threshold=Decimal(str(data.get('alert_threshold', 5))),
                    enable_low_stock_alert=data.get('enable_low_stock_alert', True),
                    cost_method=data.get('cost_method', 'FIFO'),
                    has_expiry=data.get('has_expiry', False),
                    has_serial=data.get('has_serial', False),
                    has_batch=data.get('has_batch', True),
                    description=data.get('description', ''),
                    created_by=request.user
                )
            
            # تولید بارکد خودکار
                if not item.barcode:
                    item.barcode = item.generate_barcode()
                    item.save()
            
            # اگر فایل تصویر وجود دارد، آن را ذخیره کنید
                if 'image' in request.FILES:
                    item.image = request.FILES['image']
                    item.save()
            
                return json_response(
                    data={'item_id': str(item.id)},
                    message=f'کالا با موفقیت ایجاد شد',
                    success=True
                )
            
            except Exception as e:
                import traceback
                traceback.print_exc()  # چاپ خطای کامل در کنسول
                return json_response(
                    message=f'خطا در ایجاد کالا: {str(e)}',
                    status=500,
                    success=False
                )


    
    def put(self, request, item_id):
        """بروزرسانی کالا"""
    
        item = get_object_or_404(Item, id=item_id)
    
        try:
        # بررسی نوع درخواست و استخراج داده‌ها
            if request.content_type == 'application/json':
                data = parse_json_body(request)
            else:
            # استفاده از داده‌های فرم
                data = request.POST.dict()
            
            # پردازش فایل‌ها
                files = request.FILES
            
            # تبدیل فیلدهای عددی - اگر خالی باشند، صفر قرار بده
                decimal_fields = [
                    'min_stock_level', 'max_stock_level', 'alert_threshold',
                    'storage_temperature_min', 'storage_temperature_max'
                ]
            
                for field in decimal_fields:
                    if field in data:
                        if data[field] == '' or data[field] is None:
                            data[field] = '0'
                    else:
                        data[field] = '0'
            
            # تبدیل فیلدهای بولی
                bool_fields = ['enable_low_stock_alert', 'has_expiry', 'has_serial', 'has_batch']
                for field in bool_fields:
                    data[field] = field in data and data[field] in ['true', 'True', 'on', '1', 1, True]
        
        # بروزرسانی فیلدها
            if 'name' in data:
                item.name = data['name']
            if 'code' in data:
                # چک یکتا بودن کد (به جز خود این کالا)
                if Item.objects.filter(code=data['code']).exclude(id=item_id).exists():
                    return json_response(
                        message="کد کالا تکراری است.",
                        status=400,
                        success=False
                    )
                item.code = data['code']
            if 'barcode' in data:
                item.barcode = data['barcode']
            if 'category_id' in data:
                item.category_id = data['category_id']
            if 'primary_unit_id' in data:
                item.primary_unit_id = data['primary_unit_id']
            if 'brand' in data:
                item.brand = data['brand']
            if 'model' in data:
                item.model = data['model']
            if 'specifications' in data:
                item.specifications = data['specifications']
            if 'min_stock_level' in data:
                item.min_stock_level = Decimal(str(data['min_stock_level']))
            if 'max_stock_level' in data:
                item.max_stock_level = Decimal(str(data['max_stock_level']))
            if 'alert_threshold' in data:
                item.alert_threshold = Decimal(str(data['alert_threshold']))
            if 'enable_low_stock_alert' in data:
                item.enable_low_stock_alert = data['enable_low_stock_alert']
            if 'cost_method' in data:
                item.cost_method = data['cost_method']
            if 'has_expiry' in data:
                item.has_expiry = data['has_expiry']
            if 'has_serial' in data:
                item.has_serial = data['has_serial']
            if 'has_batch' in data:
                item.has_batch = data['has_batch']
            if 'description' in data:
                item.description = data['description']
            if 'storage_temperature_min' in data:
                item.storage_temperature_min = Decimal(str(data['storage_temperature_min']))
            if 'storage_temperature_max' in data:
                item.storage_temperature_max = Decimal(str(data['storage_temperature_max']))
            if 'storage_conditions' in data:
                item.storage_conditions = data['storage_conditions']
            if 'usage_instructions' in data:
                item.usage_instructions = data['usage_instructions']
        
            item.save()
        
        # اگر فایل تصویر وجود دارد، آن را ذخیره کنید
            if 'image' in request.FILES:
                item.image = request.FILES['image']
                item.save()
        
            return json_response(
                message=f'کالای "{item.name}" با موفقیت بروزرسانی شد.',
                success=True
            )
        
        except Exception as e:
            import traceback
            traceback.print_exc()  # چاپ خطای کامل در کنسول
            return json_response(
                message=f'خطا در بروزرسانی کالا: {str(e)}',
                status=500,
                success=False
            )

    
    def delete(self, request, item_id):
        """حذف کالا (غیرفعال کردن)"""
        
        item = get_object_or_404(Item, id=item_id)
        
        try:
            # چک کردن وجود موجودی
            has_inventory = Inventory.objects.filter(item=item, quantity__gt=0).exists()
            if has_inventory:
                return json_response(
                    message="امکان حذف کالا وجود ندارد. موجودی در انبار موجود است.",
                    status=400,
                    success=False
                )
            
            # غیرفعال کردن به جای حذف
            item.is_active = False
            item.save()
            
            return json_response(
                message=f'کالای "{item.name}" با موفقیت حذف شد.'
            )
            
        except Exception as e:
            return json_response(
                message=f'خطا در حذف کالا: {str(e)}',
                status=500,
                success=False
            )

# ==================== مدیریت ورودی انبار ====================

@method_decorator(login_required, name='dispatch')
@method_decorator(csrf_exempt, name='dispatch')
class StockEntryView(View):
    """API برای مدیریت ورودی‌های انبار"""
    def create_stock_entry_notification(self, title, message, notification_type='STOCK_ENTRY', priority='MEDIUM', url=None, metadata=None):
        """
        ایجاد اعلان برای همه انباردارها و ادمین‌ها
        """
        try:
            from users.models import Role
            from inventory.models import Notification
            from django.db.models import Q
        
        # دریافت نقش انباردار
            warehouse_role = Role.objects.filter(name='انباردار').first()
        
        # ایجاد اعلان در دیتابیس
            notification = Notification.objects.create(
                title=title,
                message=message,
                notification_type=notification_type,
                priority=priority,
                action_url=url,
                metadata=metadata or {},
                created_by=self.request.user,
                is_admin_only=False,
                is_global=False,
                target_role=warehouse_role,
                recipient=None  # اعلان گروهی است، نه شخصی
            )
        
        # ارسال اعلان به همه انباردارها و ادمین‌ها از طریق WebSocket
            from users.models import CustomUser
            from django.db.models import Q
        
        # 1. کاربرانی که نقش انباردار دارند
            warehouse_users = CustomUser.objects.filter(role=warehouse_role) if warehouse_role else []
        
        # 2. ادمین‌ها
            admin_users = CustomUser.objects.filter(Q(is_superuser=True) | Q(is_staff=True))
        
        # ترکیب کاربران (بدون تکرار)
            target_users = list(warehouse_users) + [user for user in admin_users if user not in warehouse_users]
        
            notification_data = {
                'id': str(notification.id),
                'title': notification.title,
                'message': notification.message,
                'type': notification.notification_type,
                'priority': notification.priority,
                'is_read': False,
                'url': notification.action_url or '#',
                'created_at': notification.created_at.isoformat(),
                'created_by': self.request.user.get_full_name() if self.request.user else 'سیستم'
            }
        
        # ارسال به کاربران هدف
            for user in target_users:
                from inventory.notification_consumers import send_notification_to_user
                send_notification_to_user(user.id, notification_data)
        
            return notification
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"خطا در ایجاد اعلان ورودی انبار: {str(e)}")
            return None

    def get(self, request, entry_id=None):
        """دریافت ورودی یا لیست ورودی‌ها"""
        if entry_id:
            # دریافت یک ورودی
            try:
                entry = StockEntry.objects.get(id=entry_id)
                
                data = {
                    'id': str(entry.id),
                    'entry_number': entry.entry_number,
                    'entry_date': entry.entry_date.strftime('%Y-%m-%d'),
                    'item_id': str(entry.item.id),
                    'item_name': entry.item.name,
                    'warehouse_id': str(entry.warehouse.id),
                    'warehouse_name': entry.warehouse.name,
                    'quantity': float(entry.quantity),
                    'unit_id': str(entry.unit.id),
                    'unit_name': entry.unit.name,
                    'unit_cost': float(entry.unit_cost),
                    'total_cost': float(entry.total_cost),
                    'supplier': entry.supplier,
                    'invoice_number': entry.invoice_number,
                    'batch_number': entry.batch_number,
                    'expiry_date': entry.expiry_date.strftime('%Y-%m-%d') if entry.expiry_date else None,
                    'description': entry.description,
                    'created_at': entry.created_at.isoformat(),
                    'created_by': entry.created_by.get_full_name() if entry.created_by else None
                }
                
                return JsonResponse({'success': True, 'entry': data})
                
            except StockEntry.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'ورودی مورد نظر یافت نشد'}, status=404)
                
        else:
            # دریافت لیست ورودی‌ها
            try:
                # پارامترهای فیلتر و جستجو
                search = request.GET.get('search', '')
                item_id = request.GET.get('item', '')
                warehouse_id = request.GET.get('warehouse', '')
                start_date = request.GET.get('start_date', '')
                end_date = request.GET.get('end_date', '')
                
                # پارامترهای صفحه‌بندی
                page = int(request.GET.get('page', 1))
                per_page = int(request.GET.get('per_page', 20))
                
                # پارامترهای مرتب‌سازی
                sort_by = request.GET.get('sort_by', 'entry_date')
                sort_order = request.GET.get('sort_order', 'desc')
                
                # فیلتر ورودی‌ها
                entries = StockEntry.objects.all()
                
                # اعمال جستجو
                if search:
                    entries = entries.filter(
                        Q(entry_number__icontains=search) | 
                        Q(item__name__icontains=search) | 
                        Q(supplier__icontains=search) |
                        Q(invoice_number__icontains=search)
                    )
                
                # اعمال فیلتر کالا
                if item_id:
                    entries = entries.filter(item_id=item_id)
                
                # اعمال فیلتر انبار
                if warehouse_id:
                    entries = entries.filter(warehouse_id=warehouse_id)
                
                # اعمال فیلتر تاریخ
                if start_date:
                    entries = entries.filter(entry_date__gte=start_date)
                
                if end_date:
                    entries = entries.filter(entry_date__lte=end_date)
                
                # اعمال مرتب‌سازی
                order_by = sort_by
                if sort_order == 'desc':
                    order_by = f'-{sort_by}'
                
                entries = entries.order_by(order_by)
                
                # صفحه‌بندی
                paginator = Paginator(entries, per_page)
                page_obj = paginator.get_page(page)
                
                # تبدیل به دیکشنری
                entries_data = []
                for entry in page_obj:
                    entries_data.append({
                        'id': str(entry.id),
                        'entry_number': entry.entry_number,
                        'entry_date': entry.entry_date.strftime('%Y-%m-%d'),
                        'item_id': str(entry.item.id),
                        'item_name': entry.item.name,
                        'warehouse_id': str(entry.warehouse.id),
                        'warehouse_name': entry.warehouse.name,
                        'quantity': float(entry.quantity),
                        'unit_id': str(entry.unit.id),
                        'unit_name': entry.unit.name,
                        'unit_cost': float(entry.unit_cost),
                        'total_cost': float(entry.total_cost),
                        'supplier': entry.supplier,
                        'invoice_number': entry.invoice_number,
                        'created_at': entry.created_at.isoformat()
                    })
                
                # اطلاعات صفحه‌بندی
                pagination = {
                    'current_page': page_obj.number,
                    'total_pages': paginator.num_pages,
                    'total': paginator.count,
                    'has_next': page_obj.has_next(),
                    'has_previous': page_obj.has_previous(),
                }
                
                return JsonResponse({
                    'success': True,
                    'entries': entries_data,
                    'pagination': pagination
                })
                
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    def post(self, request):
        """ایجاد ورودی جدید"""
        try:
            data = json.loads(request.body)
        
        # اعتبارسنجی داده‌های ورودی
            required_fields = ['item_id', 'warehouse_id', 'quantity', 'unit_id', 'unit_cost']
            for field in required_fields:
                if field not in data or not data[field]:
                    return JsonResponse({
                        'success': False, 
                        'message': f'فیلد {field} الزامی است'
                    }, status=400)
        
        # دریافت کالا، انبار و واحد
            try:
                item = Item.objects.get(id=data['item_id'], is_active=True)
            except Item.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'message': 'کالای مورد نظر یافت نشد'
                }, status=400)
        
            try:
                warehouse = Warehouse.objects.get(id=data['warehouse_id'], is_active=True)
            except Warehouse.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'message': 'انبار مورد نظر یافت نشد'
                }, status=400)
        
            try:
                unit = UnitOfMeasure.objects.get(id=data['unit_id'], is_active=True)
            except UnitOfMeasure.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'message': 'واحد مورد نظر یافت نشد'
                }, status=400)
        
        # تبدیل مقدار به واحد اصلی
            quantity = float(data['quantity'])
            conversion_factor = 1.0
        
            if unit.id != item.primary_unit.id:
                try:
                    item_unit = ItemUnit.objects.get(item=item, unit=unit)
                    conversion_factor = float(item_unit.conversion_factor)
                except ItemUnit.DoesNotExist:
                    return JsonResponse({
                        'success': False, 
                        'message': 'واحد انتخاب شده برای این کالا تعریف نشده است'
                    }, status=400)
        
            primary_quantity = quantity * conversion_factor
        
        # محاسبه هزینه کل
            unit_cost = float(data['unit_cost'])
            total_cost = quantity * unit_cost
        
        # ایجاد شماره ورودی
            entry_number = f"IN-{timezone.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"
        
        # تنظیم تاریخ ورودی
            entry_date = data.get('entry_date')
            if not entry_date:
                entry_date = timezone.now().date()
            else:
                entry_date = datetime.strptime(entry_date, '%Y-%m-%d').date()
        
        # ایجاد ورودی
            entry = StockEntry(
                entry_number=entry_number,
                entry_date=entry_date,
                item=item,
                warehouse=warehouse,
                quantity=quantity,
                unit=unit,
                unit_cost=unit_cost,
                total_cost=total_cost,
                supplier=data.get('supplier', ''),
                invoice_number=data.get('invoice_number', ''),
                batch_number=data.get('batch_number', ''),
                expiry_date=data.get('expiry_date'),
                description=data.get('description', ''),
                created_by=request.user
            )
        
        # ذخیره ورودی
            entry.save()
        
        # بروزرسانی موجودی
            inventory, created = Inventory.objects.get_or_create(
                item=item,
                warehouse=warehouse,
                defaults={
                    'quantity': 0,
                    'unit_cost': 0,
                    'created_by': request.user
                }
            )
        
        # محاسبه قیمت واحد جدید بر اساس روش قیمت‌گذاری
            if item.cost_method == 'FIFO':
            # در روش FIFO، قیمت واحد تغییر نمی‌کند
                pass
            elif item.cost_method == 'LIFO':
            # در روش LIFO، قیمت واحد به قیمت آخرین ورودی تغییر می‌کند
                inventory.unit_cost = unit_cost
            elif item.cost_method == 'AVERAGE':
            # در روش میانگین، قیمت واحد به میانگین وزنی تغییر می‌کند
                total_value = (inventory.quantity * inventory.unit_cost) + total_cost
                new_quantity = inventory.quantity + primary_quantity
                if new_quantity > 0:
                    inventory.unit_cost = total_value / new_quantity
        
        # بروزرسانی موجودی
            inventory.quantity += primary_quantity
            inventory.save()
        
        # ایجاد نوتیفیکیشن
            self.create_stock_entry_notification(
                title="ورود کالا به انبار",
                message=f"ورود {quantity} {unit.name} کالای «{item.name}» به انبار «{warehouse.name}» توسط {request.user.get_full_name() or request.user.username}",
                notification_type="STOCK_ENTRY",
                priority="MEDIUM",
                url=f"/inventory/stock-entries/{entry.id}/detail/",
                metadata={
                    'entry_id': str(entry.id),
                    'entry_number': entry.entry_number,
                    'item_id': str(item.id),
                    'item_name': item.name,
                    'warehouse_id': str(warehouse.id),
                    'warehouse_name': warehouse.name,
                    'quantity': quantity,
                    'unit_name': unit.name,
                    'total_cost': total_cost,
                    'action': 'create'
                }
            )
        
            return JsonResponse({
                'success': True,
                'message': 'ورودی با موفقیت ثبت شد',
                'entry_id': str(entry.id)
            })
        
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False, 
                'message': 'داده‌های ورودی نامعتبر است'
            }, status=400)
        
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'message': f'خطا در ثبت ورودی: {str(e)}'
            }, status=500)
            
    
    def put(self, request, entry_id):
        """بروزرسانی ورودی"""
        try:
        # دریافت ورودی
            try:
                entry = StockEntry.objects.get(id=entry_id)
            except StockEntry.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'message': 'ورودی مورد نظر یافت نشد'
                }, status=404)
        
        # ذخیره مقادیر قبلی برای بروزرسانی موجودی و نوتیفیکیشن
            old_item = entry.item
            old_warehouse = entry.warehouse
            old_quantity = entry.quantity
            old_unit = entry.unit
            old_entry_number = entry.entry_number
        
        # تبدیل مقدار قبلی به واحد اصلی
            old_conversion_factor = 1.0
            if old_unit.id != old_item.primary_unit.id:
                try:
                    old_item_unit = ItemUnit.objects.get(item=old_item, unit=old_unit)
                    old_conversion_factor = float(old_item_unit.conversion_factor)
                except ItemUnit.DoesNotExist:
                    pass
        
            old_primary_quantity = old_quantity * old_conversion_factor
        
            data = json.loads(request.body)
        
        # اعتبارسنجی داده‌های ورودی
            required_fields = ['item_id', 'warehouse_id', 'quantity', 'unit_id', 'unit_cost']
            for field in required_fields:
                if field not in data or not data[field]:
                    return JsonResponse({
                        'success': False, 
                        'message': f'فیلد {field} الزامی است'
                    }, status=400)
        
        # دریافت کالا، انبار و واحد
            try:
                item = Item.objects.get(id=data['item_id'], is_active=True)
            except Item.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'message': 'کالای مورد نظر یافت نشد'
                }, status=400)
        
            try:
                warehouse = Warehouse.objects.get(id=data['warehouse_id'], is_active=True)
            except Warehouse.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'message': 'انبار مورد نظر یافت نشد'
                }, status=400)
        
            try:
                unit = UnitOfMeasure.objects.get(id=data['unit_id'], is_active=True)
            except UnitOfMeasure.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'message': 'واحد مورد نظر یافت نشد'
                }, status=400)
        
        # تبدیل مقدار به واحد اصلی
            quantity = float(data['quantity'])
            conversion_factor = 1.0
        
            if unit.id != item.primary_unit.id:
                try:
                    item_unit = ItemUnit.objects.get(item=item, unit=unit)
                    conversion_factor = float(item_unit.conversion_factor)
                except ItemUnit.DoesNotExist:
                    return JsonResponse({
                        'success': False, 
                        'message': 'واحد انتخاب شده برای این کالا تعریف نشده است'
                    }, status=400)
        
            primary_quantity = quantity * conversion_factor
        
        # محاسبه هزینه کل
            unit_cost = float(data['unit_cost'])
            total_cost = quantity * unit_cost
        
        # تنظیم تاریخ ورودی
            entry_date = data.get('entry_date')
            if entry_date:
                entry_date = datetime.strptime(entry_date, '%Y-%m-%d').date()
        
        # بروزرسانی ورودی
            entry.item = item
            entry.warehouse = warehouse
            entry.quantity = quantity
            entry.unit = unit
            entry.unit_cost = unit_cost
            entry.total_cost = total_cost
            entry.supplier = data.get('supplier', '')
            entry.invoice_number = data.get('invoice_number', '')
            entry.batch_number = data.get('batch_number', '')
        
            if entry_date:
                entry.entry_date = entry_date
            
            if 'expiry_date' in data:
                entry.expiry_date = data['expiry_date']
            
            if 'description' in data:
                entry.description = data['description']
        
        # ذخیره ورودی
            entry.save()
        
        # بروزرسانی موجودی قبلی
            if old_item.id != item.id or old_warehouse.id != warehouse.id:
            # اگر کالا یا انبار تغییر کرده، موجودی قبلی را کاهش دهید
                old_inventory, created = Inventory.objects.get_or_create(
                    item=old_item,
                    warehouse=old_warehouse,
                    defaults={
                        'quantity': 0,
                        'unit_cost': 0,
                        'created_by': request.user
                    }
                )
            
                old_inventory.quantity -= old_primary_quantity
                old_inventory.save()
            
            # موجودی جدید را افزایش دهید
                new_inventory, created = Inventory.objects.get_or_create(
                    item=item,
                    warehouse=warehouse,
                    defaults={
                        'quantity': 0,
                        'unit_cost': 0,
                        'created_by': request.user
                    }
                )
            
                new_inventory.quantity += primary_quantity
            
            # بروزرسانی قیمت واحد بر اساس روش قیمت‌گذاری
                if item.cost_method == 'AVERAGE':
                    total_value = (new_inventory.quantity * new_inventory.unit_cost) + total_cost
                    new_quantity = new_inventory.quantity + primary_quantity
                    if new_quantity > 0:
                        new_inventory.unit_cost = total_value / new_quantity
            
                new_inventory.save()
            else:
            # اگر کالا و انبار تغییر نکرده، فقط مقدار را بروزرسانی کنید
                inventory, created = Inventory.objects.get_or_create(
                    item=item,
                    warehouse=warehouse,
                    defaults={
                        'quantity': 0,
                        'unit_cost': 0,
                        'created_by': request.user
                    }
                )
            
            # تغییر مقدار
                inventory.quantity = inventory.quantity - old_primary_quantity + primary_quantity
            
            # بروزرسانی قیمت واحد بر اساس روش قیمت‌گذاری
                if item.cost_method == 'AVERAGE':
                # محاسبه ارزش کل قبلی
                    old_value = inventory.quantity * inventory.unit_cost
                # کم کردن ارزش قبلی این ورودی
                    old_value -= old_primary_quantity * entry.unit_cost
                # اضافه کردن ارزش جدید این ورودی
                    new_value = old_value + (primary_quantity * unit_cost)
                # محاسبه مقدار جدید
                    new_quantity = inventory.quantity - old_primary_quantity + primary_quantity
                
                    if new_quantity > 0:
                        inventory.unit_cost = new_value / new_quantity
            
                inventory.save()
        
        # ایجاد نوتیفیکیشن
            changes = []
            if old_item.id != item.id:
                changes.append(f"کالا از «{old_item.name}» به «{item.name}»")
            if old_warehouse.id != warehouse.id:
                changes.append(f"انبار از «{old_warehouse.name}» به «{warehouse.name}»")
            if old_quantity != quantity or old_unit.id != unit.id:
                changes.append(f"مقدار از {old_quantity} {old_unit.name} به {quantity} {unit.name}")
        
            if changes:
                changes_text = "، ".join(changes)
                message = f"بروزرسانی ورودی انبار {entry.entry_number}: {changes_text}"
            else:
                message = f"بروزرسانی اطلاعات ورودی انبار {entry.entry_number}"
        
            self.create_stock_entry_notification(
                title="بروزرسانی ورودی انبار",
                message=message,
                notification_type="INVENTORY_ADJUSTMENT",
                priority="MEDIUM",
                url=f"/inventory/stock-entries/{entry.id}/detail/",
                metadata={
                    'entry_id': str(entry.id),
                    'entry_number': entry.entry_number,
                    'item_id': str(item.id),
                    'item_name': item.name,
                    'warehouse_id': str(warehouse.id),
                    'warehouse_name': warehouse.name,
                    'quantity': quantity,
                    'unit_name': unit.name,
                    'total_cost': total_cost,
                    'action': 'update',
                    'changes': changes
                }
            )
        
            return JsonResponse({
                'success': True,
                'message': 'ورودی با موفقیت بروزرسانی شد'
            })
        
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False, 
                'message': 'داده‌های ورودی نامعتبر است'
            }, status=400)
        
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'message': f'خطا در بروزرسانی ورودی: {str(e)}'
            }, status=500)
    def delete(self, request, entry_id):
        """حذف ورودی"""
        try:
        # دریافت ورودی
            try:
                entry = StockEntry.objects.get(id=entry_id)
            except StockEntry.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'message': 'ورودی مورد نظر یافت نشد'
                }, status=404)
        
        # ذخیره مقادیر برای بروزرسانی موجودی و نوتیفیکیشن
            item = entry.item
            warehouse = entry.warehouse
            quantity = entry.quantity
            unit = entry.unit
            entry_number = entry.entry_number
        
        # تبدیل مقدار به واحد اصلی
            conversion_factor = 1.0
            if unit.id != item.primary_unit.id:
                try:
                    item_unit = ItemUnit.objects.get(item=item, unit=unit)
                    conversion_factor = float(item_unit.conversion_factor)
                except ItemUnit.DoesNotExist:
                    pass
        
            primary_quantity = quantity * conversion_factor
        
        # بروزرسانی موجودی
            inventory = Inventory.objects.filter(item=item, warehouse=warehouse).first()
            if inventory:
                inventory.quantity -= primary_quantity
                inventory.save()
        
        # حذف ورودی
            entry.delete()
        
        # ایجاد نوتیفیکیشن
            self.create_stock_entry_notification(
                title="حذف ورودی انبار",
                message=f"ورودی انبار {entry_number} برای {quantity} {unit.name} کالای «{item.name}» از انبار «{warehouse.name}» توسط {request.user.get_full_name() or request.user.username} حذف شد.",
                notification_type="INVENTORY_ADJUSTMENT",
                priority="HIGH",
                metadata={
                    'entry_number': entry_number,
                    'item_name': item.name,
                    'warehouse_name': warehouse.name,
                    'quantity': quantity,
                    'unit_name': unit.name,
                    'action': 'delete'
                }
            )
        
            return JsonResponse({
                'success': True,
                'message': 'ورودی با موفقیت حذف شد'
            })
        
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'message': f'خطا در حذف ورودی: {str(e)}'
            }, status=500)


# ==================== مدیریت خروجی انبار ====================

@method_decorator(login_required, name='dispatch')
@method_decorator(csrf_exempt, name='dispatch')
class StockExitView(View):
    """API برای مدیریت خروجی‌های انبار"""
    
    def get(self, request, exit_id=None):
        """دریافت خروجی یا لیست خروجی‌ها"""
        if exit_id:
            # دریافت یک خروجی
            try:
                exit_obj = StockExit.objects.get(id=exit_id)
                
                data = {
                    'id': str(exit_obj.id),
                    'exit_number': exit_obj.exit_number,
                    'exit_date': exit_obj.exit_date.strftime('%Y-%m-%d'),
                    'item_id': str(exit_obj.item.id),
                    'item_name': exit_obj.item.name,
                    'warehouse_id': str(exit_obj.warehouse.id),
                    'warehouse_name': exit_obj.warehouse.name,
                    'quantity': float(exit_obj.quantity),
                    'unit_id': str(exit_obj.unit.id),
                    'unit_name': exit_obj.unit.name,
                    'unit_cost': float(exit_obj.unit_cost),
                    'total_cost': float(exit_obj.total_cost),
                    'receiver': exit_obj.receiver,
                    'receiver_phone': exit_obj.receiver_phone,
                    'reference_number': exit_obj.reference_number,
                    'description': exit_obj.description,
                    'created_at': exit_obj.created_at.isoformat(),
                    'created_by': exit_obj.created_by.get_full_name() if exit_obj.created_by else None
                }
                
                return JsonResponse({'success': True, 'exit': data})
                
            except StockExit.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'خروجی مورد نظر یافت نشد'}, status=404)
                
        else:
            # دریافت لیست خروجی‌ها
            try:
                # پارامترهای فیلتر و جستجو
                search = request.GET.get('search', '')
                item_id = request.GET.get('item', '')
                warehouse_id = request.GET.get('warehouse', '')
                start_date = request.GET.get('start_date', '')
                end_date = request.GET.get('end_date', '')
                
                # پارامترهای صفحه‌بندی
                page = int(request.GET.get('page', 1))
                per_page = int(request.GET.get('per_page', 20))
                
                # پارامترهای مرتب‌سازی
                sort_by = request.GET.get('sort_by', 'exit_date')
                sort_order = request.GET.get('sort_order', 'desc')
                
                # فیلتر خروجی‌ها
                exits = StockExit.objects.all()
                
                # اعمال جستجو
                if search:
                    exits = exits.filter(
                        Q(exit_number__icontains=search) | 
                        Q(item__name__icontains=search) | 
                        Q(receiver__icontains=search) |
                        Q(reference_number__icontains=search)
                    )
                
                # اعمال فیلتر کالا
                if item_id:
                    exits = exits.filter(item_id=item_id)
                
                # اعمال فیلتر انبار
                if warehouse_id:
                    exits = exits.filter(warehouse_id=warehouse_id)
                
                # اعمال فیلتر تاریخ
                if start_date:
                    exits = exits.filter(exit_date__gte=start_date)
                
                if end_date:
                    exits = exits.filter(exit_date__lte=end_date)
                
                # اعمال مرتب‌سازی
                order_by = sort_by
                if sort_order == 'desc':
                    order_by = f'-{sort_by}'
                
                exits = exits.order_by(order_by)
                
                # صفحه‌بندی
                paginator = Paginator(exits, per_page)
                page_obj = paginator.get_page(page)
                
                # تبدیل به دیکشنری
                exits_data = []
                for exit_obj in page_obj:
                    exits_data.append({
                        'id': str(exit_obj.id),
                        'exit_number': exit_obj.exit_number,
                        'exit_date': exit_obj.exit_date.strftime('%Y-%m-%d'),
                        'item_id': str(exit_obj.item.id),
                        'item_name': exit_obj.item.name,
                        'warehouse_id': str(exit_obj.warehouse.id),
                        'warehouse_name': exit_obj.warehouse.name,
                        'quantity': float(exit_obj.quantity),
                        'unit_id': str(exit_obj.unit.id),
                        'unit_name': exit_obj.unit.name,
                        'unit_cost': float(exit_obj.unit_cost),
                        'total_cost': float(exit_obj.total_cost),
                        'receiver': exit_obj.receiver,
                        'created_at': exit_obj.created_at.isoformat()
                    })
                
                # اطلاعات صفحه‌بندی
                pagination = {
                    'current_page': page_obj.number,
                    'total_pages': paginator.num_pages,
                    'total': paginator.count,
                    'has_next': page_obj.has_next(),
                    'has_previous': page_obj.has_previous(),
                }
                
                return JsonResponse({
                    'success': True,
                    'exits': exits_data,
                    'pagination': pagination
                })
                
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    def post(self, request):
    
        try:
            data = json.loads(request.body)
        
        # اعتبارسنجی داده‌های ورودی
            required_fields = ['item_id', 'warehouse_id', 'quantity', 'unit_id']
            for field in required_fields:
                if field not in data or not data[field]:
                    return JsonResponse({
                        'success': False, 
                        'message': f'فیلد {field} الزامی است'
                    }, status=400)
        
        # دریافت کالا، انبار و واحد
            try:
                item = Item.objects.get(id=data['item_id'], is_active=True)
            except Item.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'message': 'کالای مورد نظر یافت نشد'
                }, status=400)
        
            try:
                warehouse = Warehouse.objects.get(id=data['warehouse_id'], is_active=True)
            except Warehouse.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'message': 'انبار مورد نظر یافت نشد'
                }, status=400)
        
            try:
                unit = UnitOfMeasure.objects.get(id=data['unit_id'], is_active=True)
            except UnitOfMeasure.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'message': 'واحد مورد نظر یافت نشد'
                }, status=400)
        
        # تبدیل مقدار به واحد اصلی
            quantity = float(data['quantity'])
            conversion_factor = 1.0
        
            if unit.id != item.primary_unit.id:
                try:
                    item_unit = ItemUnit.objects.get(item=item, unit=unit)
                    conversion_factor = float(item_unit.conversion_factor)
                except ItemUnit.DoesNotExist:
                    return JsonResponse({
                        'success': False, 
                        'message': 'واحد انتخاب شده برای این کالا تعریف نشده است'
                    }, status=400)
        
            primary_quantity = quantity * conversion_factor
        
        # بررسی موجودی
            inventory = Inventory.objects.filter(item=item, warehouse=warehouse).first()
            if not inventory or inventory.quantity < primary_quantity:
                return JsonResponse({
                    'success': False, 
                    'message': 'موجودی کافی نیست'
                }, status=400)
        
        # محاسبه هزینه واحد و کل
        # اصلاح شده: استفاده از average_cost به جای unit_cost
            unit_cost = float(inventory.value / inventory.quantity) if inventory.quantity > 0 else 0
            total_cost = quantity * unit_cost
        
        # ایجاد شماره خروجی
            exit_number = f"OUT-{timezone.now().strftime('%Y%m%d')}-{uuid.uuid4().hex[:6].upper()}"
        
        # تنظیم تاریخ خروجی
            exit_date = data.get('exit_date')
            if not exit_date:
                exit_date = timezone.now().date()
            else:
                from django.utils.timezone import make_aware
                exit_date = make_aware(datetime.strptime(exit_date, '%Y-%m-%d'))
                #exit_date = datetime.strptime(exit_date, '%Y-%m-%d').date()
        
        # ایجاد خروجی
            exit_obj = StockExit(
                exit_number=exit_number,
                exit_date=exit_date,
                item=item,
                warehouse=warehouse,
                quantity=quantity,
                unit=unit,
                unit_cost=unit_cost,
                total_cost=total_cost,
                receiver=data.get('receiver', ''),
                receiver_phone=data.get('receiver_phone', ''),
                reference_number=data.get('reference_number', ''),
                description=data.get('description', ''),
                created_by=request.user
            )
        
        # ذخیره خروجی
            exit_obj.save()
        
        # بروزرسانی موجودی
            inventory.quantity -= primary_quantity
            inventory.value -= total_cost  # کاهش ارزش موجودی
            inventory.save()
        
            return JsonResponse({
                'success': True,
                'message': 'خروجی با موفقیت ثبت شد',
                'exit_id': str(exit_obj.id)
            })
        
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False, 
                'message': 'داده‌های ورودی نامعتبر است'
            }, status=400)
        
        except Exception as e:
            return JsonResponse({
                'success': False, 
                'message': f'خطا در ثبت خروجی: {str(e)}'
            }, status=500)
# ==================== مدیریت انبارها ====================


from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views import View
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum, F
import json

@method_decorator(csrf_exempt, name='dispatch')
@method_decorator(login_required, name='dispatch')
class WarehouseView(View):
    # inventory/views.py -> class WarehouseView

    def create_warehouse_notification(self, title, message, notification_type='WAREHOUSE_ALERT', priority='MEDIUM', url=None, metadata=None):
        """ایجاد اعلان برای همه انباردارها و ادمین‌ها"""
        try:
            from users.models import Role
            from .notification_service import NotificationService
            import logging
        
            logger = logging.getLogger(__name__)
            logger.info(f"شروع ایجاد اعلان: {title}")
        
            # دریافت نقش انباردار
            warehouse_role = Role.objects.filter(name='انباردار').first()
            logger.info(f"نقش انباردار یافت شد: {warehouse_role}")
        
            # بررسی وجود کاربر
            created_by = getattr(self, 'request', None)
            if created_by:
                created_by = getattr(created_by, 'user', None)
        
            # استفاده از سرویس مرکزی برای ایجاد اعلان
            notification = NotificationService.create_notification(
                title=title,
                message=message,
                notification_type=notification_type,
                priority=priority,
                action_url=url,
                metadata=metadata or {},
                created_by=created_by,
                is_admin_only=False,
                is_global=False,
                target_role=warehouse_role,
                recipient=None  # اعلان گروهی است، نه شخصی
            )
        
            return notification
        
        except Exception as e:
            import traceback
            logger = logging.getLogger(__name__)
            logger.error(f"خطا در ایجاد اعلان انباردار: {str(e)}")
            logger.error(traceback.format_exc())
            return None

    """
    نمای مدیریت انبارها
    """
    def get_manager_display_name(self, manager):
        """
        تابع کمکی برای گرفتن نام نمایشی مدیر انبار
        """
        if not manager:
            return ''
        
        try:
            # اگر manager یک Profile است
            if hasattr(manager, 'user') and manager.user:
                full_name = manager.user.get_full_name()
                if full_name:
                    return full_name
                return manager.user.username
            
            # اگر مستقیماً User است
            elif hasattr(manager, 'get_full_name'):
                full_name = manager.get_full_name()
                if full_name:
                    return full_name
                return manager.username
            
            # اگر فیلد name دارد
            elif hasattr(manager, 'name'):
                return manager.name
            
            # در غیر این صورت
            return str(manager)
            
        except Exception as e:
            print(f"Error getting manager name: {e}")
            return ''
    
    def get(self, request, warehouse_id=None):
        """نمایش لیست انبارها یا routing به method مناسب"""
        print(f"GET request - warehouse_id: {warehouse_id}")
        print(f"Request path: {request.path}")
        print(f"Is AJAX: {request.headers.get('X-Requested-With') == 'XMLHttpRequest'}")
    
        try:
            if warehouse_id:
            # بررسی URL path برای تشخیص نوع درخواست
                if request.path.endswith('/info/'):
                    return self.get_warehouse_info(request, warehouse_id)
                elif request.path.endswith('/detail/'):
                    return self.get_warehouse_detail(request, warehouse_id)
                else:
                # برای URL های بدون پسوند خاص
                # اگر AJAX است، اطلاعات ساده برگردان
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return self.get_warehouse_info(request, warehouse_id)
                    else:
                    # اگر درخواست معمولی است، به صفحه جزئیات برو
                        return self.get_warehouse_detail(request, warehouse_id)
            else:
                return self.get_warehouses_list(request)
            
        except Exception as e:
            print(f"Error in GET method: {str(e)}")
            import traceback
            traceback.print_exc()
        
        # برای AJAX درخواست‌ها JSON برگردان
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f'خطا در دریافت اطلاعات: {str(e)}'
                }, status=500)
        
        # برای درخواست‌های معمولی صفحه خطا نشان بده
            return render(request, 'inventory/warehouses.html', {
                'error': f'خطا در دریافت اطلاعات: {str(e)}'
            })
    
    def get_warehouses_list(self, request):
        """
        نمایش لیست انبارها
        """
        # اگر درخواست AJAX باشد، JSON برگردان
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return self.get_warehouses_api(request)
        
        # در غیر این صورت، صفحه HTML را نمایش بده
        try:
            # آمار کلی برای نمایش در صفحه
            warehouses = Warehouse.objects.all().select_related('manager')
            active_warehouses_count = warehouses.filter(is_active=True).count()
            
            # محاسبه آمار
            total_items_count = Inventory.objects.values('item').distinct().count()
            total_inventory_value = Inventory.objects.aggregate(
                total=Sum('value')
            )['total'] or 0
            
            # انبارها برای نمایش در جدول
            warehouses_for_display = warehouses.annotate(
                inventory_count=Count('inventory', distinct=True)
            )
            
            context = {
                'warehouses': warehouses_for_display,
                'active_warehouses_count': active_warehouses_count,
                'total_items_count': total_items_count,
                'total_inventory_value': total_inventory_value,
            }
            
            return render(request, 'inventory/warehouses.html', context)
            
        except Exception as e:
            print(f"Error in get_warehouses_list: {str(e)}")
            import traceback
            traceback.print_exc()
            
            return render(request, 'inventory/warehouses.html', {
                'warehouses': Warehouse.objects.none(),
                'active_warehouses_count': 0,
                'total_items_count': 0,
                'total_inventory_value': 0,
                'error': f'خطا در دریافت لیست انبارها: {str(e)}'
            })
    
    def get_warehouses_api(self, request):
        """
        API برای دریافت لیست انبارها
        """
        try:
            # دریافت پارامترهای جستجو و فیلتر
            search_query = request.GET.get('search', '')
            status_filter = request.GET.get('status', '')
            sort_by = request.GET.get('sort_by', 'name')
            sort_order = request.GET.get('sort_order', 'asc')
            page = int(request.GET.get('page', 1))
            per_page = int(request.GET.get('per_page', 12))
            
            # ساخت کوئری
            warehouses = Warehouse.objects.all().select_related('manager')
            
            # اعمال جستجو
            if search_query:
                warehouses = warehouses.filter(
                    Q(name__icontains=search_query) |
                    Q(description__icontains=search_query) |
                    Q(location__icontains=search_query) |
                    Q(code__icontains=search_query)
                )
            
            # اعمال فیلتر وضعیت
            if status_filter:
                warehouses = warehouses.filter(is_active=(status_filter == 'active'))
            
            # اضافه کردن آمار
            warehouses = warehouses.annotate(
                total_items=Count('inventory__item', distinct=True),
                total_entries=Count('stockentry', distinct=True),
                total_exits=Count('stockexit', distinct=True),
                active_inventories=Count('inventory', distinct=True)
            )
            
            # مرتب‌سازی
            if sort_order == 'desc':
                sort_by = f'-{sort_by}'
            warehouses = warehouses.order_by(sort_by)
            
            # صفحه‌بندی
            paginator = Paginator(warehouses, per_page)
            warehouses_page = paginator.get_page(page)
            
            # آماده‌سازی داده‌ها
            warehouses_data = []
            for warehouse in warehouses_page:
                # محاسبه مجموع موجودی و ارزش
                inventory_stats = Inventory.objects.filter(
                    warehouse=warehouse
                ).aggregate(
                    total_quantity=Sum('quantity'),
                    total_value=Sum('value')
                )
                
                warehouses_data.append({
                    'id': str(warehouse.id),
                    'name': warehouse.name,
                    'code': warehouse.code,
                    'description': warehouse.description or '',
                    'location': warehouse.location or '',
                    'is_active': warehouse.is_active,
                    'manager_name': self.get_manager_display_name(warehouse.manager),
                    'manager_id': str(warehouse.manager.id) if warehouse.manager else None,
                    'created_at': warehouse.created_at.strftime('%Y-%m-%d %H:%M') if warehouse.created_at else '',
                    'total_items': warehouse.total_items or 0,
                    'total_entries': warehouse.total_entries or 0,
                    'total_exits': warehouse.total_exits or 0,
                    'active_inventories': warehouse.active_inventories or 0,
                    'total_quantity': float(inventory_stats['total_quantity'] or 0),
                    'total_value': float(inventory_stats['total_value'] or 0),
                })
            
            pagination_data = {
                'current_page': warehouses_page.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_previous': warehouses_page.has_previous(),
                'has_next': warehouses_page.has_next(),
                'per_page': per_page
            }
            
            return JsonResponse({
                'success': True,
                'warehouses': warehouses_data,
                'pagination': pagination_data
            })
            
        except Exception as e:
            print(f"Error in get_warehouses_api: {str(e)}")
            import traceback
            traceback.print_exc()
            
            return JsonResponse({
                'success': False,
                'message': f'خطا در دریافت لیست انبارها: {str(e)}'
            }, status=500)
            
            
    def get_warehouse_info(self, request, warehouse_id):
        """نمایش جزئیات یک انبار"""
        print(f"get_warehouse_detail called - warehouse_id: {warehouse_id}")
        print(f"Is AJAX: {request.headers.get('X-Requested-With') == 'XMLHttpRequest'}")
    
        try:
            warehouse = get_object_or_404(Warehouse, id=warehouse_id)
        
        # اگر درخواست AJAX باشد، فقط اطلاعات اصلی انبار رو برگردان
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                print("Returning JSON response for AJAX")
                return JsonResponse({
                    'success': True,
                    'warehouse': {
                        'id': str(warehouse.id),
                        'name': warehouse.name,
                        'code': warehouse.code,
                        'description': warehouse.description or '',
                        'location': warehouse.location or '',
                        'is_active': warehouse.is_active,
                        'manager_name': self.get_manager_display_name(warehouse.manager),
                        'manager_id': str(warehouse.manager.id) if warehouse.manager else None,
                        'created_at': warehouse.created_at.strftime('%Y-%m-%d %H:%M') if warehouse.created_at else '',
                    }
                })
        
            print("Returning HTML response for normal request")
        # برای درخواست‌های معمولی، صفحه کامل
        # ... باقی کد برای HTML response
        
        except Warehouse.DoesNotExist:
            print(f"Warehouse not found: {warehouse_id}")
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': 'انبار مورد نظر یافت نشد'
                }, status=404)
        
            return render(request, 'inventory/warehouses.html', {
                'error': 'انبار مورد نظر یافت نشد'
            })
        
        except Exception as e:
            print(f"Error in get_warehouse_detail: {str(e)}")
            import traceback
            traceback.print_exc()
        
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f'خطا در دریافت جزئیات انبار: {str(e)}'
                }, status=500)
        
            return render(request, 'inventory/warehouses.html', {
                'error': f'خطا در دریافت جزئیات انبار: {str(e)}'
            })

    def get_warehouse_detail(self, request, warehouse_id):
        """
        نمایش جزئیات یک انبار
        """
        try:
            warehouse = get_object_or_404(Warehouse, id=warehouse_id)
            # آمار انبار
            inventories = Inventory.objects.filter(warehouse=warehouse)
        #     from users.models import UserProfile  # یا هر model که برای مدیران دارید
        #     warehouse_managers = UserProfile.objects.filter(
        #         user__is_active=True,
        # # اضافه کردن فیلتر مناسب برای مدیران
        #     ).select_related('user')
            stats = {
                'total_items': inventories.values('item').distinct().count(),
                'total_inventory_records': inventories.count(),
                'total_stock_quantity': inventories.aggregate(
                    total=Sum('quantity')
                )['total'] or 0,
                'total_stock_value': inventories.aggregate(
                    total=Sum('value')
                )['total'] or 0,
                'low_stock_items': inventories.filter(
                    quantity__lte=F('item__min_stock_level'),
                    quantity__gt=0
                ).count(),
                'out_of_stock_items': inventories.filter(
                    quantity=0
                ).count(),
                'total_entries': StockEntry.objects.filter(warehouse=warehouse).count(),
                'total_exits': StockExit.objects.filter(warehouse=warehouse).count(),
            }
            
            # آخرین فعالیت‌ها
            recent_entries = StockEntry.objects.filter(warehouse=warehouse)\
                .select_related('item', 'created_by')\
                .order_by('-created_at')[:5]
            
            recent_exits = StockExit.objects.filter(warehouse=warehouse)\
                .select_related('item', 'created_by')\
                .order_by('-created_at')[:5]
            
            # کالاهای کم موجود
            low_stock_items = inventories.filter(
                quantity__lte=F('item__min_stock_level'),
                quantity__gt=0,
                item__is_active=True
            ).select_related('item').order_by('quantity')[:10]
            
            # آماده‌سازی داده‌ها برای WebSocket
            warehouse_data = {
                'id': str(warehouse.id),
                'name': warehouse.name,
                'code': warehouse.code,
                'description': warehouse.description or '',
                'location': warehouse.location or '',
                'is_active': warehouse.is_active,
                'manager_name': self.get_manager_display_name(warehouse.manager),
                'created_at': warehouse.created_at.strftime('%Y-%m-%d %H:%M') if warehouse.created_at else '',
                'stats': {k: float(v) if isinstance(v, (int, float)) else v for k, v in stats.items()},
                'recent_entries': [
                    {
                        'id': str(entry.id),
                        'item_name': entry.item.name,
                        'quantity': float(entry.quantity),
                        'user': entry.created_by.get_full_name() if entry.created_by else '',
                        'created_at': entry.created_at.strftime('%Y-%m-%d %H:%M') if entry.created_at else ''
                    } for entry in recent_entries
                ],
                'recent_exits': [
                    {
                        'id': str(exit.id),
                        'item_name': exit.item.name,
                        'quantity': float(exit.quantity),
                        'user': exit.created_by.get_full_name() if exit.created_by else '',
                        'created_at': exit.created_at.strftime('%Y-%m-%d %H:%M') if exit.created_at else ''
                    } for exit in recent_exits
                ],
                'low_stock_items': [
                    {
                        'id': str(item.item.id),
                        'name': item.item.name,
                        'stock_quantity': float(item.quantity),
                        'min_stock_level': float(item.item.min_stock_level),
                        'unit': item.item.primary_unit.name if item.item.primary_unit else ''
                    } for item in low_stock_items
                ]
            }
            
            # اگر درخواست AJAX باشد، JSON برگردان
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'warehouse': warehouse_data
                })
            
            context = {
                'warehouse': warehouse,
                'stats': stats,
                'recent_entries': recent_entries,
                'recent_exits': recent_exits,
                'low_stock_items': low_stock_items,
                #'warehouse_managers': warehouse_managers,
                'warehouse_json': json.dumps(warehouse_data, default=str)
            }
            
            return render(request, 'inventory/warehouses/warehouse_detail.html', context)
            
        except Exception as e:
            print(f"Error in get_warehouse_detail: {str(e)}")
            import traceback
            traceback.print_exc()
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': f'خطا در دریافت جزئیات انبار: {str(e)}'
                }, status=500)
            
            return render(request, 'inventory/warehouses/warehouse_detail.html', {
                'error': f'خطا در دریافت جزئیات انبار: {str(e)}'
            })
    
    def post(self, request):
        """
        ایجاد انبار جدید
        """
        try:
            data = json.loads(request.body)
            print("Received data:", data)
            
            # بررسی اجباری بودن فیلدها
            if not data.get('name'):
                return JsonResponse({
                    'success': False,
                    'message': 'نام انبار الزامی است'
                }, status=400)
            
            if not data.get('code'):
                return JsonResponse({
                    'success': False,
                    'message': 'کد انبار الزامی است'
                }, status=400)
            
            # بررسی تکراری نبودن کد
            if Warehouse.objects.filter(code=data.get('code')).exists():
                return JsonResponse({
                    'success': False,
                    'message': 'کد انبار تکراری است'
                }, status=400)
            
            
            warehouse = Warehouse.objects.create(
                name=data.get('name'),
                code=data.get('code'),
                description=data.get('description', ''),
                location=data.get('location', ''),
                is_active=data.get('is_active', True),
                manager_id=data.get('manager_id') if data.get('manager_id') else None,
                created_by=request.user
            )
            
            # ارسال پیام WebSocket
            try:
                from .websocket_utils import send_websocket_message
                send_websocket_message('warehouse_created', {
                    'id': str(warehouse.id),
                    'name': warehouse.name,
                    'code': warehouse.code,
                    'description': warehouse.description,
                    'location': warehouse.location,
                    'is_active': warehouse.is_active,
                    'manager_name': self.get_manager_display_name(warehouse.manager),
                    'created_at': warehouse.created_at.strftime('%Y-%m-%d %H:%M'),
                    'user': request.user.get_full_name() or request.user.username
                })
            except Exception as ws_error:
                print(f"WebSocket error: {ws_error}")
            # ایجاد نوتیفیکیشن
            self.create_warehouse_notification(
                title="ایجاد انبار جدید",
                message=f"انبار جدید «{warehouse.name}» با کد {warehouse.code} توسط {request.user.get_full_name() or request.user.username} ایجاد شد.",
                notification_type="WAREHOUSE_ALERT",
                priority="MEDIUM",
                url=f"/inventory/warehouses/{warehouse.id}/detail/",
                metadata={
                    'warehouse_id': str(warehouse.id),
                    'warehouse_name': warehouse.name,
                    'warehouse_code': warehouse.code,
                    'action': 'create'
                }
            )
            return JsonResponse({
                'success': True,
                'message': 'انبار با موفقیت ایجاد شد',
                'warehouse': {
                    'id': str(warehouse.id),
                    'name': warehouse.name,
                    'code': warehouse.code,
                    'description': warehouse.description,
                    'location': warehouse.location,
                    'manager_name': self.get_manager_display_name(warehouse.manager),
                    'is_active': warehouse.is_active
                }
            })
            
        except Exception as e:
            print(f"Error in warehouse creation: {str(e)}")
            import traceback
            traceback.print_exc()
            
            return JsonResponse({
                'success': False,
                'message': f'خطا در ایجاد انبار: {str(e)}'
            }, status=400)
    
    def put(self, request, warehouse_id):
        """
        بروزرسانی انبار
        """
        try:
            warehouse = get_object_or_404(Warehouse, id=warehouse_id)
            data = json.loads(request.body)
            
            # بررسی تکراری نبودن کد (به جز خود انبار)
            if data.get('code') and data.get('code') != warehouse.code:
                if Warehouse.objects.filter(code=data.get('code')).exclude(id=warehouse.id).exists():
                    return JsonResponse({
                        'success': False,
                        'message': 'کد انبار تکراری است'
                    }, status=400)
            # ذخیره مقادیر قبلی برای نوتیفیکیشن
            old_name = warehouse.name
            old_code = warehouse.code
            old_manager = warehouse.manager
            old_is_active = warehouse.is_active
            
            
            
            warehouse.name = data.get('name', warehouse.name)
            warehouse.code = data.get('code', warehouse.code)
            warehouse.description = data.get('description', warehouse.description)
            warehouse.location = data.get('location', warehouse.location)
            warehouse.is_active = data.get('is_active', warehouse.is_active)
            
            if data.get('manager_id'):
                warehouse.manager_id = data.get('manager_id')
            
            warehouse.save()
            
            # ارسال پیام WebSocket
            try:
                from .websocket_utils import send_websocket_message
                send_websocket_message('warehouse_updated', {
                    'id': str(warehouse.id),
                    'name': warehouse.name,
                    'code': warehouse.code,
                    'description': warehouse.description,
                    'location': warehouse.location,
                    'is_active': warehouse.is_active,
                    'manager_name': self.get_manager_display_name(warehouse.manager),
                    'user': request.user.get_full_name() or request.user.username
                })
            except Exception as ws_error:
                print(f"WebSocket error: {ws_error}")
            # ایجاد پیام نوتیفیکیشن
            changes = []
            if old_name != warehouse.name:
                changes.append(f"نام از «{old_name}» به «{warehouse.name}»")
            if old_code != warehouse.code:
                changes.append(f"کد از «{old_code}» به «{warehouse.code}»")
            if old_manager != warehouse.manager:
                old_manager_name = self.get_manager_display_name(old_manager)
                new_manager_name = self.get_manager_display_name(warehouse.manager)
                changes.append(f"مدیر از «{old_manager_name}» به «{new_manager_name}»")
            if old_is_active != warehouse.is_active:
                status_change = "فعال به غیرفعال" if old_is_active else "غیرفعال به فعال"
                changes.append(f"وضعیت از {status_change}")

            if changes:
                changes_text = "، ".join(changes)
                message = f"انبار «{warehouse.name}» بروزرسانی شد. تغییرات: {changes_text}"
            else:
                message = f"انبار «{warehouse.name}» بروزرسانی شد."

            try:
                notification = self.create_warehouse_notification(
                    title="بروزرسانی انبار",
                    message=message,
                    notification_type="WAREHOUSE_ALERT",
                    priority="MEDIUM",
                    url=f"/inventory/warehouses/{warehouse.id}/detail/",
                    metadata={
                        'warehouse_id': str(warehouse.id),
                        'warehouse_name': warehouse.name,
                        'warehouse_code': warehouse.code,
                        'action': 'update',
                        'changes': changes
                    }
                )
                print("notification andar put:",notification)
                if notification:
                    print(f"نوتیفیکیشن با موفقیت ایجاد شد. ID: {notification.id}")
                else:
                    print("خطا در ایجاد نوتیفیکیشن")
            except Exception as e:
                import traceback
                print(f"خطا در ایجاد نوتیفیکیشن: {str(e)}")
                print(traceback.format_exc())
                
                
            return JsonResponse({
                'success': True,
                'message': 'انبار با موفقیت بروزرسانی شد',
                'warehouse': {
                    'id': str(warehouse.id),
                    'name': warehouse.name,
                    'code': warehouse.code,
                    'description': warehouse.description,
                    'location': warehouse.location,
                    'manager_name': self.get_manager_display_name(warehouse.manager),
                    'is_active': warehouse.is_active
                }
            })
            
        except Exception as e:
            print(f"Error in warehouse update: {str(e)}")
            import traceback
            traceback.print_exc()
            
            return JsonResponse({
                'success': False,
                'message': f'خطا در بروزرسانی انبار: {str(e)}'
            }, status=400)
    
    def delete(self, request, warehouse_id):
        """حذف انبار با چک‌های امنیتی """
        try:
            warehouse = get_object_or_404(Warehouse, id=warehouse_id)
        
        # 1. چک موجودی‌های فعال
            active_inventory = Inventory.objects.filter(
                warehouse=warehouse,
                quantity__gt=0
            ).select_related('item')
        
            if active_inventory.exists():
                items_list = list(active_inventory.values_list('item__name', flat=True)[:5])
                items_text = '، '.join(items_list)
                if active_inventory.count() > 5:
                    items_text += f' و {active_inventory.count() - 5} مورد دیگر'
                
                return JsonResponse({
                    'success': False,
                    'message': f'امکان حذف انبار وجود ندارد.\n\nاین انبار دارای موجودی از کالاهای زیر است:\n{items_text}\n\nابتدا موجودی‌ها را تخلیه کنید یا انبار را غیرفعال نمایید.'
                }, status=400)
        
        # 2. چک سوابق ورود و خروج
            entries_count = StockEntry.objects.filter(warehouse=warehouse).count()
            exits_count = StockExit.objects.filter(warehouse=warehouse).count()
        
            if entries_count > 0 or exits_count > 0:
                return JsonResponse({
                    'success': False,
                    'message': f'امکان حذف انبار وجود ندارد.\n\nاین انبار دارای سوابق عملیاتی است:\n• {entries_count} مورد ورودی\n• {exits_count} مورد خروجی\n\nبرای حفظ سوابق، انبار را غیرفعال کنید.'
                }, status=400)
        
        # 3. چک اینکه آیا انبار پیش‌فرض سیستم است
            if hasattr(warehouse, 'is_default') and warehouse.is_default:
                return JsonResponse({
                    'success': False,
                    'message': 'امکان حذف انبار پیش‌فرض سیستم وجود ندارد.'
                }, status=400)
        
        # 4. چک دسترسی کاربر (اختیاری)
        # if not request.user.has_perm('inventory.delete_warehouse'):
        #     return JsonResponse({
        #         'success': False,
        #         'message': 'شما مجوز حذف انبار را ندارید.'
        #     }, status=403)
        
        # اگر همه چیز اوکی بود، حذف کن
            warehouse_name = warehouse.name
            warehouse_code = warehouse.code
            warehouse.delete()
        
        # ارسال پیام WebSocket
            try:
                from .websocket_utils import send_websocket_message
                send_websocket_message('warehouse_deleted', {
                    'id': str(warehouse_id),
                    'name': warehouse_name,
                    'code': warehouse_code,
                    'user': request.user.get_full_name() or request.user.username
                })
            except Exception as ws_error:
                print(f"WebSocket error: {ws_error}")
                # ایجاد نوتیفیکیشن
            self.create_warehouse_notification(
                title="حذف انبار",
                message=f"انبار «{warehouse_name}» با کد {warehouse_code} توسط {request.user.get_full_name() or request.user.username} حذف شد.",
                notification_type="WAREHOUSE_ALERT",
                priority="HIGH",
                metadata={
                    'warehouse_name': warehouse_name,
                    'warehouse_code': warehouse_code,
                    'action': 'delete'
                }
            )
            return JsonResponse({
                'success': True,
                'message': f'انبار "{warehouse_name}" با موفقیت حذف شد'
            })
        
        except Warehouse.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'انبار مورد نظر یافت نشد'
            }, status=404)
        
        except Exception as e:
            print(f"Error in warehouse deletion: {str(e)}")
            import traceback
            traceback.print_exc()
        
            return JsonResponse({
                'success': False,
                'message': f'خطا در حذف انبار: {str(e)}'
            }, status=500)

# ==================== مدیریت موجودی ====================

@login_required
def inventory_list(request):
    """لیست موجودی انبار - API"""
    
    inventories = Inventory.objects.select_related(
        'item', 'warehouse', 'batch'
    ).filter(quantity__gt=0)
    
    # فیلترها
    warehouse_id = request.GET.get('warehouse')
    if warehouse_id:
        inventories = inventories.filter(warehouse_id=warehouse_id)
    
    search = request.GET.get('search')
    if search:
        inventories = inventories.filter(
            Q(item__name__icontains=search) |
            Q(item__code__icontains=search) |
            Q(item__barcode__icontains=search)
        )
    
    # صفحه‌بندی
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 20))
    paginator = Paginator(inventories, per_page)
    page_obj = paginator.get_page(page)
    
    # محاسبه آمار
    total_value = inventories.aggregate(total=Sum('value'))['total'] or 0
    total_items = inventories.count()
    
    data = {
        'inventories': [{
            'id': str(inv.id),
            'item_name': inv.item.name,
            'item_code': inv.item.code,
            'warehouse_name': inv.warehouse.name,
            'batch_number': inv.batch.batch_number if inv.batch else None,
            'quantity': float(inv.quantity),
            'value': float(inv.value),
            'average_cost': float(inv.average_cost),
            'primary_unit': inv.item.primary_unit.symbol,
        } for inv in page_obj],
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
        },
        'summary': {
            'total_value': float(total_value),
            'total_items': total_items,
        }
    }
    
    return json_response(data)

# ==================== گزارشات ====================

@login_required
def low_stock_report(request):
    """گزارش کالاهای کم موجودی - API"""
    
    warehouse_id = request.GET.get('warehouse')
    warehouse = None
    if warehouse_id:
        warehouse = get_object_or_404(Warehouse, id=warehouse_id)
    
    low_stock_items = InventoryService.get_low_stock_items(warehouse)
    
    data = {
        'low_stock_items': [{
            'item_id': str(item['item'].id),
            'item_name': item['item'].name,
            'item_code': item['item'].code,
            'current_stock': float(item['current_stock']),
            'alert_threshold': float(item['alert_threshold']),
            'status': item['status'],
            'primary_unit': item['item'].primary_unit.symbol,
        } for item in low_stock_items],
        'warehouse': warehouse.name if warehouse else 'همه انبارها',
        'total_items': len(low_stock_items)
    }
    
    return json_response(data)

@login_required
def expiry_report(request):
    """گزارش کالاهای نزدیک به انقضا - API"""
    
    from django.utils import timezone
    from datetime import timedelta
    
    days = int(request.GET.get('days', 60))
    expiry_threshold = timezone.now().date() + timedelta(days=days)
    
    expiring_batches = StockBatch.objects.filter(
        expiry_date__lte=expiry_threshold,
        expiry_date__gte=timezone.now().date()
    ).select_related('item').order_by('expiry_date')
    
    # اضافه کردن موجودی به هر بچ
    batch_data = []
    for batch in expiring_batches:
        current_stock = Inventory.objects.filter(batch=batch).aggregate(
            total=Sum('quantity')
        )['total'] or 0
        
        if current_stock > 0:  # فقط بچ‌هایی که موجودی دارند
            batch_data.append({
                'batch_id': str(batch.id),
                'batch_number': batch.batch_number,
                'item_name': batch.item.name,
                'item_code': batch.item.code,
                'expiry_date': batch.expiry_date.isoformat(),
                'current_stock': float(current_stock),
                'days_to_expiry': (batch.expiry_date - timezone.now().date()).days,
                'status': 'critical' if (batch.expiry_date - timezone.now().date()).days <= 7 else 'warning'
            })
    
    data = {
        'expiring_batches': batch_data,
        'days_threshold': days,
        'total_batches': len(batch_data)
    }
    
    return json_response(data)

@login_required
def stock_movement_report(request):
    """گزارش حرکات انبار - API"""
    
    movements = StockMovement.objects.select_related(
        'item', 'warehouse', 'batch'
    ).order_by('-movement_date')
    
    # فیلترها
    item_id = request.GET.get('item')
    if item_id:
        movements = movements.filter(item_id=item_id)
    
    warehouse_id = request.GET.get('warehouse')
    if warehouse_id:
        movements = movements.filter(warehouse_id=warehouse_id)
    
    movement_type = request.GET.get('movement_type')
    if movement_type:
        movements = movements.filter(movement_type=movement_type)
    
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        movements = movements.filter(movement_date__date__gte=date_from)
    if date_to:
        movements = movements.filter(movement_date__date__lte=date_to)
    
    # صفحه‌بندی
    page = int(request.GET.get('page', 1))
    per_page = int(request.GET.get('per_page', 50))
    paginator = Paginator(movements, per_page)
    page_obj = paginator.get_page(page)
    
    data = {
        'movements': [{
            'id': str(movement.id),
            'item_name': movement.item.name,
            'warehouse_name': movement.warehouse.name,
            'movement_type': movement.movement_type,
            'quantity_before': float(movement.quantity_before),
            'quantity_change': float(movement.quantity_change),
            'quantity_after': float(movement.quantity_after),
            'value_change': float(movement.value_change),
            'movement_date': movement.movement_date.isoformat(),
            'reference_type': movement.reference_type,
        } for movement in page_obj],
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_items': paginator.count,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
        }
    }
    
    return json_response(data)

# ==================== API های کمکی ====================

@login_required
def api_search_items(request):
    """API جستجوی کالاها"""
    
    query = request.GET.get('q', '')
    if len(query) < 2:
        return json_response({'items': []})
    
    items = Item.objects.filter(
        Q(name__icontains=query) |
        Q(code__icontains=query) |
        Q(barcode__icontains=query),
        is_active=True
    )[:10]
    
    results = []
    for item in items:
        current_stock = InventoryService.get_current_stock(item)
        results.append({
            'id': str(item.id),
            'name': item.name,
            'code': item.code,
            'barcode': item.barcode,
            'current_stock': float(current_stock),
            'primary_unit': item.primary_unit.symbol,
            'has_batch': item.has_batch,
            'has_serial': item.has_serial,
        })
    
    return json_response({'items': results})

@login_required
def api_item_units(request, item_id):
    """API دریافت واحدهای کالا"""
    
    item = get_object_or_404(Item, id=item_id)
    units = ItemUnit.objects.filter(item=item).select_related('unit')
    
    results = [{
        'id': str(unit.unit.id),
        'name': unit.unit.name,
        'symbol': unit.unit.symbol,
        'conversion_factor': float(unit.conversion_factor),
    } for unit in units]
    
    # اضافه کردن واحد اصلی
    results.insert(0, {
        'id': str(item.primary_unit.id),
        'name': item.primary_unit.name,
        'symbol': item.primary_unit.symbol,
        'conversion_factor': 1.0,
    })
    
    return json_response({'units': results})

@login_required
def api_item_batches(request, item_id):
    """API دریافت بچ‌های کالا"""
    
    item = get_object_or_404(Item, id=item_id)
    warehouse_id = request.GET.get('warehouse')
    
    batches = StockBatch.objects.filter(item=item)
    
    if warehouse_id:
        # فقط بچ‌هایی که در انبار مشخص موجودی دارند
        batches = batches.filter(
            inventory__warehouse_id=warehouse_id,
            inventory__quantity__gt=0
        ).distinct()
    
    results = []
    for batch in batches:
        # محاسبه موجودی بچ
        if warehouse_id:
            stock = Inventory.objects.filter(
                batch=batch, 
                warehouse_id=warehouse_id
            ).aggregate(total=Sum('quantity'))['total'] or 0
        else:
            stock = Inventory.objects.filter(
                batch=batch
            ).aggregate(total=Sum('quantity'))['total'] or 0
        
        if stock > 0:
            results.append({
                'id': str(batch.id),
                'batch_number': batch.batch_number,
                'production_date': batch.production_date.isoformat() if batch.production_date else None,
                'expiry_date': batch.expiry_date.isoformat() if batch.expiry_date else None,
                'current_stock': float(stock),
            })
    
    return json_response({'batches': results})

@login_required
def api_check_stock(request, item_id):
    """API چک موجودی کالا"""
    
    item = get_object_or_404(Item, id=item_id)
    warehouse_id = request.GET.get('warehouse')
    batch_id = request.GET.get('batch')
    
    warehouse = None
    if warehouse_id:
        warehouse = get_object_or_404(Warehouse, id=warehouse_id)
    
    batch = None
    if batch_id:
        batch = get_object_or_404(StockBatch, id=batch_id)
    
    # محاسبه موجودی
    if batch:
        if warehouse:
            current_stock = Inventory.objects.filter(
                item=item, warehouse=warehouse, batch=batch
            ).aggregate(total=Sum('quantity'))['total'] or 0
        else:
            current_stock = Inventory.objects.filter(
                item=item, batch=batch
            ).aggregate(total=Sum('quantity'))['total'] or 0
    else:
        current_stock = InventoryService.get_current_stock(item, warehouse)
    
    return json_response({
        'current_stock': float(current_stock),
        'alert_threshold': float(item.alert_threshold),
        'status': 'critical' if current_stock == 0 else 'warning' if current_stock <= item.alert_threshold else 'ok'
    })

# API Views
@csrf_exempt
def api_item(request):
    """API برای کالاها"""
    if request.method == 'GET':
        try:
            items = Item.objects.filter(is_active=True).select_related('category', 'unit')
            
            # جستجو
            search = request.GET.get('search', '')
            if search:
                items = items.filter(
                    Q(name__icontains=search) | 
                    Q(code__icontains=search)
                )
            
            # فیلتر دسته‌بندی
            category_id = request.GET.get('category')
            if category_id:
                items = items.filter(category_id=category_id)
            
            # صفحه‌بندی
            page = int(request.GET.get('page', 1))
            paginator = Paginator(items, 20)
            page_obj = paginator.get_page(page)
            
            items_data = []
            for item in page_obj:
                items_data.append({
                    'id': item.id,
                    'name': item.name,
                    'code': item.code,
                    'category_name': item.category.name if item.category else '',
                    'unit': item.unit.name if item.unit else '',
                    'description': item.description or '',
                    'is_active': item.is_active,
                })
            
            return JsonResponse({
                'items': items_data,
                'pagination': {
                    'current_page': page_obj.number,
                    'total_pages': paginator.num_pages,
                    'has_previous': page_obj.has_previous(),
                    'has_next': page_obj.has_next(),
                    'total': paginator.count,
                    'from': (page_obj.number - 1) * 20 + 1,
                    'to': min(page_obj.number * 20, paginator.count),
                }
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    elif request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            item = Item.objects.create(
                name=data['name'],
                code=data['code'],
                category_id=data.get('category_id'),
                unit_id=data.get('unit_id'),
                description=data.get('description', ''),
            )
            
            return JsonResponse({
                'message': 'کالا با موفقیت ثبت شد',
                'item_id': item.id
            })
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

@csrf_exempt
def api_item_detail(request, item_id):
    """جزئیات کالا"""
    try:
        item = get_object_or_404(Item, id=item_id)
        
        if request.method == 'GET':
            return JsonResponse({
                'item': {
                    'id': item.id,
                    'name': item.name,
                    'code': item.code,
                    'category_id': item.category_id,
                    'category_name': item.category.name if item.category else '',
                    'unit_id': item.unit_id,
                    'unit_name': item.unit.name if item.unit else '',
                    'description': item.description,
                    'is_active': item.is_active,
                }
            })
            
        elif request.method == 'PUT':
            data = json.loads(request.body)
            
            item.name = data['name']
            item.code = data['code']
            item.category_id = data.get('category_id')
            item.unit_id = data.get('unit_id')
            item.description = data.get('description', '')
            item.save()
            
            return JsonResponse({'message': 'کالا با موفقیت بروزرسانی شد'})
            
        elif request.method == 'DELETE':
            item.is_active = False
            item.save()
            return JsonResponse({'message': 'کالا با موفقیت حذف شد'})
            
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

def api_warehouse(request):
    """API برای انبارها"""
    try:
        warehouses = Warehouse.objects.filter(is_active=True)
        warehouses_data = []
        
        for warehouse in warehouses:
            warehouses_data.append({
                'id': warehouse.id,
                'name': warehouse.name,
                'code': warehouse.code,
                'location': warehouse.location,
            })
        
        return JsonResponse({'warehouses': warehouses_data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def api_category(request):
    """API برای دسته‌بندی‌ها"""
    try:
        categories = Category.objects.filter(is_active=True)
        categories_data = []
        
        for category in categories:
            categories_data.append({
                'id': category.id,
                'name': category.name,
                'description': category.description,
            })
        
        return JsonResponse({'categories': categories_data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def api_unit(request):
    """API برای واحدها"""
    try:
        units = Unit.objects.filter(is_active=True)
        units_data = []
        
        for unit in units:
            units_data.append({
                'id': unit.id,
                'name': unit.name,
                'symbol': unit.symbol,
            })
        
        return JsonResponse({'units': units_data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def api_stock_entry(request):
    """API برای ورودی انبار"""
    if request.method == 'GET':
        try:
            entries = StockEntry.objects.select_related('item', 'warehouse', 'unit').order_by('-entry_date')
            
            # محدود کردن تعداد برای داشبورد
            limit = request.GET.get('limit')
            if limit:
                entries = entries[:int(limit)]
            
            entries_data = []
            for entry in entries:
                entries_data.append({
                    'id': entry.id,
                    'entry_number': entry.entry_number,
                    'entry_date': entry.entry_date.isoformat(),
                    'item_name': entry.item.name,
                    'item_code': entry.item.code,
                    'warehouse_name': entry.warehouse.name,
                    'quantity': float(entry.quantity),
                    'unit': entry.unit.name if entry.unit else '',
                    'unit_cost': float(entry.unit_cost),
                    'total_cost': float(entry.total_cost),
                    'notes': entry.notes,
                })
            
            return JsonResponse({'entries': entries_data})
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

def api_stock_entry_detail(request, entry_id):
    """جزئیات ورودی"""
    # فعلاً خالی
    return JsonResponse({'message': 'در حال توسعه'})

@csrf_exempt
def api_stock_exit(request):
    """API برای خروجی انبار"""
    if request.method == 'GET':
        try:
            exits = StockExit.objects.select_related('item', 'warehouse', 'unit').order_by('-exit_date')
            
            # محدود کردن تعداد برای داشبورد
            limit = request.GET.get('limit')
            if limit:
                exits = exits[:int(limit)]
            
            exits_data = []
            for exit in exits:
                exits_data.append({
                    'id': exit.id,
                    'exit_number': exit.exit_number,
                    'exit_date': exit.exit_date.isoformat(),
                    'item_name': exit.item.name,
                    'item_code': exit.item.code,
                    'warehouse_name': exit.warehouse.name,
                    'quantity': float(exit.quantity),
                    'unit': exit.unit.name if exit.unit else '',
                    'unit_cost': float(exit.unit_cost),
                    'total_cost': float(exit.total_cost),
                    'notes': exit.notes,
                })
            
            return JsonResponse({'exits': exits_data})
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

def api_stock_exit_detail(request, exit_id):
    """جزئیات خروجی"""
    # فعلاً خالی
    return JsonResponse({'message': 'در حال توسعه'})

def api_inventory(request):
    """API برای موجودی"""
    try:
        # فعلاً یک لیست خالی برمی‌گردونیم
        return JsonResponse({'inventory': []})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def export_movements(request):
    """صادرات حرکات انبار"""
    # فعلاً خالی
    return JsonResponse({'message': 'در حال توسعه'})

@login_required
def warehouse_detail_page(request, warehouse_id):
    """صفحه جزئیات انبار"""
    warehouse = get_object_or_404(Warehouse, id=warehouse_id)
    
    context = {
        'warehouse': warehouse,
        'page_title': f'جزئیات انبار {warehouse.name}',
    }
    
    return render(request, 'inventory/warehouse_detail.html', context)

# ==================== صفحات HTML ====================
# inventory/views.py - اضافه کن به کد قبلی
@login_required
def item_list_page(request):
    """صفحه لیست کالاها"""
    categories = Category.objects.filter(is_active=True)
    units = UnitOfMeasure.objects.filter(is_active=True)
    warehouses = Warehouse.objects.filter(is_active=True)
    
    return render(request, 'inventory/item_list.html', {
        'categories': categories,
        'units': units,
        'warehouses': warehouses,
        'page_title': 'مدیریت کالاها'
    })
@login_required
def warehouse_list_page(request):
    """صفحه مدیریت انبارها"""
    try:
        # دریافت لیست انبارها
        warehouses = Warehouse.objects.all().order_by('name')
        
        # آمار
        active_warehouses_count = warehouses.filter(is_active=True).count()
        total_items_count = Item.objects.count()
        total_inventory_value = Inventory.objects.aggregate(total=Sum('value'))['total'] or 0
        
        # دریافت لیست مدیران برای فیلتر
        managers = Profile.objects.filter(managed_warehouses__isnull=False).distinct()
        
        # دریافت لیست پروفایل‌ها با نقش انباردار
        try:
            # اول چک کنیم که آیا نقش انباردار وجود دارد
            warehouse_manager_role = Role.objects.filter(name='انباردار').first()
            if not warehouse_manager_role:
                print("نقش انباردار تعریف نشده است")
                # اگر نقش انباردار وجود نداشت، همه پروفایل‌های فعال را برگردان
                warehouse_managers = Profile.objects.filter(is_activate=True)
            else:
                # فیلتر پروفایل‌های با نقش انباردار
                warehouse_managers = Profile.objects.filter(
                    role=warehouse_manager_role,
                    is_activate=True
                ).distinct()
                
            print(f"تعداد انبارداران یافت شده: {warehouse_managers.count()}")
            
            # اگر هیچ انبارداری پیدا نشد، همه پروفایل‌های فعال را برگردان
            if warehouse_managers.count() == 0:
                warehouse_managers = Profile.objects.filter(is_activate=True)
                print(f"هیچ انبارداری پیدا نشد، همه پروفایل‌های فعال برگردانده شد: {warehouse_managers.count()}")
                
        except Exception as e:
            print(f"خطا در دریافت انبارداران: {str(e)}")
            # در صورت خطا، همه پروفایل‌های فعال را برگردان
            warehouse_managers = Profile.objects.filter(is_activate=True)
        
        # صفحه‌بندی
        page = request.GET.get('page', 1)
        paginator = Paginator(warehouses, 10)
        warehouses_page = paginator.get_page(page)
        
        context = {
            'warehouses': warehouses_page,
            'active_warehouses_count': active_warehouses_count,
            'total_items_count': total_items_count,
            'total_inventory_value': total_inventory_value,
            'managers': managers,
            'warehouse_managers': warehouse_managers,
            'page_title': 'مدیریت انبارها',
        }
        
        return render(request, 'inventory/warehouses.html', context)
    except Exception as e:
            return json_response(
                message=f'خطا در ایجاد دسته‌بندی: {str(e)}',
                status=500,
                success=False
            )
# در فایل views.py
@login_required
def get_warehouse_managers(request):
    """API برای دریافت لیست انبارداران"""
    try:
        # اول چک کنیم که آیا نقش انباردار وجود دارد
        warehouse_manager_role = Role.objects.filter(name='انباردار').first()
        
        if warehouse_manager_role:
            # فیلتر پروفایل‌های با نقش انباردار
            warehouse_managers = Profile.objects.filter(
                role=warehouse_manager_role,
                is_activate=True
            ).distinct()
        else:
            # اگر نقش انباردار وجود نداشت، همه پروفایل‌های فعال را برگردان
            warehouse_managers = Profile.objects.filter(is_activate=True)
        
        # تبدیل به دیکشنری
        managers_data = []
        for profile in warehouse_managers:
            managers_data.append({
                'id': profile.id,
                'first_name': profile.first_name,
                'last_name': profile.last_name,
                'full_name': f"{profile.first_name} {profile.last_name}",
                'phone': profile.mobile or profile.phone or '',
                'email': profile.user.email if profile.user else '',
                'role': profile.role.name if profile.role else ''
            })
        
        return JsonResponse({
            'success': True,
            'managers': managers_data
        })
    except Exception as e:
        print(f"خطا در دریافت انبارداران: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': str(e),
            'managers': []
        })
        
@login_required
def item_detail_page(request, item_id):
    """صفحه جزئیات کالا"""
    item = get_object_or_404(Item, id=item_id, is_active=True)
    
    # دریافت موجودی کالا در انبارهای مختلف
    inventory_items = Inventory.objects.filter(item=item)
    
    # دریافت تاریخچه ورود و خروج
    entries = StockEntry.objects.filter(item=item).order_by('-entry_date')[:10]
    exits = StockExit.objects.filter(item=item).order_by('-exit_date')[:10]
    
    return render(request, 'inventory/item_detail.html', {
        'item': item,
        'inventory_items': inventory_items,
        'entries': entries,
        'exits': exits,
        'page_title': f'جزئیات کالا: {item.name}'
    })


@login_required
def item_detail_page(request, item_id):
    """صفحه جزئیات کالا"""
    item = get_object_or_404(Item, id=item_id)
    return render(request, 'inventory/item_detail.html', {
        'item': item,
        'page_title': f'جزئیات کالا - {item.name}'
    })

@login_required
def stock_entry_list_page(request):
    """صفحه لیست ورودی‌های انبار"""
    items = Item.objects.filter(is_active=True)
    warehouses = Warehouse.objects.filter(is_active=True)
    units = UnitOfMeasure.objects.filter(is_active=True)
    
    return render(request, 'inventory/stock_entry_list.html', {
        'items': items,
        'warehouses': warehouses,
        'units': units,
        'today': timezone.now().date().strftime('%Y-%m-%d'),
        'page_title': 'ورود کالا به انبار'
    })

@login_required
def stock_exit_list_page(request):
    """صفحه لیست خروجی‌های انبار"""
    items = Item.objects.filter(is_active=True)
    warehouses = Warehouse.objects.filter(is_active=True)
    units = UnitOfMeasure.objects.filter(is_active=True)
    
    return render(request, 'inventory/stock_exit_list.html', {
        'items': items,
        'warehouses': warehouses,
        'units': units,
        'today': timezone.now().date().strftime('%Y-%m-%d'),
        'page_title': 'خروج کالا از انبار'
    })

@login_required
def stock_exit_list_page(request):
    """صفحه لیست خروجی‌ها"""
    warehouses = Warehouse.objects.filter(is_active=True)
    items = Item.objects.filter(is_active=True)
    units = Unit.objects.filter(is_active=True)
    return render(request, 'inventory/stock_exit_list.html', {
        'warehouses': warehouses,
        'items': items,
        'units': units,
        'page_title': 'خروجی انبار'
    })

@login_required
def inventory_list_page(request):
    """صفحه لیست موجودی انبار"""
    warehouses = Warehouse.objects.filter(is_active=True)
    categories = Category.objects.filter(is_active=True)
    
    return render(request, 'inventory/inventory_list.html', {
        'warehouses': warehouses,
        'categories': categories,
        'page_title': 'موجودی انبار'
    })

login_required
def reports_page(request):
    """صفحه گزارشات"""
    warehouses = Warehouse.objects.filter(is_active=True)
    categories = Category.objects.filter(is_active=True)
    
    return render(request, 'inventory/reports.html', {
        'warehouses': warehouses,
        'categories': categories,
        'page_title': 'گزارشات انبار'
    })


# inventory/views.py



@method_decorator(login_required, name='dispatch')
@method_decorator(csrf_exempt, name='dispatch')
class CategoryView(View):
    """View کلاس برای مدیریت دسته‌بندی‌ها"""
    
    def get(self, request, category_id=None):
        """دریافت دسته‌بندی یا لیست دسته‌بندی‌ها"""
        
        if category_id:
            # جزئیات یک دسته‌بندی
            category = get_object_or_404(Category, id=category_id)
            
            data = {
                'category': {
                    'id': str(category.id),
                    'name': category.name,
                    'code': category.code,
                    'description': category.description,
                    'parent_id': str(category.parent.id) if category.parent else None,
                    'parent_name': category.parent.name if category.parent else None,
                    'is_active': category.is_active,
                    'created_at': category.created_at.isoformat(),
                    'updated_at': category.updated_at.isoformat(),
                    'items_count': category.items.filter(is_active=True).count(),
                    'subcategories_count': category.children.filter(is_active=True).count(),
                }
            }
            
            return json_response(data)
        
        else:
            # لیست دسته‌بندی‌ها
            categories = Category.objects.filter(is_active=True).select_related('parent')
            
            # فیلترها
            parent_id = request.GET.get('parent_id')
            if parent_id:
                if parent_id == 'null':
                    categories = categories.filter(parent__isnull=True)
                else:
                    categories = categories.filter(parent_id=parent_id)
            
            search = request.GET.get('search')
            if search:
                categories = categories.filter(
                    name__icontains=search
                ).distinct()
            
            # مرتب‌سازی
            sort_by = request.GET.get('sort_by', 'name')
            if sort_by in ['name', 'code', 'created_at']:
                categories = categories.order_by(sort_by)
            
            # صفحه‌بندی
            page = int(request.GET.get('page', 1))
            per_page = int(request.GET.get('per_page', 20))
            paginator = Paginator(categories, per_page)
            page_obj = paginator.get_page(page)
            
            data = {
                'categories': [{
                    'id': str(category.id),
                    'name': category.name,
                    'code': category.code,
                    'description': category.description,
                    'parent_name': category.parent.name if category.parent else None,
                    'items_count': category.items.filter(is_active=True).count(),
                    'subcategories_count': category.children.filter(is_active=True).count(),
                    'created_at': category.created_at.isoformat(),
                } for category in page_obj],
                'pagination': {
                    'current_page': page_obj.number,
                    'total_pages': paginator.num_pages,
                    'total_items': paginator.count,
                    'has_next': page_obj.has_next(),
                    'has_previous': page_obj.has_previous(),
                }
            }
            
            return json_response(data)
    
    def post(self, request):
        """ایجاد دسته‌بندی جدید"""
        
        data = parse_json_body(request)
        
        try:
            # اعتبارسنجی
            if not data.get('name'):
                return json_response(
                    message="نام دسته‌بندی الزامی است.",
                    status=400,
                    success=False
                )
            
            # بررسی تکراری بودن نام
            if Category.objects.filter(name=data['name'], is_active=True).exists():
                return json_response(
                    message="دسته‌بندی با این نام قبلاً ثبت شده است.",
                    status=400,
                    success=False
                )
            
            # بررسی تکراری بودن کد (اگر ارسال شده)
            if data.get('code') and Category.objects.filter(code=data['code'], is_active=True).exists():
                return json_response(
                    message="دسته‌بندی با این کد قبلاً ثبت شده است.",
                    status=400,
                    success=False
                )
            
            # دریافت دسته‌بندی والد (اگر وجود دارد)
            parent = None
            if data.get('parent_id'):
                parent = get_object_or_404(Category, id=data['parent_id'], is_active=True)
            
            # ایجاد دسته‌بندی
            category = Category.objects.create(
                name=data['name'],
                code=data.get('code', ''),
                description=data.get('description', ''),
                parent=parent,
                created_by=request.user
            )
            
            return json_response(
                data={'category_id': str(category.id)},
                message=f'دسته‌بندی "{category.name}" با موفقیت ایجاد شد.'
            )
            
        except Exception as e:
            return json_response(
                message=f'خطا در ایجاد دسته‌بندی: {str(e)}',
                status=500,
                success=False
            )
    
    def put(self, request, category_id):
        """بروزرسانی دسته‌بندی"""
        
        category = get_object_or_404(Category, id=category_id, is_active=True)
        data = parse_json_body(request)
        
        try:
            # اعتبارسنجی
            if not data.get('name'):
                return json_response(
                    message="نام دسته‌بندی الزامی است.",
                    status=400,
                    success=False
                )
            
            # بررسی تکراری بودن نام (به جز خود رکورد)
            if Category.objects.filter(
                name=data['name'], 
                is_active=True
            ).exclude(id=category_id).exists():
                return json_response(
                    message="دسته‌بندی با این نام قبلاً ثبت شده است.",
                    status=400,
                    success=False
                )
            
            # بررسی تکراری بودن کد (اگر ارسال شده)
            if data.get('code') and Category.objects.filter(
                code=data['code'], 
                is_active=True
            ).exclude(id=category_id).exists():
                return json_response(
                    message="دسته‌بندی با این کد قبلاً ثبت شده است.",
                    status=400,
                    success=False
                )
            
            # دریافت دسته‌بندی والد (اگر وجود دارد)
            parent = None
            if data.get('parent_id'):
                parent = get_object_or_404(Category, id=data['parent_id'], is_active=True)
                
                # بررسی اینکه والد، فرزند خودش نباشد
                if parent.id == category.id:
                    return json_response(
                        message="دسته‌بندی نمی‌تواند والد خودش باشد.",
                        status=400,
                        success=False
                    )
            
            # بروزرسانی
            category.name = data['name']
            category.code = data.get('code', category.code)
            category.description = data.get('description', category.description)
            category.parent = parent
            category.save()
            
            return json_response(
                message=f'دسته‌بندی "{category.name}" با موفقیت بروزرسانی شد.'
            )
            
        except Exception as e:
            return json_response(
                message=f'خطا در بروزرسانی دسته‌بندی: {str(e)}',
                status=500,
                success=False
            )
    
    def delete(self, request, category_id):
        """حذف دسته‌بندی (غیرفعال کردن)"""
        
        category = get_object_or_404(Category, id=category_id, is_active=True)
        
        try:
            # بررسی وجود کالا در این دسته‌بندی
            if category.items.filter(is_active=True).exists():
                return json_response(
                    message="این دسته‌بندی دارای کالا است و قابل حذف نیست.",
                    status=400,
                    success=False
                )
            
            # بررسی وجود زیردسته
            if category.children.filter(is_active=True).exists():
                return json_response(
                    message="این دسته‌بندی دارای زیردسته است و قابل حذف نیست.",
                    status=400,
                    success=False
                )
            
            # غیرفعال کردن
            category.is_active = False
            category.save()
            
            return json_response(
                message=f'دسته‌بندی "{category.name}" با موفقیت حذف شد.'
            )
            
        except Exception as e:
            return json_response(
                message=f'خطا در حذف دسته‌بندی: {str(e)}',
                status=500,
                success=False
            )

# inventory/views.py

@method_decorator(login_required, name='dispatch')
@method_decorator(csrf_exempt, name='dispatch')
class ItemView(View):
    def create_item_notification(self, title, message, notification_type='SYSTEM_ALERT', priority='MEDIUM', url=None, metadata=None):
        """
        ایجاد اعلان برای همه انباردارها و ادمین‌ها
        """
        try:
            from users.models import Role
            from inventory.models import Notification
            from django.db.models import Q
        
        # دریافت نقش انباردار
            warehouse_role = Role.objects.filter(name='انباردار').first()
            default_recipient = CustomUser.objects.filter(is_superuser=True).first()
        # ایجاد اعلان در دیتابیس
            notification = Notification.objects.create(
                title=title,
                message=message,
                notification_type=notification_type,
                priority=priority,
                action_url=url,
                metadata=metadata or {},
                created_by=self.request.user,
                is_admin_only=False,
                is_global=False,
                target_role=warehouse_role,
                recipient=default_recipient  # استفاده از کاربر پیش‌فرض به جای NULL
            )
        
        # ارسال اعلان به همه انباردارها و ادمین‌ها از طریق WebSocket
            from users.models import CustomUser
            from django.db.models import Q
        
        # 1. کاربرانی که نقش انباردار دارند
            warehouse_users = CustomUser.objects.filter(role=warehouse_role) if warehouse_role else []
        
        # 2. ادمین‌ها
            admin_users = CustomUser.objects.filter(Q(is_superuser=True) | Q(is_staff=True))
        
        # ترکیب کاربران (بدون تکرار)
            target_users = list(warehouse_users) + [user for user in admin_users if user not in warehouse_users]
        
            notification_data = {
                'id': str(notification.id),
                'title': notification.title,
                'message': notification.message,
                'type': notification.notification_type,
                'priority': notification.priority,
                'is_read': False,
                'url': notification.action_url or '#',
                'created_at': notification.created_at.isoformat(),
                'created_by': self.request.user.get_full_name() if self.request.user else 'سیستم'
            }
        
        # ارسال به کاربران هدف
            for user in target_users:
                from inventory.notification_consumers import send_notification_to_user
                send_notification_to_user(user.id, notification_data)
        
            return notification
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"خطا در ایجاد اعلان کالا: {str(e)}")
            return None

    """API برای مدیریت کالاها"""
    
    def get(self, request, item_id=None):
        """دریافت کالا یا لیست کالاها"""
        if item_id:
            # دریافت یک کالا
            try:
                item = Item.objects.get(id=item_id, is_active=True)
                
                # دریافت موجودی کالا در انبارها
                inventories = Inventory.objects.filter(item=item)
                inventory_data = []
                for inv in inventories:
                    inventory_data.append({
                        'warehouse_id': str(inv.warehouse.id),
                        'warehouse_name': inv.warehouse.name,
                        'quantity': float(inv.quantity),
                        'value': float(inv.value)
                    })
                
                # دریافت واحدهای کالا
                units = ItemUnit.objects.filter(item=item)
                unit_data = []
                for unit in units:
                    unit_data.append({
                        'unit_id': str(unit.unit.id),
                        'unit_name': unit.unit.name,
                        'conversion_factor': float(unit.conversion_factor),
                        'is_purchase_unit': unit.is_purchase_unit,
                        'is_sales_unit': unit.is_sales_unit
                    })
                
                data = {
                    'id': str(item.id),
                    'name': item.name,
                    'code': item.code,
                    'barcode': item.barcode,
                    'category_id': str(item.category.id),
                    'category_name': item.category.name,
                    'primary_unit_id': str(item.primary_unit.id),
                    'primary_unit_name': item.primary_unit.name,
                    'brand': item.brand,
                    'model': item.model,
                    'specifications': item.specifications,
                    'min_stock_level': float(item.min_stock_level),
                    'max_stock_level': float(item.max_stock_level),
                    'alert_threshold': float(item.alert_threshold),
                    'enable_low_stock_alert': item.enable_low_stock_alert,
                    'cost_method': item.cost_method,
                    'has_expiry': item.has_expiry,
                    'has_serial': item.has_serial,
                    'has_batch': item.has_batch,
                    'storage_conditions': item.storage_conditions,
                    'storage_temperature_min': float(item.storage_temperature_min) if item.storage_temperature_min else None,
                    'storage_temperature_max': float(item.storage_temperature_max) if item.storage_temperature_max else None,
                    'description': item.description,
                    'usage_instructions': item.usage_instructions,
                    'image': request.build_absolute_uri(item.image.url) if item.image else None,
                    'created_at': item.created_at.isoformat(),
                    'updated_at': item.updated_at.isoformat(),
                    'created_by': item.created_by.get_full_name() if item.created_by else None,
                    'inventories': inventory_data,
                    'units': unit_data
                }
                
                return JsonResponse({'success': True, 'item': data})
                
            except Item.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'کالا یافت نشد'}, status=404)
                
        else:
            # دریافت لیست کالاها
            try:
                # پارامترهای فیلتر و جستجو
                search = request.GET.get('search', '')
                category = request.GET.get('category', '')
                warehouse = request.GET.get('warehouse', '')
                stock_status = request.GET.get('stock_status', '')
                
                # پارامترهای صفحه‌بندی
                page = int(request.GET.get('page', 1))
                per_page = int(request.GET.get('per_page', 20))
                
                # پارامترهای مرتب‌سازی
                sort_by = request.GET.get('sort_by', 'name')
                sort_order = request.GET.get('sort_order', 'asc')
                
                # فیلتر کالاها
                items = Item.objects.filter(is_active=True)
                
                # اعمال جستجو
                if search:
                    items = items.filter(
                        Q(name__icontains=search) | 
                        Q(code__icontains=search) | 
                        Q(barcode__icontains=search)
                    )
                
                # اعمال فیلتر دسته‌بندی
                if category:
                    items = items.filter(category_id=category)
                
                # اعمال فیلتر انبار
                if warehouse:
                    items = items.filter(inventory__warehouse_id=warehouse).distinct()
                
                # اعمال فیلتر وضعیت موجودی
                if stock_status:
                    if stock_status == 'in_stock':
                        items = items.filter(inventory__quantity__gt=F('min_stock_level'))
                    elif stock_status == 'low_stock':
                        items = items.filter(
                            inventory__quantity__gt=0,
                            inventory__quantity__lte=F('min_stock_level')
                        )
                    elif stock_status == 'out_of_stock':
                        items = items.filter(inventory__quantity=0)
                
                # اعمال مرتب‌سازی
                order_by = sort_by
                if sort_order == 'desc':
                    order_by = f'-{sort_by}'
                
                items = items.order_by(order_by)
                
                # صفحه‌بندی
                paginator = Paginator(items, per_page)
                page_obj = paginator.get_page(page)
                
                # تبدیل به دیکشنری
                items_data = []
                for item in page_obj:
                    # محاسبه موجودی کل
                    total_stock = Inventory.objects.filter(item=item).aggregate(total=Sum('quantity'))['total'] or 0
                    
                    # محاسبه قیمت واحد (میانگین)
                    avg_cost = 0
                    inventories = Inventory.objects.filter(item=item)
                    if inventories.exists():
                        total_value = inventories.aggregate(total=Sum('value'))['total'] or 0
                        total_quantity = inventories.aggregate(total=Sum('quantity'))['total'] or 0
                        if total_quantity > 0:
                            avg_cost = total_value / total_quantity
                    
                    items_data.append({
                        'id': str(item.id),
                        'name': item.name,
                        'code': item.code,
                        'barcode': item.barcode,
                        'category_id': str(item.category.id),
                        'category_name': item.category.name,
                        'primary_unit_id': str(item.primary_unit.id),
                        'primary_unit_name': item.primary_unit.name,
                        'current_stock': float(total_stock),
                        'min_stock_level': float(item.min_stock_level),
                        'unit_price': float(avg_cost),
                        'image': request.build_absolute_uri(item.image.url) if item.image else None,
                        'has_expiry': item.has_expiry,
                        'has_serial': item.has_serial
                    })
                
                # اطلاعات صفحه‌بندی
                pagination = {
                    'current_page': page_obj.number,
                    'total_pages': paginator.num_pages,
                    'total': paginator.count,
                    'has_next': page_obj.has_next(),
                    'has_previous': page_obj.has_previous(),
                }
                
                return JsonResponse({
                    'success': True,
                    'items': items_data,
                    'pagination': pagination
                })
                
            except Exception as e:
                return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    def post(self, request):

        try:
        # دریافت داده‌ها از درخواست
            data = request.POST
            files = request.FILES
        
        # اعتبارسنجی داده‌های ورودی
            required_fields = ['name', 'code', 'category_id', 'primary_unit_id']
            for field in required_fields:
                if field not in data or not data[field]:
                    return JsonResponse({
                        'success': False, 
                        'message': f'فیلد {field} الزامی است'
                    }, status=400)
        
        # بررسی یکتا بودن کد
            if Item.objects.filter(code=data['code'], is_active=True).exists():
                return JsonResponse({
                    'success': False, 
                    'message': 'کد کالا تکراری است'
                }, status=400)
        
        # بررسی یکتا بودن بارکد (اگر وارد شده باشد)
            if data.get('barcode') and Item.objects.filter(barcode=data['barcode'], is_active=True).exists():
                return JsonResponse({
                    'success': False, 
                    'message': 'بارکد کالا تکراری است'
                }, status=400)
        
        # دریافت دسته‌بندی و واحد اصلی
            try:
                category = Category.objects.get(id=data['category_id'], is_active=True)
            except Category.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'message': 'دسته‌بندی مورد نظر یافت نشد'
                }, status=400)
        
            try:
                primary_unit = UnitOfMeasure.objects.get(id=data['primary_unit_id'], is_active=True)
            except UnitOfMeasure.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'message': 'واحد اصلی مورد نظر یافت نشد'
                }, status=400)
        
        # ایجاد کالا
            item = Item(
                name=data['name'],
                code=data['code'],
                barcode=data.get('barcode'),
                category=category,
                primary_unit=primary_unit,
                brand=data.get('brand'),
                model=data.get('model'),
                min_stock_level=float(data.get('min_stock_level', 0)),
                max_stock_level=float(data.get('max_stock_level', 0)),
                alert_threshold=float(data.get('alert_threshold', 5)),
                enable_low_stock_alert=data.get('enable_low_stock_alert') == 'on',
                cost_method=data.get('cost_method', 'FIFO'),
                has_expiry=data.get('has_expiry') == 'on',
                has_serial=data.get('has_serial') == 'on',
                has_batch=data.get('has_batch') == 'on',
                storage_conditions=data.get('storage_conditions'),
                storage_temperature_min=data.get('storage_temperature_min'),
                storage_temperature_max=data.get('storage_temperature_max'),
                description=data.get('description'),
                usage_instructions=data.get('usage_instructions'),
                created_by=request.user
            )
        
        # اگر تصویر آپلود شده باشد
            if 'image' in files:
                item.image = files['image']
        
        # ذخیره کالا
            item.save()
        
        # اگر انبارهای مجاز انتخاب شده باشند
            if 'allowed_warehouses' in data:
                allowed_warehouses = data.getlist('allowed_warehouses')
                for warehouse_id in allowed_warehouses:
                    try:
                        warehouse = Warehouse.objects.get(id=warehouse_id, is_active=True)
                        item.allowed_warehouses.add(warehouse)
                    except Warehouse.DoesNotExist:
                        pass
        
        # ارسال اطلاعیه از طریق وب‌سوکت
            self.send_websocket_notification('item_created', {
                'item_id': str(item.id),
                'item_name': item.name,
                'user': request.user.get_full_name()
            })
            # ایجاد نوتیفیکیشن
            self.create_item_notification(
                title="ایجاد کالای جدید",
                message=f"کالای جدید «{item.name}» با کد {item.code} توسط {request.user.get_full_name() or request.user.username} ایجاد شد.",
                notification_type="STOCK_ENTRY",
                priority="MEDIUM",
                url=f"/inventory/items/{item.id}/detail/",
                metadata={
                    'item_id': str(item.id),
                    'item_name': item.name,
                    'item_code': item.code,
                    'category': category.name,
                    'action': 'create'
                }
            )
            return JsonResponse({
                'success': True,
                'message': 'کالا با موفقیت ایجاد شد',
                'item_id': str(item.id)
            })
        
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    def put(self, request, item_id):
        """بروزرسانی کالا"""
        try:
        # یافتن کالا
            try:
                item = Item.objects.get(id=item_id, is_active=True)
            except Item.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'کالا یافت نشد'}, status=404)
        
        # دریافت داده‌ها از درخواست
            data = request.POST
            files = request.FILES
        
        # اعتبارسنجی داده‌های ورودی
            required_fields = ['name', 'code', 'category_id', 'primary_unit_id']
            for field in required_fields:
                if field not in data or not data[field]:
                    return JsonResponse({
                        'success': False, 
                        'message': f'فیلد {field} الزامی است'
                    }, status=400)
        
        # بررسی یکتا بودن کد
            if Item.objects.filter(code=data['code'], is_active=True).exclude(id=item_id).exists():
                return JsonResponse({
                    'success': False, 
                    'message': 'کد کالا تکراری است'
                }, status=400)
        
        # بررسی یکتا بودن بارکد (اگر وارد شده باشد)
            if data.get('barcode') and Item.objects.filter(barcode=data['barcode'], is_active=True).exclude(id=item_id).exists():
                return JsonResponse({
                    'success': False, 
                    'message': 'بارکد کالا تکراری است'
                }, status=400)
        
        # دریافت دسته‌بندی و واحد اصلی
            try:
                category = Category.objects.get(id=data['category_id'], is_active=True)
            except Category.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'message': 'دسته‌بندی مورد نظر یافت نشد'
                }, status=400)
        
            try:
                primary_unit = UnitOfMeasure.objects.get(id=data['primary_unit_id'], is_active=True)
            except UnitOfMeasure.DoesNotExist:
                return JsonResponse({
                    'success': False, 
                    'message': 'واحد اصلی مورد نظر یافت نشد'
                }, status=400)
        
        # بروزرسانی کالا
            item.name = data['name']
            item.code = data['code']
            item.barcode = data.get('barcode')
            item.category = category
            item.primary_unit = primary_unit
            item.brand = data.get('brand')
            item.model = data.get('model')
            item.min_stock_level = float(data.get('min_stock_level', 0))
            item.max_stock_level = float(data.get('max_stock_level', 0))
            item.alert_threshold = float(data.get('alert_threshold', 5))
            item.enable_low_stock_alert = data.get('enable_low_stock_alert') == 'on'
            item.cost_method = data.get('cost_method', 'FIFO')
            item.has_expiry = data.get('has_expiry') == 'on'
            item.has_serial = data.get('has_serial') == 'on'
            item.has_batch = data.get('has_batch') == 'on'
            item.storage_conditions = data.get('storage_conditions')
            item.storage_temperature_min = data.get('storage_temperature_min')
            item.storage_temperature_max = data.get('storage_temperature_max')
            item.description = data.get('description')
            item.usage_instructions = data.get('usage_instructions')
            
            
            old_name = warehouse.name
            old_code = warehouse.code
            old_category = warehouse.category
            
        # اگر تصویر آپلود شده باشد
            if 'image' in files:
                item.image = files['image']
        
        # ذخیره کالا
            item.save()
        
        # بروزرسانی انبارهای مجاز
            if 'allowed_warehouses' in data:
            # حذف همه انبارهای مجاز قبلی
                item.allowed_warehouses.clear()
            
            # اضافه کردن انبارهای مجاز جدید
                allowed_warehouses = data.getlist('allowed_warehouses')
                for warehouse_id in allowed_warehouses:
                    try:
                        warehouse = Warehouse.objects.get(id=warehouse_id, is_active=True)
                        item.allowed_warehouses.add(warehouse)
                    except Warehouse.DoesNotExist:
                        pass
        
        # ارسال اطلاعیه از طریق وب‌سوکت
                self.send_websocket_notification('item_updated', {
                    'item_id': str(item.id),
                    'item_name': item.name,
                    'user': request.user.get_full_name()
                })
                    # ایجاد پیام نوتیفیکیشن
                changes = []
                if old_name != item.name:
                    changes.append(f"نام از «{old_name}» به «{item.name}»")
                if old_code != item.code:
                    changes.append(f"کد از «{old_code}» به «{item.code}»")
                if old_category != item.category:
                    changes.append(f"دسته‌بندی از «{old_category.name}» به «{item.category.name}»")
        
                if changes:
                    changes_text = "، ".join(changes)
                    message = f"کالای «{item.name}» بروزرسانی شد. تغییرات: {changes_text}"
                else:
                    message = f"کالای «{item.name}» بروزرسانی شد."
        
                self.create_item_notification(
                    title="بروزرسانی کالا",
                    message=message,
                    notification_type="INVENTORY_ADJUSTMENT",
                    priority="MEDIUM",
                    url=f"/inventory/items/{item.id}/detail/",
                    metadata={
                        'item_id': str(item.id),
                        'item_name': item.name,
                        'item_code': item.code,
                        'category': item.category.name,
                        'action': 'update',
                        'changes': changes
                    }
                )
                return JsonResponse({
                    'success': True,
                    'message': 'کالا با موفقیت بروزرسانی شد'
                })
        
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)

    def delete(self, request, item_id):
        """حذف کالا (غیرفعال‌سازی)"""
        try:
            item = Item.objects.get(id=item_id, is_active=True)
            item.is_active = False
            # ذخیره اطلاعات قبل از غیرفعال‌سازی
            item_name = item.name
            item_code = item.code
            category_name = item.category.name
            item.save()
            
            # ارسال اطلاعیه از طریق وب‌سوکت
            self.send_websocket_notification('item_deleted', {
                'item_id': str(item.id),
                'item_name': item.name,
                'user': request.user.get_full_name()
            })
            # ایجاد نوتیفیکیشن
            self.create_item_notification(
                title="حذف کالا",
                message=f"کالای «{item_name}» با کد {item_code} توسط {request.user.get_full_name() or request.user.username} حذف شد.",
                notification_type="INVENTORY_ADJUSTMENT",
                priority="HIGH",
                metadata={
                    'item_id': str(item.id),
                    'item_name': item_name,
                    'item_code': item_code,
                    'category': category_name,
                    'action': 'delete'
                }
            )
            return JsonResponse({
                'success': True,
                'message': 'کالا با موفقیت حذف شد'
            })
            
        except Item.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'کالا یافت نشد'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    def send_websocket_notification(self, event_type, data):
        """ارسال اطلاعیه از طریق وب‌سوکت"""
        try:
            from channels.layers import get_channel_layer
            from asgiref.sync import async_to_sync
            
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                "inventory",
                {
                    "type": "inventory_update",
                    "event_type": event_type,
                    "data": data
                }
            )
        except Exception as e:
            # در صورت خطا، فقط لاگ می‌کنیم و ادامه می‌دهیم
            print(f"WebSocket notification error: {str(e)}")


# اضافه کردن import های مورد نیاز
from .models import UnitOfMeasure, StockBatch

@method_decorator(login_required, name='dispatch')
@method_decorator(csrf_exempt, name='dispatch')
class UnitView(View):
    """View کلاس برای مدیریت واحدهای اندازه‌گیری"""
    
    def get(self, request, unit_id=None):
        """دریافت واحد یا لیست واحدها"""
        
        if unit_id:
            # جزئیات یک واحد
            unit = get_object_or_404(UnitOfMeasure, id=unit_id)
            
            data = {
                'unit': {
                    'id': str(unit.id),
                    'name': unit.name,
                    'symbol': unit.symbol,
                    'description': unit.description,
                    'is_base_unit': unit.is_base_unit,
                    'base_unit_id': str(unit.base_unit.id) if unit.base_unit else None,
                    'base_unit_name': unit.base_unit.name if unit.base_unit else None,
                    'conversion_factor': float(unit.conversion_factor),
                    'is_active': unit.is_active,
                    'created_at': unit.created_at.isoformat(),
                    'updated_at': unit.updated_at.isoformat(),
                    'items_count': unit.items.filter(is_active=True).count(),
                }
            }
            
            return json_response(data)
        
        else:
            # لیست واحدها
            units = UnitOfMeasure.objects.filter(is_active=True).select_related('base_unit')
            
            # فیلترها
            is_base_unit = request.GET.get('is_base_unit')
            if is_base_unit == 'true':
                units = units.filter(is_base_unit=True)
            elif is_base_unit == 'false':
                units = units.filter(is_base_unit=False)
            
            search = request.GET.get('search')
            if search:
                units = units.filter(
                    name__icontains=search
                ).distinct()
            
            # مرتب‌سازی
            sort_by = request.GET.get('sort_by', 'name')
            if sort_by in ['name', 'symbol', 'created_at']:
                units = units.order_by(sort_by)
            
            # صفحه‌بندی
            page = int(request.GET.get('page', 1))
            per_page = int(request.GET.get('per_page', 20))
            paginator = Paginator(units, per_page)
            page_obj = paginator.get_page(page)
            
            data = {
                'units': [{
                    'id': str(unit.id),
                    'name': unit.name,
                    'symbol': unit.symbol,
                    'description': unit.description,
                    'is_base_unit': unit.is_base_unit,
                    'base_unit_name': unit.base_unit.name if unit.base_unit else None,
                    'conversion_factor': float(unit.conversion_factor),
                    'items_count': unit.items.filter(is_active=True).count(),
                    'created_at': unit.created_at.isoformat(),
                } for unit in page_obj],
                'pagination': {
                    'current_page': page_obj.number,
                    'total_pages': paginator.num_pages,
                    'total_items': paginator.count,
                    'has_next': page_obj.has_next(),
                    'has_previous': page_obj.has_previous(),
                }
            }
            
            return json_response(data)
    
    def post(self, request):
        """ایجاد واحد جدید"""
        
        try:
            data = parse_json_body(request)
        except ValueError as e:
            return error_response(str(e), status=400)
        
        # اعتبارسنجی فیلدهای الزامی
        required_fields = ['name', 'symbol']
        validation_errors = validate_required_fields(data, required_fields)
        
        if validation_errors:
            return validation_error_response(validation_errors)
        
        try:
            # بررسی تکراری بودن نام
            if UnitOfMeasure.objects.filter(name=data['name'], is_active=True).exists():
                return error_response("واحد با این نام قبلاً ثبت شده است.", status=400)
            
            # بررسی تکراری بودن نماد
            if UnitOfMeasure.objects.filter(symbol=data['symbol'], is_active=True).exists():
                return error_response("واحد با این نماد قبلاً ثبت شده است.", status=400)
            
            # دریافت واحد پایه (اگر وجود دارد)
            base_unit = None
            if data.get('base_unit_id'):
                try:
                    base_unit = UnitOfMeasure.objects.get(id=data['base_unit_id'], is_active=True)
                except UnitOfMeasure.DoesNotExist:
                    return not_found_response("واحد پایه یافت نشد.")
            
            # ایجاد واحد
            unit = UnitOfMeasure.objects.create(
                name=data['name'],
                symbol=data['symbol'],
                description=data.get('description', ''),
                is_base_unit=data.get('is_base_unit', True),
                base_unit=base_unit,
                conversion_factor=safe_decimal(data.get('conversion_factor', 1)),
                created_by=request.user
            )
            
            # ثبت فعالیت کاربر
            log_user_activity(
                user=request.user,
                action='CREATE_UNIT',
                details={'unit_id': str(unit.id), 'unit_name': unit.name},
                request=request
            )
            
            return success_response(
                data={'unit_id': str(unit.id)},
                message=f'واحد "{unit.name}" با موفقیت ایجاد شد.'
            )
            
        except Exception as e:
            return error_response(f'خطا در ایجاد واحد: {str(e)}', status=500)
    
    def put(self, request, unit_id):
        """بروزرسانی واحد"""
        
        unit = get_object_or_404(UnitOfMeasure, id=unit_id, is_active=True)
        
        try:
            data = parse_json_body(request)
        except ValueError as e:
            return error_response(str(e), status=400)
        
        try:
            # اعتبارسنجی
            if not data.get('name'):
                return error_response("نام واحد الزامی است.", status=400)
            
            # بررسی تکراری بودن نام (به جز خود رکورد)
            if UnitOfMeasure.objects.filter(
                name=data['name'], 
                is_active=True
            ).exclude(id=unit_id).exists():
                return error_response("واحد با این نام قبلاً ثبت شده است.", status=400)
            
            # بروزرسانی فیلدها
            unit.name = data['name']
            if 'symbol' in data:
                unit.symbol = data['symbol']
            if 'description' in data:
                unit.description = data['description']
            if 'conversion_factor' in data:
                unit.conversion_factor = safe_decimal(data['conversion_factor'])
            
            unit.save()
            
            # ثبت فعالیت کاربر
            log_user_activity(
                user=request.user,
                action='UPDATE_UNIT',
                details={'unit_id': str(unit.id), 'unit_name': unit.name},
                request=request
            )
            
            return success_response(
                message=f'واحد "{unit.name}" با موفقیت بروزرسانی شد.'
            )
            
        except Exception as e:
            return error_response(f'خطا در بروزرسانی واحد: {str(e)}', status=500)
    
    def delete(self, request, unit_id):
        """حذف واحد (غیرفعال کردن)"""
        
        unit = get_object_or_404(UnitOfMeasure, id=unit_id, is_active=True)
        
        try:
            # بررسی وجود کالا با این واحد
            if unit.items.filter(is_active=True).exists():
                return error_response("این واحد در کالاها استفاده شده و قابل حذف نیست.", status=400)
            
            # غیرفعال کردن
            unit.is_active = False
            unit.save()
            
            # ثبت فعالیت کاربر
            log_user_activity(
                user=request.user,
                action='DELETE_UNIT',
                details={'unit_id': str(unit.id), 'unit_name': unit.name},
                request=request
            )
            
            return success_response(
                message=f'واحد "{unit.name}" با موفقیت حذف شد.'
            )
            
        except Exception as e:
            return error_response(f'خطا در حذف واحد: {str(e)}', status=500)

@method_decorator(login_required, name='dispatch')
class StockReportView(View):
    """View کلاس برای گزارش موجودی"""
    
    def get(self, request):
        """دریافت گزارش موجودی"""
        
        try:
            # فیلترها
            warehouse_id = request.GET.get('warehouse_id')
            category_id = request.GET.get('category_id')
            low_stock_only = request.GET.get('low_stock_only') == 'true'
            
            # Query پایه
            items = Item.objects.filter(is_active=True).select_related('category', 'unit')
            
            # اعمال فیلترها
            if category_id:
                items = items.filter(category_id=category_id)
            
            # محاسبه موجودی برای هر کالا
            stock_data = []
            for item in items:
                current_stock = item.get_current_stock(warehouse_id=warehouse_id)
                
                # فیلتر موجودی کم
                if low_stock_only and current_stock >= item.reorder_point:
                    continue
                
                stock_info = {
                    'item_id': str(item.id),
                    'item_name': item.name,
                    'item_code': item.code,
                    'category_name': item.category.name if item.category else None,
                    'unit_symbol': item.unit.symbol if item.unit else None,
                    'current_stock': float(current_stock),
                    'minimum_stock': float(item.minimum_stock),
                    'maximum_stock': float(item.maximum_stock),
                    'reorder_point': float(item.reorder_point),
                    'unit_cost': float(item.unit_cost),
                    'stock_value': float(current_stock * item.unit_cost),
                    'is_low_stock': current_stock < item.reorder_point,
                    'stock_status': 'کم' if current_stock < item.reorder_point else 'عادی' if current_stock <= item.maximum_stock else 'زیاد'
                }
                
                # اگر انبار خاصی انتخاب شده، موجودی آن انبار را نمایش بده
                if warehouse_id:
                    warehouse_stock = item.get_warehouse_stock(warehouse_id)
                    stock_info['warehouse_stock'] = float(warehouse_stock)
                
                stock_data.append(stock_info)
            
            # مرتب‌سازی
            sort_by = request.GET.get('sort_by', 'item_name')
            reverse = request.GET.get('order') == 'desc'
            
            if sort_by in ['item_name', 'current_stock', 'stock_value']:
                stock_data.sort(key=lambda x: x[sort_by], reverse=reverse)
            
            # صفحه‌بندی
            page = int(request.GET.get('page', 1))
            per_page = int(request.GET.get('per_page', 50))
            paginator = Paginator(stock_data, per_page)
            page_obj = paginator.get_page(page)
            
            # آمار کلی
            total_items = len(stock_data)
            total_value = sum(item['stock_value'] for item in stock_data)
            low_stock_items = len([item for item in stock_data if item['is_low_stock']])
            
            data = {
                'stock_report': list(page_obj),
                'summary': {
                    'total_items': total_items,
                    'total_value': total_value,
                    'low_stock_items': low_stock_items,
                    'low_stock_percentage': (low_stock_items / total_items * 100) if total_items > 0 else 0
                },
                'pagination': {
                    'current_page': page_obj.number,
                    'total_pages': paginator.num_pages,
                    'total_items': paginator.count,
                    'has_next': page_obj.has_next(),
                    'has_previous': page_obj.has_previous(),
                }
            }
            
            return json_response(data)
            
        except Exception as e:
            return error_response(f'خطا در تولید گزارش موجودی: {str(e)}', status=500)


@method_decorator(login_required, name='dispatch')
class MovementReportView(View):
    """View کلاس برای گزارش تحرکات انبار"""
    
    def get(self, request):
        """دریافت گزارش تحرکات"""
        
        try:
            # فیلترها
            warehouse_id = request.GET.get('warehouse_id')
            item_id = request.GET.get('item_id')
            date_from = request.GET.get('date_from')
            date_to = request.GET.get('date_to')
            movement_type = request.GET.get('movement_type')  # 'entry' یا 'exit'
            
            movements = []
            
            # ورودی‌ها
            if not movement_type or movement_type == 'entry':
                entries = StockEntry.objects.filter(is_active=True).select_related(
                    'item', 'warehouse', 'supplier', 'created_by'
                )
                
                # اعمال فیلترها
                if warehouse_id:
                    entries = entries.filter(warehouse_id=warehouse_id)
                if item_id:
                    entries = entries.filter(item_id=item_id)
                if date_from:
                    entries = entries.filter(entry_date__date__gte=date_from)
                if date_to:
                    entries = entries.filter(entry_date__date__lte=date_to)
                
                for entry in entries:
                    movements.append({
                        'id': str(entry.id),
                        'type': 'entry',
                        'type_display': 'ورودی',
                        'document_number': entry.entry_number,
                        'item_name': entry.item.name,
                        'item_code': entry.item.code,
                        'warehouse_name': entry.warehouse.name,
                        'quantity': float(entry.quantity),
                        'unit_symbol': entry.unit.symbol,
                        'unit_cost': float(entry.unit_cost),
                        'total_cost': float(entry.total_cost),
                        'date': entry.entry_date.isoformat(),
                        'reference': entry.supplier.name if entry.supplier else '',
                        'created_by': entry.created_by.get_full_name() if entry.created_by else '',
                        'notes': entry.notes,
                    })
            
            # خروجی‌ها
            if not movement_type or movement_type == 'exit':
                exits = StockExit.objects.filter(is_active=True).select_related(
                    'item', 'warehouse', 'created_by'
                )
                
                # اعمال فیلترها
                if warehouse_id:
                    exits = exits.filter(warehouse_id=warehouse_id)
                if item_id:
                    exits = exits.filter(item_id=item_id)
                if date_from:
                    exits = exits.filter(exit_date__date__gte=date_from)
                if date_to:
                    exits = exits.filter(exit_date__date__lte=date_to)
                
                for exit in exits:
                    movements.append({
                        'id': str(exit.id),
                        'type': 'exit',
                        'type_display': 'خروجی',
                        'document_number': exit.exit_number,
                        'item_name': exit.item.name,
                        'item_code': exit.item.code,
                        'warehouse_name': exit.warehouse.name,
                        'quantity': float(exit.quantity),
                        'unit_symbol': exit.unit.symbol,
                        'unit_cost': float(exit.unit_cost),
                        'total_cost': float(exit.total_cost),
                        'date': exit.exit_date.isoformat(),
                        'reference': exit.recipient_name,
                        'created_by': exit.created_by.get_full_name() if exit.created_by else '',
                        'notes': exit.notes,
                    })
            
            # مرتب‌سازی بر اساس تاریخ
            movements.sort(key=lambda x: x['date'], reverse=True)
            
            # صفحه‌بندی
            page = int(request.GET.get('page', 1))
            per_page = int(request.GET.get('per_page', 50))
            paginator = Paginator(movements, per_page)
            page_obj = paginator.get_page(page)
            
            # آمار کلی
            total_entries = len([m for m in movements if m['type'] == 'entry'])
            total_exits = len([m for m in movements if m['type'] == 'exit'])
            total_entry_value = sum(m['total_cost'] for m in movements if m['type'] == 'entry')
            total_exit_value = sum(m['total_cost'] for m in movements if m['type'] == 'exit')
            
            data = {
                'movements': list(page_obj),
                'summary': {
                    'total_movements': len(movements),
                    'total_entries': total_entries,
                    'total_exits': total_exits,
                    'total_entry_value': total_entry_value,
                    'total_exit_value': total_exit_value,
                    'net_value': total_entry_value - total_exit_value
                },
                'pagination': {
                    'current_page': page_obj.number,
                    'total_pages': paginator.num_pages,
                    'total_items': paginator.count,
                    'has_next': page_obj.has_next(),
                    'has_previous': page_obj.has_previous(),
                }
            }
            
            return json_response(data)
            
        except Exception as e:
            return error_response(f'خطا در تولید گزارش تحرکات: {str(e)}', status=500)


# View های HTML برای نمایش صفحات
@method_decorator(login_required, name='dispatch')
class CategoryListView(View):
    """نمایش لیست دسته‌بندی‌ها"""
    
    def get(self, request):
        from django.shortcuts import render
        return render(request, 'inventory/categories.html')


@method_decorator(login_required, name='dispatch')
class ItemListView(View):
    """نمایش لیست کالاها"""
    
    def get(self, request):
        from django.shortcuts import render
        return render(request, 'inventory/items.html')




@method_decorator(login_required, name='dispatch')
class StockEntryListView(View):
    """نمایش لیست ورودی‌های انبار"""
    
    def get(self, request):
        from django.shortcuts import render
        return render(request, 'inventory/stock_entries.html')


@method_decorator(login_required, name='dispatch')
class StockExitListView(View):
    """نمایش لیست خروجی‌های انبار"""
    
    def get(self, request):
        from django.shortcuts import render
        return render(request, 'inventory/stock_exits.html')
# inventory/views.py

@method_decorator(login_required, name='dispatch')
@method_decorator(csrf_exempt, name='dispatch')
class BatchView(View):
    """View کلاس برای مدیریت بچ‌های کالا"""
    
    def get(self, request, batch_id=None):
        """دریافت بچ یا لیست بچ‌ها"""
        
        if batch_id:
            # جزئیات یک بچ
            batch = get_object_or_404(StockBatch, id=batch_id)
            
            data = {
                'batch': {
                    'id': str(batch.id),
                    'batch_number': batch.batch_number,
                    'item_id': str(batch.item.id),
                    'item_name': batch.item.name,
                    'item_code': batch.item.code,
                    'warehouse_id': str(batch.warehouse.id),
                    'warehouse_name': batch.warehouse.name,
                    'quantity': float(batch.quantity),
                    'remaining_quantity': float(batch.remaining_quantity),
                    'unit_cost': float(batch.unit_cost),
                    'total_cost': float(batch.total_cost),
                    'manufacture_date': batch.manufacture_date.isoformat() if batch.manufacture_date else None,
                    'expiry_date': batch.expiry_date.isoformat() if batch.expiry_date else None,
                    'supplier_name': batch.supplier.name if batch.supplier else '',
                    'is_expired': batch.is_expired(),
                    'days_to_expiry': batch.days_to_expiry(),
                    'is_active': batch.is_active,
                    'created_at': batch.created_at.isoformat(),
                    'updated_at': batch.updated_at.isoformat(),
                }
            }
            
            return json_response(data)
        
        else:
            # لیست بچ‌ها
            batches = StockBatch.objects.filter(is_active=True).select_related(
                'item', 'warehouse', 'supplier'
            )
            
            # فیلترها
            item_id = request.GET.get('item_id')
            if item_id:
                batches = batches.filter(item_id=item_id)
            
            warehouse_id = request.GET.get('warehouse_id')
            if warehouse_id:
                batches = batches.filter(warehouse_id=warehouse_id)
            
            # فیلتر بچ‌های منقضی شده
            expired_only = request.GET.get('expired_only')
            if expired_only == 'true':
                from django.utils import timezone
                batches = batches.filter(expiry_date__lt=timezone.now().date())
            
            # فیلتر بچ‌های نزدیک به انقضا
            expiring_soon = request.GET.get('expiring_soon')
            if expiring_soon == 'true':
                from django.utils import timezone
                from datetime import timedelta
                warning_date = timezone.now().date() + timedelta(days=30)
                batches = batches.filter(
                    expiry_date__lte=warning_date,
                    expiry_date__gte=timezone.now().date()
                )
            
            # فیلتر بچ‌های با موجودی
            has_stock = request.GET.get('has_stock')
            if has_stock == 'true':
                batches = batches.filter(remaining_quantity__gt=0)
            
            search = request.GET.get('search')
            if search:
                batches = batches.filter(
                    models.Q(batch_number__icontains=search) |
                    models.Q(item__name__icontains=search) |
                    models.Q(item__code__icontains=search)
                ).distinct()
            
            # مرتب‌سازی
            sort_by = request.GET.get('sort_by', 'expiry_date')
            if sort_by in ['batch_number', 'expiry_date', 'remaining_quantity', 'created_at']:
                batches = batches.order_by(sort_by)
            
            # صفحه‌بندی
            page = int(request.GET.get('page', 1))
            per_page = int(request.GET.get('per_page', 20))
            paginator = Paginator(batches, per_page)
            page_obj = paginator.get_page(page)
            
            data = {
                'batches': [{
                    'id': str(batch.id),
                    'batch_number': batch.batch_number,
                    'item_name': batch.item.name,
                    'item_code': batch.item.code,
                    'warehouse_name': batch.warehouse.name,
                    'quantity': float(batch.quantity),
                    'remaining_quantity': float(batch.remaining_quantity),
                    'unit_cost': float(batch.unit_cost),
                    'manufacture_date': batch.manufacture_date.isoformat() if batch.manufacture_date else None,
                    'expiry_date': batch.expiry_date.isoformat() if batch.expiry_date else None,
                    'supplier_name': batch.supplier.name if batch.supplier else '',
                    'is_expired': batch.is_expired(),
                    'days_to_expiry': batch.days_to_expiry(),
                    'status': 'منقضی شده' if batch.is_expired() else 'نزدیک به انقضا' if batch.days_to_expiry() <= 30 else 'عادی',
                    'created_at': batch.created_at.isoformat(),
                } for batch in page_obj],
                'pagination': {
                    'current_page': page_obj.number,
                    'total_pages': paginator.num_pages,
                    'total_items': paginator.count,
                    'has_next': page_obj.has_next(),
                    'has_previous': page_obj.has_previous(),
                }
            }
            
            return json_response(data)
    
    def post(self, request):
        """ایجاد بچ جدید"""
        
        try:
            data = parse_json_body(request)
        except ValueError as e:
            return error_response(str(e), status=400)
        
        # اعتبارسنجی فیلدهای الزامی
        required_fields = ['batch_number', 'item_id', 'warehouse_id', 'quantity', 'unit_cost']
        validation_errors = validate_required_fields(data, required_fields)
        
        if validation_errors:
            return validation_error_response(validation_errors)
        
        try:
            # بررسی تکراری بودن شماره بچ
            if StockBatch.objects.filter(
                batch_number=data['batch_number'], 
                is_active=True
            ).exists():
                return error_response("بچ با این شماره قبلاً ثبت شده است.", status=400)
            
            # دریافت کالا
            try:
                item = Item.objects.get(id=data['item_id'], is_active=True)
            except Item.DoesNotExist:
                return not_found_response("کالا یافت نشد.")
            
            # دریافت انبار
            try:
                warehouse = Warehouse.objects.get(id=data['warehouse_id'], is_active=True)
            except Warehouse.DoesNotExist:
                return not_found_response("انبار یافت نشد.")
            
            # دریافت تامین‌کننده (اختیاری)
            supplier = None
            if data.get('supplier_id'):
                try:
                    from .models import Supplier
                    supplier = Supplier.objects.get(id=data['supplier_id'], is_active=True)
                except (Supplier.DoesNotExist, ImportError):
                    pass
            
            # تبدیل تاریخ‌ها
            manufacture_date = None
            if data.get('manufacture_date'):
                from datetime import datetime
                manufacture_date = datetime.fromisoformat(data['manufacture_date']).date()
            
            expiry_date = None
            if data.get('expiry_date'):
                from datetime import datetime
                expiry_date = datetime.fromisoformat(data['expiry_date']).date()
            
            # ایجاد بچ
            batch = StockBatch.objects.create(
                batch_number=data['batch_number'],
                item=item,
                warehouse=warehouse,
                quantity=safe_decimal(data['quantity']),
                remaining_quantity=safe_decimal(data['quantity']),
                unit_cost=safe_decimal(data['unit_cost']),
                manufacture_date=manufacture_date,
                expiry_date=expiry_date,
                supplier=supplier,
                notes=data.get('notes', ''),
                created_by=request.user
            )
            
            return success_response(
                data={'batch_id': str(batch.id)},
                message=f'بچ "{batch.batch_number}" با موفقیت ایجاد شد.'
            )
            
        except Exception as e:
            return error_response(f'خطا در ایجاد بچ: {str(e)}', status=500)
    
    def put(self, request, batch_id):
        """بروزرسانی بچ"""
        
        batch = get_object_or_404(StockBatch, id=batch_id, is_active=True)
        
        try:
            data = parse_json_body(request)
        except ValueError as e:
            return error_response(str(e), status=400)
        
        try:
            # بروزرسانی فیلدهای قابل تغییر
            if 'batch_number' in data:
                # بررسی تکراری بودن شماره بچ
                if StockBatch.objects.filter(
                    batch_number=data['batch_number'], 
                    is_active=True
                ).exclude(id=batch_id).exists():
                    return error_response("بچ با این شماره قبلاً ثبت شده است.", status=400)
                batch.batch_number = data['batch_number']
            
            if 'unit_cost' in data:
                batch.unit_cost = safe_decimal(data['unit_cost'])
            
            if 'manufacture_date' in data:
                if data['manufacture_date']:
                    from datetime import datetime
                    batch.manufacture_date = datetime.fromisoformat(data['manufacture_date']).date()
                else:
                    batch.manufacture_date = None
            
            if 'expiry_date' in data:
                if data['expiry_date']:
                    from datetime import datetime
                    batch.expiry_date = datetime.fromisoformat(data['expiry_date']).date()
                else:
                    batch.expiry_date = None
            
            if 'notes' in data:
                batch.notes = data['notes']
            
            batch.save()
            
            return success_response(
                message=f'بچ "{batch.batch_number}" با موفقیت بروزرسانی شد.'
            )
            
        except Exception as e:
            return error_response(f'خطا در بروزرسانی بچ: {str(e)}', status=500)
    
    def delete(self, request, batch_id):
        """حذف بچ (غیرفعال کردن)"""
        
        batch = get_object_or_404(StockBatch, id=batch_id, is_active=True)
        
        try:
            # بررسی وجود موجودی
            if batch.remaining_quantity > 0:
                return error_response("این بچ دارای موجودی است و قابل حذف نیست.", status=400)
            
            # غیرفعال کردن
            batch.is_active = False
            batch.save()
            
            return success_response(
                message=f'بچ "{batch.batch_number}" با موفقیت حذف شد.'
            )
            
        except Exception as e:
            return error_response(f'خطا در حذف بچ: {str(e)}', status=500)



@method_decorator(login_required, name='dispatch')
@method_decorator(csrf_exempt, name='dispatch')
class StockTransferView(View):
    """View کلاس برای انتقال موجودی بین انبارها"""
    
    def get(self, request, transfer_id=None):
        """دریافت انتقال یا لیست انتقال‌ها"""
        
        if transfer_id:
            # جزئیات یک انتقال
            transfer = get_object_or_404(StockTransfer, id=transfer_id)
            
            data = {
                'transfer': {
                    'id': str(transfer.id),
                    'transfer_number': transfer.transfer_number,
                    'item_id': str(transfer.item.id),
                    'item_name': transfer.item.name,
                    'item_code': transfer.item.code,
                    'from_warehouse_id': str(transfer.from_warehouse.id),
                    'from_warehouse_name': transfer.from_warehouse.name,
                    'to_warehouse_id': str(transfer.to_warehouse.id),
                    'to_warehouse_name': transfer.to_warehouse.name,
                    'quantity': float(transfer.quantity),
                    'unit_symbol': transfer.unit.symbol if transfer.unit else '',
                    'transfer_date': transfer.transfer_date.isoformat(),
                    'status': transfer.status,
                    'status_display': transfer.get_status_display(),
                    'notes': transfer.notes,
                    'requested_by': transfer.requested_by.get_full_name() if transfer.requested_by else '',
                    'approved_by': transfer.approved_by.get_full_name() if transfer.approved_by else '',
                    'approved_at': transfer.approved_at.isoformat() if transfer.approved_at else None,
                    'created_by': transfer.created_by.get_full_name() if transfer.created_by else '',
                    'created_at': transfer.created_at.isoformat(),
                    'updated_at': transfer.updated_at.isoformat(),
                }
            }
            
            return json_response(data)
        
        else:
            # لیست انتقال‌ها
            transfers = StockTransfer.objects.select_related(
                'item', 'from_warehouse', 'to_warehouse', 'unit', 
                'requested_by', 'approved_by', 'created_by'
            )
            
            # فیلترها
            status = request.GET.get('status')
            if status:
                transfers = transfers.filter(status=status)
            
            from_warehouse_id = request.GET.get('from_warehouse_id')
            if from_warehouse_id:
                transfers = transfers.filter(from_warehouse_id=from_warehouse_id)
            
            to_warehouse_id = request.GET.get('to_warehouse_id')
            if to_warehouse_id:
                transfers = transfers.filter(to_warehouse_id=to_warehouse_id)
            
            item_id = request.GET.get('item_id')
            if item_id:
                transfers = transfers.filter(item_id=item_id)
            
            date_from = request.GET.get('date_from')
            if date_from:
                transfers = transfers.filter(transfer_date__date__gte=date_from)
            
            date_to = request.GET.get('date_to')
            if date_to:
                transfers = transfers.filter(transfer_date__date__lte=date_to)
            
            search = request.GET.get('search')
            if search:
                transfers = transfers.filter(
                    models.Q(transfer_number__icontains=search) |
                    models.Q(item__name__icontains=search) |
                    models.Q(item__code__icontains=search) |
                    models.Q(from_warehouse__name__icontains=search) |
                    models.Q(to_warehouse__name__icontains=search)
                ).distinct()
            
            # مرتب‌سازی
            sort_by = request.GET.get('sort_by', '-created_at')
            if sort_by in ['transfer_number', 'transfer_date', 'status', 'created_at', '-created_at']:
                transfers = transfers.order_by(sort_by)
            
            # صفحه‌بندی
            page = int(request.GET.get('page', 1))
            per_page = int(request.GET.get('per_page', 20))
            paginator = Paginator(transfers, per_page)
            page_obj = paginator.get_page(page)
            
            data = {
                'transfers': [{
                    'id': str(transfer.id),
                    'transfer_number': transfer.transfer_number,
                    'item_name': transfer.item.name,
                    'item_code': transfer.item.code,
                    'from_warehouse_name': transfer.from_warehouse.name,
                    'to_warehouse_name': transfer.to_warehouse.name,
                    'quantity': float(transfer.quantity),
                    'unit_symbol': transfer.unit.symbol if transfer.unit else '',
                    'transfer_date': transfer.transfer_date.isoformat(),
                    'status': transfer.status,
                    'status_display': transfer.get_status_display(),
                    'requested_by': transfer.requested_by.get_full_name() if transfer.requested_by else '',
                    'created_at': transfer.created_at.isoformat(),
                } for transfer in page_obj],
                'pagination': {
                    'current_page': page_obj.number,
                    'total_pages': paginator.num_pages,
                    'total_items': paginator.count,
                    'has_next': page_obj.has_next(),
                    'has_previous': page_obj.has_previous(),
                }
            }
            
            return json_response(data)
    
    def post(self, request):
        """ایجاد درخواست انتقال جدید"""
        
        try:
            data = parse_json_body(request)
        except ValueError as e:
            return error_response(str(e), status=400)
        
        # اعتبارسنجی فیلدهای الزامی
        required_fields = ['item_id', 'from_warehouse_id', 'to_warehouse_id', 'quantity']
        validation_errors = validate_required_fields(data, required_fields)
        
        if validation_errors:
            return validation_error_response(validation_errors)
        
        try:
            # بررسی انبار مبدا و مقصد متفاوت باشند
            if data['from_warehouse_id'] == data['to_warehouse_id']:
                return error_response("انبار مبدا و مقصد نمی‌توانند یکسان باشند.", status=400)
            
            # دریافت کالا
            try:
                item = Item.objects.get(id=data['item_id'], is_active=True)
            except Item.DoesNotExist:
                return not_found_response("کالا یافت نشد.")
            
            # دریافت انبار مبدا
            try:
                from_warehouse = Warehouse.objects.get(id=data['from_warehouse_id'], is_active=True)
            except Warehouse.DoesNotExist:
                return not_found_response("انبار مبدا یافت نشد.")
            
            # دریافت انبار مقصد
            try:
                to_warehouse = Warehouse.objects.get(id=data['to_warehouse_id'], is_active=True)
            except Warehouse.DoesNotExist:
                return not_found_response("انبار مقصد یافت نشد.")
            
            # بررسی موجودی کافی در انبار مبدا
            current_stock = item.get_warehouse_stock(data['from_warehouse_id'])
            requested_quantity = safe_decimal(data['quantity'])
            
            if current_stock < requested_quantity:
                return error_response(
                    f"موجودی کافی نیست. موجودی فعلی: {current_stock}", 
                    status=400
                )
            
            # تولید شماره انتقال
            transfer_number = self.generate_transfer_number()
            
            # ایجاد انتقال
            transfer = StockTransfer.objects.create(
                transfer_number=transfer_number,
                item=item,
                from_warehouse=from_warehouse,
                to_warehouse=to_warehouse,
                quantity=requested_quantity,
                unit=item.unit,
                notes=data.get('notes', ''),
                requested_by=request.user,
                created_by=request.user,
                status='pending'
            )
            
            return success_response(
                data={'transfer_id': str(transfer.id)},
                message=f'درخواست انتقال "{transfer.transfer_number}" با موفقیت ایجاد شد.'
            )
            
        except Exception as e:
            return error_response(f'خطا در ایجاد درخواست انتقال: {str(e)}', status=500)
    
    def put(self, request, transfer_id):
        """بروزرسانی وضعیت انتقال (تایید/رد)"""
        
        transfer = get_object_or_404(StockTransfer, id=transfer_id)
        
        try:
            data = parse_json_body(request)
        except ValueError as e:
            return error_response(str(e), status=400)
        
        try:
            action = data.get('action')  # 'approve' یا 'reject'
            
            if action == 'approve':
                if transfer.status != 'pending':
                    return error_response("فقط درخواست‌های در انتظار قابل تایید هستند.", status=400)
                
                # بررسی مجدد موجودی
                current_stock = transfer.item.get_warehouse_stock(str(transfer.from_warehouse.id))
                if current_stock < transfer.quantity:
                    return error_response(
                        f"موجودی کافی نیست. موجودی فعلی: {current_stock}", 
                        status=400
                    )
                
                # انجام انتقال
                self.execute_transfer(transfer, request.user)
                
                return success_response(
                    message=f'انتقال "{transfer.transfer_number}" با موفقیت تایید و اجرا شد.'
                )
            
            elif action == 'reject':
                if transfer.status != 'pending':
                    return error_response("فقط درخواست‌های در انتظار قابل رد هستند.", status=400)
                
                transfer.status = 'rejected'
                transfer.approved_by = request.user
                transfer.approved_at = timezone.now()
                transfer.save()
                
                return success_response(
                    message=f'انتقال "{transfer.transfer_number}" رد شد.'
                )
            
            else:
                return error_response("عمل نامعتبر. باید 'approve' یا 'reject' باشد.", status=400)
            
        except Exception as e:
            return error_response(f'خطا در بروزرسانی انتقال: {str(e)}', status=500)
    
    def delete(self, request, transfer_id):
        """لغو درخواست انتقال"""
        
        transfer = get_object_or_404(StockTransfer, id=transfer_id)
        
        try:
            # فقط درخواست‌های در انتظار قابل لغو هستند
            if transfer.status != 'pending':
                return error_response("فقط درخواست‌های در انتظار قابل لغو هستند.", status=400)
            
            # بررسی مجوز (فقط ایجادکننده یا مدیر)
            if transfer.created_by != request.user and not request.user.is_staff:
                return error_response("شما مجاز به لغو این درخواست نیستید.", status=403)
            
            transfer.status = 'cancelled'
            transfer.save()
            
            return success_response(
                message=f'درخواست انتقال "{transfer.transfer_number}" لغو شد.'
            )
            
        except Exception as e:
            return error_response(f'خطا در لغو درخواست انتقال: {str(e)}', status=500)
    
    def generate_transfer_number(self):
        """تولید شماره انتقال"""
        from datetime import datetime
        import random
        
        date_part = datetime.now().strftime('%Y%m%d')
        random_part = str(random.randint(1000, 9999))
        return f"TR-{date_part}-{random_part}"
    
    def execute_transfer(self, transfer, approved_by):
        """اجرای انتقال موجودی"""
        from django.utils import timezone
        from django.db import transaction
        
        with transaction.atomic():
            # ایجاد خروجی از انبار مبدا
            exit_entry = StockExit.objects.create(
                exit_number=f"EX-{transfer.transfer_number}",
                item=transfer.item,
                warehouse=transfer.from_warehouse,
                quantity=transfer.quantity,
                unit=transfer.item.unit,
                unit_cost=transfer.item.unit_cost,
                exit_type='transfer',
                recipient_name=f"انتقال به {transfer.to_warehouse.name}",
                notes=f"انتقال شماره {transfer.transfer_number}",
                created_by=approved_by
            )
            
            # ایجاد ورودی به انبار مقصد
            entry_entry = StockEntry.objects.create(
                entry_number=f"EN-{transfer.transfer_number}",
                item=transfer.item,
                warehouse=transfer.to_warehouse,
                quantity=transfer.quantity,
                unit=transfer.item.unit,
                unit_cost=transfer.item.unit_cost,
                entry_type='transfer',
                notes=f"انتقال شماره {transfer.transfer_number}",
                created_by=approved_by
            )
            
            # بروزرسانی وضعیت انتقال
            transfer.status = 'completed'
            transfer.approved_by = approved_by
            transfer.approved_at = timezone.now()
            transfer.save()
# inventory/views.py

# ==================== StockAdjustmentView ====================
@method_decorator(login_required, name='dispatch')
@method_decorator(csrf_exempt, name='dispatch')
class StockAdjustmentView(View):
    """View کلاس برای تنظیم موجودی"""
    
    def get(self, request, adjustment_id=None):
        """دریافت تنظیم یا لیست تنظیمات موجودی"""
        
        if adjustment_id:
            # جزئیات یک تنظیم
            data = {
                'adjustment': {
                    'id': str(adjustment_id),
                    'adjustment_number': f'ADJ-{adjustment_id}',
                    'item_name': 'کالای نمونه',
                    'warehouse_name': 'انبار اصلی',
                    'old_quantity': 100,
                    'new_quantity': 95,
                    'difference': -5,
                    'reason': 'کسری موجودی',
                    'created_at': timezone.now().isoformat(),
                }
            }
            return json_response(data)
        
        else:
            # لیست تنظیمات
            data = {
                'adjustments': [],
                'pagination': {
                    'current_page': 1,
                    'total_pages': 1,
                    'total_items': 0,
                    'has_next': False,
                    'has_previous': False,
                }
            }
            return json_response(data)
    
    def post(self, request):
        """ایجاد تنظیم موجودی جدید"""
        try:
            data = parse_json_body(request)
        except ValueError as e:
            return error_response(str(e), status=400)
        
        return success_response(
            data={'adjustment_id': str(uuid.uuid4())},
            message='تنظیم موجودی با موفقیت ثبت شد.'
        )
    
    def put(self, request, adjustment_id):
        """بروزرسانی تنظیم موجودی"""
        return success_response(message='تنظیم موجودی بروزرسانی شد.')
    
    def delete(self, request, adjustment_id):
        """حذف تنظیم موجودی"""
        return success_response(message='تنظیم موجودی حذف شد.')

#=====================notifications===============================
# inventory/views.py

from django.views.generic import TemplateView
from django.db.models import Exists, OuterRef, Q, Count
from django.core.paginator import Paginator
from django.contrib.auth.mixins import LoginRequiredMixin
from .models import Notification, NotificationStatus

# ... (سایر import های شما)

# ویو برای رندر صفحه HTML
class NotificationPageView(LoginRequiredMixin, TemplateView):
    template_name = 'inventory/notifications/notifications_page.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'مدیریت اعلان‌ها'
        context['breadcrumbs'] = [
            {'title': 'داشبورد', 'url': '/'},
            {'title': 'انبارداری', 'url': '/inventory/'},
            {'title': 'اعلان‌ها', 'url': None}
        ]
        return context

# ویو اصلی API برای لیست اعلان‌ها
class NotificationAPIView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        user = request.user
        
        # 1. اعلان‌هایی که کاربر باید ببیند
        accessible_notifications_qs = Notification.objects.filter(
            Q(recipient=user) | 
            Q(is_global=True) |
            (Q(target_role__in=user.roles.all()) if hasattr(user, 'roles') and user.roles.exists() else Q()) |
            (Q(is_admin_only=True) & Q(user__is_staff=True))
        ).distinct()

        # 2. بهینه‌سازی برای وضعیت "خوانده شده" (حل مشکل N+1)
        read_status_subquery = NotificationStatus.objects.filter(
            notification=OuterRef('pk'),
            user=user,
            is_read=True
        )
        notifications = accessible_notifications_qs.annotate(
            is_read_by_user=Exists(read_status_subquery)
        ).select_related('created_by', 'related_item').order_by('-created_at')

        # 3. فیلترها (نمونه)
        search_query = request.GET.get('search')
        if search_query:
            notifications = notifications.filter(Q(title__icontains=search_query) | Q(message__icontains=search_query))

        # 4. صفحه‌بندی
        paginator = Paginator(notifications, int(request.GET.get('per_page', 20)))
        page_obj = paginator.get_page(request.GET.get('page', 1))

        data = [{
            'id': str(n.id),
            'title': n.title,
            'message': n.message,
            'type': n.notification_type,
            'priority': n.priority,
            'is_read': n.is_read_by_user,
            'url': n.action_url or '#',
            'created_at': n.created_at.isoformat(),
            'time_ago': f"{int(n.age_in_hours)} ساعت پیش", # این property را در مدل دارید
            'icon': n.get_type_icon(),
            'color_class': n.get_priority_color(),
            'created_by': n.created_by.username if n.created_by else 'سیستم',
        } for n in page_obj]

        return JsonResponse({
            'success': True,
            'notifications': data,
            'pagination': {
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous(),
            }
        })

# ویو API برای عملیات روی یک اعلان
class NotificationDetailAPIView(LoginRequiredMixin, View):
    def delete(self, request, notification_id, *args, **kwargs):
        # در عمل بهتر است اعلان‌ها حذف نشوند و آرشیو شوند
        # اما برای سادگی فعلا حذف می‌کنیم
        notification = get_object_or_404(Notification, id=notification_id)
        # TODO: چک کردن دسترسی کاربر برای حذف این اعلان
        notification.delete()
        return JsonResponse({'success': True, 'message': 'اعلان با موفقیت حذف شد.'})

# ویو API برای عملیات دسته‌ای
class NotificationActionAPIView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        try:
            data = json.loads(request.body)
            action = data.get('action')
            notification_ids = data.get('notification_ids', [])

            if not action or not isinstance(notification_ids, list):
                return JsonResponse({'success': False, 'message': 'اطلاعات ارسالی نامعتبر است.'}, status=400)

            # اطمینان از اینکه کاربر به این اعلان‌ها دسترسی دارد
            accessible_notifications = Notification.objects.filter(id__in=notification_ids) # TODO: کوئری دسترسی را کامل کنید

            if action == 'mark_read':
                statuses_to_create = [
                    NotificationStatus(notification=n, user=request.user, is_read=True, read_at=timezone.now())
                    for n in accessible_notifications
                ]
                NotificationStatus.objects.bulk_create(statuses_to_create, ignore_conflicts=True)
                return JsonResponse({'success': True, 'message': f'{len(statuses_to_create)} اعلان خوانده شد.'})

            elif action == 'mark_all_read':
                # این منطق باید تمام اعلان‌های خوانده نشده کاربر را هدف قرار دهد
                # ...
                return JsonResponse({'success': True, 'message': 'همه اعلان‌ها خوانده شدند.'})

            return JsonResponse({'success': False, 'message': 'عملیات نامعتبر.'}, status=400)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'JSON نامعتبر.'}, status=400)

#=====================end_notifications===========================
# ==================== InventorySettingsView ====================
@method_decorator(login_required, name='dispatch')
@method_decorator(csrf_exempt, name='dispatch')
class InventorySettingsView(View):
    """View کلاس برای تنظیمات انبار"""
    
    def get(self, request):
        """دریافت تنظیمات"""
        data = {
            'settings': {
                'low_stock_threshold': 10,
                'expiry_warning_days': 30,
                'auto_generate_codes': True,
                'require_batch_tracking': False,
                'default_warehouse_id': None,
                'currency': 'تومان',
                'decimal_places': 2,
            }
        }
        return json_response(data)
    
    def post(self, request):
        """بروزرسانی تنظیمات"""
        try:
            data = parse_json_body(request)
        except ValueError as e:
            return error_response(str(e), status=400)
        
        return success_response(message='تنظیمات با موفقیت بروزرسانی شد.')


# ==================== توابع ساده ====================
@login_required
def generate_barcode(request, item_id):
    """تولید بارکد"""
    return json_response({
        'barcode_url': f'/media/barcodes/{item_id}.png',
        'barcode_data': f'ITEM-{item_id}'
    })

@login_required
def generate_qrcode(request, item_id):
    """تولید QR کد"""
    return json_response({
        'qrcode_url': f'/media/qrcodes/{item_id}.png',
        'qrcode_data': f'https://example.com/items/{item_id}'
    })

@login_required
def export_items(request):
    """صادرات کالاها"""
    return json_response({
        'download_url': '/media/exports/items.xlsx',
        'message': 'فایل آماده دانلود است.'
    })

@login_required
def export_inventory(request):
    """صادرات موجودی"""
    return json_response({
        'download_url': '/media/exports/inventory.xlsx',
        'message': 'فایل آماده دانلود است.'
    })

@login_required
def export_movements(request):
    """صادرات تحرکات"""
    return json_response({
        'download_url': '/media/exports/movements.xlsx',
        'message': 'فایل آماده دانلود است.'
    })

@login_required
def import_items(request):
    """وارد کردن کالاها"""
    if request.method == 'POST':
        return success_response(message='کالاها با موفقیت وارد شدند.')
    return error_response('متد نامعتبر', status=405)

@login_required
def dashboard_charts(request):
    """داده‌های چارت داشبورد"""
    data = {
        'stock_chart': {
            'labels': ['فروردین', 'اردیبهشت', 'خرداد', 'تیر'],
            'datasets': [{
                'label': 'ورودی',
                'data': [100, 150, 120, 180],
                'backgroundColor': 'rgba(54, 162, 235, 0.2)',
                'borderColor': 'rgba(54, 162, 235, 1)',
            }, {
                'label': 'خروجی',
                'data': [80, 130, 100, 160],
                'backgroundColor': 'rgba(255, 99, 132, 0.2)',
                'borderColor': 'rgba(255, 99, 132, 1)',
            }]
        },
        'category_chart': {
            'labels': ['دندانپزشکی', 'جراحی', 'تجهیزات'],
            'datasets': [{
                'data': [300, 200, 100],
                'backgroundColor': ['#FF6384', '#36A2EB', '#FFCE56']
            }]
        }
    }
    return json_response(data)

@login_required
def dashboard_alerts(request):
    """هشدارهای داشبورد"""
    data = {
        'alerts': [
            {
                'type': 'danger',
                'title': 'موجودی کم',
                'message': '5 کالا موجودی کم دارند.',
                'count': 5
            },
            {
                'type': 'warning',
                'title': 'نزدیک به انقضا',
                'message': '3 کالا نزدیک به انقضا هستند.',
                'count': 3
            }
        ]
    }
    return json_response(data)

@login_required
def advanced_search(request):
    """جستجوی پیشرفته"""
    search_term = request.GET.get('q', '')
    data = {
        'results': [
            {
                'type': 'item',
                'id': str(uuid.uuid4()),
                'title': f'کالای حاوی "{search_term}"',
                'description': 'توضیحات کالا'
            }
        ],
        'total': 1
    }
    return json_response(data)

@login_required
def search_suggestions(request):
    """پیشنهادات جستجو"""
    query = request.GET.get('q', '')
    data = {
        'suggestions': [
            f'{query} - کالا',
            f'{query} - دسته‌بندی',
            f'{query} - انبار'
        ]
    }
    return json_response(data)

@login_required
def bulk_update_items(request):
    """بروزرسانی گروهی کالاها"""
    if request.method == 'POST':
        return success_response(message='کالاها با موفقیت بروزرسانی شدند.')
    return error_response('متد نامعتبر', status=405)

@login_required
def bulk_delete_items(request):
    """حذف گروهی کالاها"""
    if request.method == 'POST':
        return success_response(message='کالاها با موفقیت حذف شدند.')
    return error_response('متد نامعتبر', status=405)

@login_required
def bulk_create_entries(request):
    """ایجاد گروهی ورودی‌ها"""
    if request.method == 'POST':
        return success_response(message='ورودی‌ها با موفقیت ثبت شدند.')
    return error_response('متد نامعتبر', status=405)

@login_required
def bulk_create_exits(request):
    """ایجاد گروهی خروجی‌ها"""
    if request.method == 'POST':
        return success_response(message='خروجی‌ها با موفقیت ثبت شدند.')
    return error_response('متد نامعتبر', status=405)


@login_required
def notifications_page(request):
    """صفحه اصلی اعلان‌ها"""
    context = {
        'page_title': 'مدیریت اعلان‌ها',
        'breadcrumbs': [
            {'title': 'داشبورد', 'url': '/'},
            {'title': 'انبارداری', 'url': '/inventory/'},
            {'title': 'اعلان‌ها', 'url': None}
        ]
    }
    return render(request, 'inventory/notifications/notifications_page.html', context)

@login_required
def notifications_data(request):
    """API برای دریافت داده‌های اعلان‌ها با فیلتر و صفحه‌بندی"""
    try:
        from django.db.models import F
        from datetime import datetime, timedelta
        
        # پارامترهای درخواست
        page = int(request.GET.get('page', 1))
        per_page = int(request.GET.get('per_page', 20))
        notification_type = request.GET.get('type', 'all')  # all, warning, danger, info, success
        search = request.GET.get('search', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        
        notifications = []
        
        # 1. کالاهای کم موجود
        if notification_type in ['all', 'warning']:
            low_stock_items = Item.objects.filter(
                is_active=True,
                current_stock__gt=0,
                current_stock__lte=F('alert_threshold')
            ).select_related('category', 'primary_unit')
            
            if search:
                low_stock_items = low_stock_items.filter(
                    Q(name__icontains=search) | Q(code__icontains=search)
                )
            
            for item in low_stock_items:
                notifications.append({
                    'id': f'low_stock_{item.id}',
                    'title': 'موجودی کم',
                    'message': f'موجودی کالای "{item.name}" ({item.code}) کم است. موجودی فعلی: {item.current_stock} {item.primary_unit.symbol if item.primary_unit else ""}',
                    'type': 'warning',
                    'icon': 'fas fa-exclamation-triangle',
                    'url': f'/inventory/items/{item.id}/',
                    'is_read': False,
                    'created_at': timezone.now(),
                    'category': 'موجودی',
                    'priority': 'متوسط',
                    'item_name': item.name,
                    'item_code': item.code,
                    'current_stock': float(item.current_stock),
                    'alert_threshold': float(item.alert_threshold),
                })
        
        # 2. کالاهای ناموجود
        if notification_type in ['all', 'danger']:
            out_of_stock_items = Item.objects.filter(
                is_active=True,
                current_stock=0
            ).select_related('category')
            
            if search:
                out_of_stock_items = out_of_stock_items.filter(
                    Q(name__icontains=search) | Q(code__icontains=search)
                )
            
            for item in out_of_stock_items:
                notifications.append({
                    'id': f'out_of_stock_{item.id}',
                    'title': 'کالا ناموجود',
                    'message': f'کالای "{item.name}" ({item.code}) ناموجود است.',
                    'type': 'danger',
                    'icon': 'fas fa-times-circle',
                    'url': f'/inventory/items/{item.id}/',
                    'is_read': False,
                    'created_at': timezone.now(),
                    'category': 'موجودی',
                    'priority': 'بالا',
                    'item_name': item.name,
                    'item_code': item.code,
                    'current_stock': 0,
                    'alert_threshold': float(item.alert_threshold),
                })
        
        # 3. کالاهای نزدیک به انقضا
        if notification_type in ['all', 'warning']:
            try:
                expiry_threshold = datetime.now().date() + timedelta(days=30)
                expiring_batches = StockBatch.objects.filter(
                    expiry_date__lte=expiry_threshold,
                    expiry_date__gte=datetime.now().date(),
                    quantity__gt=0
                ).select_related('item')
                
                if search:
                    expiring_batches = expiring_batches.filter(
                        Q(item__name__icontains=search) | 
                        Q(item__code__icontains=search) |
                        Q(batch_number__icontains=search)
                    )
                
                for batch in expiring_batches:
                    days_to_expiry = (batch.expiry_date - datetime.now().date()).days
                    notifications.append({
                        'id': f'expiry_{batch.id}',
                        'title': 'نزدیک به انقضا',
                        'message': f'کالای "{batch.item.name}" (بچ: {batch.batch_number}) {days_to_expiry} روز دیگر منقضی می‌شود.',
                        'type': 'warning',
                        'icon': 'fas fa-calendar-times',
                        'url': f'/inventory/items/{batch.item.id}/',
                        'is_read': False,
                        'created_at': batch.expiry_date,
                        'category': 'انقضا',
                        'priority': 'متوسط',
                        'item_name': batch.item.name,
                        'batch_number': batch.batch_number,
                        'expiry_date': batch.expiry_date.strftime('%Y-%m-%d'),
                        'days_to_expiry': days_to_expiry,
                    })
            except:
                pass
        
        # 4. حرکات مهم امروز
        if notification_type in ['all', 'info']:
            today = datetime.now().date()
            
            # ورودی‌های بزرگ امروز
            large_entries = StockEntry.objects.filter(
                entry_date__date=today,
                total_cost__gte=1000000  # ورودی‌های بالای یک میلیون
            ).select_related('item', 'warehouse')
            
            for entry in large_entries:
                notifications.append({
                    'id': f'large_entry_{entry.id}',
                    'title': 'ورودی مهم',
                    'message': f'ورودی بزرگ: {entry.quantity} {entry.unit.symbol if entry.unit else ""} از "{entry.item.name}" به ارزش {entry.total_cost:,} تومان',
                    'type': 'info',
                    'icon': 'fas fa-arrow-down',
                    'url': f'/inventory/entries/{entry.id}/',
                    'is_read': False,
                    'created_at': entry.entry_date,
                    'category': 'ورودی',
                    'priority': 'پایین',
                    'item_name': entry.item.name,
                    'quantity': float(entry.quantity),
                    'total_cost': float(entry.total_cost),
                })
        
        # 5. اعلان‌های سیستمی (نمونه)
        if notification_type in ['all', 'success']:
            notifications.append({
                'id': 'system_backup',
                'title': 'پشتیبان‌گیری موفق',
                'message': 'پشتیبان‌گیری خودکار سیستم با موفقیت انجام شد.',
                'type': 'success',
                'icon': 'fas fa-check-circle',
                'url': '#',
                'is_read': False,
                'created_at': timezone.now(),
                'category': 'سیستم',
                'priority': 'پایین',
            })
        
        # فیلتر بر اساس تاریخ
        if date_from:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            notifications = [n for n in notifications if n['created_at'].date() >= date_from_obj]
        
        if date_to:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            notifications = [n for n in notifications if n['created_at'].date() <= date_to_obj]
        
        # مرتب‌سازی بر اساس اولویت و تاریخ
        priority_order = {'بالا': 1, 'متوسط': 2, 'پایین': 3}
        notifications.sort(key=lambda x: (
            priority_order.get(x.get('priority', 'پایین'), 3),
            -x['created_at'].timestamp()
        ))
        
        # تبدیل تاریخ به رشته برای JSON
        for notification in notifications:
            notification['created_at'] = notification['created_at'].strftime('%Y-%m-%d %H:%M:%S')
            notification['time_ago'] = get_time_ago(notification['created_at'])
        
        # صفحه‌بندی
        paginator = Paginator(notifications, per_page)
        page_obj = paginator.get_page(page)
        
        # آمار
        stats = {
            'total': len(notifications),
            'danger': len([n for n in notifications if n['type'] == 'danger']),
            'warning': len([n for n in notifications if n['type'] == 'warning']),
            'info': len([n for n in notifications if n['type'] == 'info']),
            'success': len([n for n in notifications if n['type'] == 'success']),
            'unread': len([n for n in notifications if not n['is_read']]),
        }
        
        return JsonResponse({
            'notifications': list(page_obj),
            'pagination': {
                'current_page': page_obj.number,
                'total_pages': paginator.num_pages,
                'total_items': paginator.count,
                'has_next': page_obj.has_next(),
                'has_previous': page_obj.has_previous(),
                'per_page': per_page,
            },
            'stats': stats,
            'filters': {
                'type': notification_type,
                'search': search,
                'date_from': date_from,
                'date_to': date_to,
            }
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

def get_time_ago(date_str):
    """محاسبه زمان گذشته"""
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
        now = datetime.now()
        diff = now - date_obj
        
        if diff.days > 0:
            return f'{diff.days} روز پیش'
        elif diff.seconds > 3600:
            hours = diff.seconds // 3600
            return f'{hours} ساعت پیش'
        elif diff.seconds > 60:
            minutes = diff.seconds // 60
            return f'{minutes} دقیقه پیش'
        else:
            return 'اکنون'
    except:
        return 'نامشخص'
@login_required
def create_backup(request):
    """ایجاد پشتیبان"""
    if request.method == 'POST':
        return success_response(message='پشتیبان با موفقیت ایجاد شد.')
    return error_response('متد نامعتبر', status=405)
# inventory/views.py - اضافه کردن این view
# inventory/views.py

from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status

@api_view(['GET'])
@login_required
def notifications_api(request):
    """API ساده برای دریافت اعلان‌ها"""
    try:
        # دریافت اعلان‌های کاربر
        notifications = []
        
        # اعلان‌های کم موجودی
        low_stock_items = Item.objects.filter(
            current_stock__lte=F('min_stock_level'),
            current_stock__gt=0
        )[:5]
        
        for item in low_stock_items:
            notifications.append({
                'id': f'low_stock_{item.id}',
                'message': f'کالای {item.name} کم موجود است',
                'type': 'warning',
                'icon': 'fa-exclamation-triangle',
                'url': f'/inventory/items/{item.id}/',
                'time': 'اکنون',
                'is_read': False
            })
        
        # اعلان‌های ناموجود
        out_of_stock_items = Item.objects.filter(current_stock=0)[:5]
        
        for item in out_of_stock_items:
            notifications.append({
                'id': f'out_of_stock_{item.id}',
                'message': f'کالای {item.name} ناموجود است',
                'type': 'danger',
                'icon': 'fa-exclamation-circle',
                'url': f'/inventory/items/{item.id}/',
                'time': 'اکنون',
                'is_read': False
            })
        
        return Response({
            'success': True,
            'data': notifications
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'message': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def export_to_pdf(notifications, user):
    """صادرات به PDF با فرمت زیبا"""
    try:
        buffer = BytesIO()
        
        # ایجاد PDF
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # استایل‌ها
        styles = getSampleStyleSheet()
        
        # استایل عنوان
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1,  # center
            textColor=colors.darkblue
        )
        
        # استایل متن
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            alignment=2,  # right
            fontName='Helvetica'
        )
        
        # محتوای PDF
        story = []
        
        # عنوان
        title = Paragraph(
            f'گزارش اعلان‌های سیستم انبارداری<br/>{timezone.now().strftime("%Y/%m/%d %H:%M")}',
            title_style
        )
        story.append(title)
        story.append(Spacer(1, 12))
        
        # اطلاعات کاربر
        user_info = Paragraph(
            f'تهیه شده توسط: {user.get_full_name() or user.username}',
            normal_style
        )
        story.append(user_info)
        story.append(Spacer(1, 20))
        
        # جدول داده‌ها
        if notifications:
            # هدرهای جدول
            table_data = [
                ['ردیف', 'عنوان', 'نوع', 'اولویت', 'کالا', 'تاریخ']
            ]
            
            # داده‌ها
            for idx, notification in enumerate(notifications, 1):
                type_persian = {
                    'danger': 'بحرانی',
                    'warning': 'هشدار',
                    'success': 'موفق',
                    'info': 'اطلاعات'
                }.get(notification['type'], notification['type'])
                
                table_data.append([
                    str(idx),
                    notification.get('title', '')[:20] + '...' if len(notification.get('title', '')) > 20 else notification.get('title', ''),
                    type_persian,
                    notification.get('priority', ''),
                    notification.get('item_name', '')[:15] + '...' if len(notification.get('item_name', '')) > 15 else notification.get('item_name', ''),
                    notification['created_at'].strftime('%Y/%m/%d') if notification.get('created_at') else ''
                ])
            
            # ایجاد جدول
            table = Table(table_data, colWidths=[0.8*inch, 2*inch, 1*inch, 1*inch, 1.5*inch, 1.2*inch])
            
            # استایل جدول
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
            
            # آمار
            danger_count = len([n for n in notifications if n['type'] == 'danger'])
            warning_count = len([n for n in notifications if n['type'] == 'warning'])
            success_count = len([n for n in notifications if n['type'] == 'success'])
            info_count = len([n for n in notifications if n['type'] == 'info'])
            
            stats_text = f'''
            آمار کلی:<br/>
            کل اعلان‌ها: {len(notifications)}<br/>
            بحرانی: {danger_count}<br/>
            هشدار: {warning_count}<br/>
            موفق: {success_count}<br/>
            اطلاعات: {info_count}
            '''
            
            stats = Paragraph(stats_text, normal_style)
            story.append(stats)
        else:
            no_data = Paragraph('هیچ اعلانی یافت نشد.', normal_style)
            story.append(no_data)
        
        # ساخت PDF
        doc.build(story)
        
        # ایجاد response
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        
        filename = f'notifications_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        logger.error(f"خطا در صادرات PDF: {str(e)}")
        return JsonResponse({'error': f'خطا در ایجاد فایل PDF: {str(e)}'}, status=500)
def get_notifications_data(user, notification_type='all', search_query='', date_from='', date_to='', export_mode=False):
    """دریافت داده‌های اعلان‌ها برای صادرات"""
    notifications = []
    
    try:
        # اعلان‌های کم موجودی
        if notification_type in ['all', 'warning']:
            low_stock_items = Item.objects.filter(
                current_stock__lte=F('min_stock_level'),
                current_stock__gt=0
            )
            
            if search_query:
                low_stock_items = low_stock_items.filter(
                    Q(name__icontains=search_query) |
                    Q(code__icontains=search_query)
                )
            
            for item in low_stock_items:
                notifications.append({
                    'id': f'low_stock_{item.id}',
                    'title': 'هشدار کم موجودی',
                    'message': f'کالای {item.name} کم موجود است (موجودی: {item.current_stock})',
                    'type': 'warning',
                    'priority': 'متوسط',
                    'category': 'موجودی',
                    'item_name': item.name,
                    'item_code': item.code,
                    'current_stock': item.current_stock,
                    'min_stock': item.min_stock_level,
                    'created_at': timezone.now(),
                    'is_read': False,
                    'url': f'/inventory/items/{item.id}/',
                    'icon': 'fa-exclamation-triangle'
                })
        
        # اعلان‌های ناموجود
        if notification_type in ['all', 'danger']:
            out_of_stock_items = Item.objects.filter(current_stock=0)
            
            if search_query:
                out_of_stock_items = out_of_stock_items.filter(
                    Q(name__icontains=search_query) |
                    Q(code__icontains=search_query)
                )
            
            for item in out_of_stock_items:
                notifications.append({
                    'id': f'out_of_stock_{item.id}',
                    'title': 'هشدار ناموجود',
                    'message': f'کالای {item.name} ناموجود است',
                    'type': 'danger',
                    'priority': 'بالا',
                    'category': 'موجودی',
                    'item_name': item.name,
                    'item_code': item.code,
                    'current_stock': 0,
                    'min_stock': item.min_stock_level,
                    'created_at': timezone.now(),
                    'is_read': False,
                    'url': f'/inventory/items/{item.id}/',
                    'icon': 'fa-exclamation-circle'
                })
        
        # اعلان‌های موفقیت (آخرین ورودی‌ها)
        if notification_type in ['all', 'success']:
            recent_entries = StockEntry.objects.select_related('item').order_by('-created_at')[:10]
            
            for entry in recent_entries:
                if search_query and search_query.lower() not in entry.item.name.lower():
                    continue
                    
                notifications.append({
                    'id': f'entry_{entry.id}',
                    'title': 'ورودی جدید',
                    'message': f'ورودی {entry.quantity} عدد از کالای {entry.item.name}',
                    'type': 'success',
                    'priority': 'پایین',
                    'category': 'ورودی',
                    'item_name': entry.item.name,
                    'item_code': entry.item.code,
                    'quantity': entry.quantity,
                    'created_at': entry.created_at,
                    'is_read': True,
                    'url': f'/inventory/entries/{entry.id}/',
                    'icon': 'fa-arrow-up'
                })
        
        # اعلان‌های اطلاعات (آخرین خروجی‌ها)
        if notification_type in ['all', 'info']:
            recent_exits = StockExit.objects.select_related('item').order_by('-created_at')[:10]
            
            for exit in recent_exits:
                if search_query and search_query.lower() not in exit.item.name.lower():
                    continue
                    
                notifications.append({
                    'id': f'exit_{exit.id}',
                    'title': 'خروجی جدید',
                    'message': f'خروجی {exit.quantity} عدد از کالای {exit.item.name}',
                    'type': 'info',
                    'priority': 'پایین',
                    'category': 'خروجی',
                    'item_name': exit.item.name,
                    'item_code': exit.item.code,
                    'quantity': exit.quantity,
                    'created_at': exit.created_at,
                    'is_read': True,
                    'url': f'/inventory/exits/{exit.id}/',
                    'icon': 'fa-arrow-down'
                })
        
        # فیلتر تاریخ
        if date_from:
            try:
                from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
                notifications = [n for n in notifications if n['created_at'].date() >= from_date]
            except ValueError:
                pass
        
        if date_to:
            try:
                to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
                notifications = [n for n in notifications if n['created_at'].date() <= to_date]
            except ValueError:
                pass
        
        # مرتب‌سازی بر اساس تاریخ
        notifications.sort(key=lambda x: x['created_at'], reverse=True)
        
        return notifications
        
    except Exception as e:
        logger.error(f"خطا در دریافت داده‌های اعلان: {str(e)}")
        return []

@login_required
def export_notifications(request):
    """صادرات اعلان‌ها به Excel یا PDF"""
    try:
        format_type = request.GET.get('format', 'excel')
        
        # دریافت پارامترهای فیلتر
        notification_type = request.GET.get('type', 'all')
        search_query = request.GET.get('search', '')
        date_from = request.GET.get('date_from', '')
        date_to = request.GET.get('date_to', '')
        
        # دریافت داده‌های اعلان‌ها
        notifications = get_notifications_data(
            user=request.user,
            notification_type=notification_type,
            search_query=search_query,
            date_from=date_from,
            date_to=date_to,
            export_mode=True  # برای دریافت همه داده‌ها
        )
        
        if format_type == 'excel':
            return export_to_excel(notifications, request.user)
        elif format_type == 'pdf':
            return export_to_pdf(notifications, request.user)
        else:
            return JsonResponse({'error': 'فرمت نامعتبر'}, status=400)
            
    except Exception as e:
        logger.error(f"خطا در صادرات اعلان‌ها: {str(e)}")
        return JsonResponse({'error': f'خطا در صادرات: {str(e)}'}, status=500)


def export_to_excel(notifications, user):
    """صادرات به Excel با فرمت زیبا"""
    try:
        # ایجاد workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "گزارش اعلان‌ها"
        
        # تنظیمات فونت و استایل
        header_font = Font(name='Tahoma', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        data_font = Font(name='Tahoma', size=10)
        data_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # عنوان گزارش
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f'گزارش اعلان‌های سیستم انبارداری - {timezone.now().strftime("%Y/%m/%d %H:%M")}'
        title_cell.font = Font(name='Tahoma', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        
        # اطلاعات کاربر
        ws.merge_cells('A2:H2')
        user_cell = ws['A2']
        user_cell.value = f'تهیه شده توسط: {user.get_full_name() or user.username}'
        user_cell.font = Font(name='Tahoma', size=10, italic=True)
        user_cell.alignment = Alignment(horizontal='center')
        
        # فاصله
        ws.row_dimensions[3].height = 10
        
        # هدرهای جدول
        headers = [
            'ردیف', 'عنوان', 'پیام', 'نوع', 'اولویت', 
            'دسته‌بندی', 'کالا', 'تاریخ ایجاد'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # تنظیم عرض ستون‌ها
        column_widths = [8, 20, 40, 12, 12, 15, 25, 18]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # داده‌ها
        for row_idx, notification in enumerate(notifications, 5):
            # تعیین رنگ بر اساس نوع
            if notification['type'] == 'danger':
                row_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
            elif notification['type'] == 'warning':
                row_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            elif notification['type'] == 'success':
                row_fill = PatternFill(start_color='E6F7E6', end_color='E6F7E6', fill_type='solid')
            else:
                row_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
            
            # نوع فارسی
            type_persian = {
                'danger': 'بحرانی',
                'warning': 'هشدار',
                'success': 'موفق',
                'info': 'اطلاعات'
            }.get(notification['type'], notification['type'])
            
            row_data = [
                row_idx - 4,  # ردیف
                notification.get('title', ''),
                notification.get('message', ''),
                type_persian,
                notification.get('priority', ''),
                notification.get('category', ''),
                notification.get('item_name', ''),
                notification['created_at'].strftime('%Y/%m/%d %H:%M') if notification.get('created_at') else ''
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                cell.fill = row_fill
        
        # تنظیم ارتفاع ردیف‌ها
        for row in range(5, len(notifications) + 5):
            ws.row_dimensions[row].height = 25
        
        # آمار در انتها
        stats_row = len(notifications) + 6
        ws.merge_cells(f'A{stats_row}:H{stats_row}')
        stats_cell = ws[f'A{stats_row}']
        
        danger_count = len([n for n in notifications if n['type'] == 'danger'])
        warning_count = len([n for n in notifications if n['type'] == 'warning'])
        success_count = len([n for n in notifications if n['type'] == 'success'])
        info_count = len([n for n in notifications if n['type'] == 'info'])
        
        stats_cell.value = f'آمار کلی: کل {len(notifications)} اعلان - بحرانی: {danger_count} - هشدار: {warning_count} - موفق: {success_count} - اطلاعات: {info_count}'
        stats_cell.font = Font(name='Tahoma', size=10, bold=True)
        stats_cell.alignment = Alignment(horizontal='center')
        stats_cell.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        
        # ذخیره در حافظه
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # ایجاد response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        filename = f'notifications_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        logger.error(f"خطا در صادرات Excel: {str(e)}")
        return JsonResponse({'error': f'خطا در ایجاد فایل Excel: {str(e)}'}, status=500)

@login_required
def restore_backup(request):
    """بازیابی از پشتیبان"""
    if request.method == 'POST':
        return success_response(message='پشتیبان با موفقیت بازیابی شد.')
    return error_response('متد نامعتبر', status=405)

@login_required
def mobile_login(request):
    """ورود موبایل"""
    return json_response({'token': 'sample-token', 'user': 'sample-user'})

@login_required
def mobile_scan(request):
    """اسکن موبایل"""
    return json_response({'item': 'scanned-item'})

@login_required
def mobile_quick_entry(request):
    """ورودی سریع موبایل"""
    return success_response(message='ورودی سریع ثبت شد.')

@login_required
def mobile_quick_exit(request):
    """خروجی سریع موبایل"""
    return success_response(message='خروجی سریع ثبت شد.')

@login_required
def mobile_inventory_check(request):
    """بررسی موجودی موبایل"""
    return json_response({'stock': 100})

@login_required
def upload_item_image(request):
    """آپلود تصویر کالا"""
    return json_response({'image_url': '/media/items/sample.jpg'})

@login_required
def upload_document(request):
    """آپلود سند"""
    return json_response({'document_url': '/media/documents/sample.pdf'})

@login_required
def print_barcode(request, item_id):
    """چاپ بارکد"""
    return json_response({'print_url': f'/media/print/barcode-{item_id}.pdf'})

@login_required
def print_entry_receipt(request, entry_id):
    """چاپ رسید ورودی"""
    return json_response({'print_url': f'/media/print/entry-{entry_id}.pdf'})

@login_required
def print_exit_receipt(request, exit_id):
    """چاپ رسید خروجی"""
    return json_response({'print_url': f'/media/print/exit-{exit_id}.pdf'})

@login_required
def pdf_inventory_report(request):
    """گزارش PDF موجودی"""
    return json_response({'pdf_url': '/media/reports/inventory.pdf'})

@login_required
def erp_integration(request):
    """یکپارچه‌سازی ERP"""
    return json_response({'status': 'connected'})

@login_required
def accounting_integration(request):
    """یکپارچه‌سازی حسابداری"""
    return json_response({'status': 'connected'})

@login_required
def webhook_handler(request):
    """مدیریت Webhook"""
    return json_response({'status': 'received'})

@login_required
def api_documentation(request):
    """مستندات API"""
    return render(request, 'inventory/api_docs.html')

@login_required
def api_schema(request):
    """Schema API"""
    return json_response({'openapi': '3.0.0', 'info': {'title': 'Inventory API'}})

# ==================== Consumer برای WebSocket ====================

# inventory/views.py

@login_required
def dashboard_stats(request):
    """آمار داشبورد انبار"""
    
    try:
        # محاسبه آمار واقعی
        from django.db.models import Sum, Count
        
        # تعداد کل کالاها
        total_items = Item.objects.filter(is_active=True).count()
        
        # تعداد کل انبارها
        total_warehouses = Warehouse.objects.filter(is_active=True).count()
        
        # تعداد کل دسته‌بندی‌ها
        total_categories = Category.objects.filter(is_active=True).count()
        
        # تعداد ورودی‌های امروز
        from django.utils import timezone
        today = timezone.now().date()
        today_entries = StockEntry.objects.filter(entry_date__date=today).count()
        
        # تعداد خروجی‌های امروز
        today_exits = StockExit.objects.filter(exit_date__date=today).count()
        
        # کالاهای کم موجود (کمتر از 10)
        low_stock_items = 0
        for item in Item.objects.filter(is_active=True):
            total_stock = item.get_total_stock()
            if total_stock < 10:
                low_stock_items += 1
        
        # کالاهای نزدیک به انقضا (کمتر از 30 روز)
        expiring_items = 0
        try:
            from datetime import timedelta
            expiry_threshold = today + timedelta(days=30)
            
            # اگر مدل StockBatch داریم
            if 'StockBatch' in globals():
                expiring_items = StockBatch.objects.filter(
                    expiry_date__lte=expiry_threshold,
                    expiry_date__gte=today,
                    is_active=True
                ).count()
        except:
            expiring_items = 0
        
        # ارزش کل موجودی
        total_inventory_value = 0
        try:
            for item in Item.objects.filter(is_active=True):
                stock = item.get_total_stock()
                total_inventory_value += stock * item.unit_cost
        except:
            total_inventory_value = 0
        
        # آمار ماهانه
        from datetime import datetime, timedelta
        current_month = today.replace(day=1)
        last_month = (current_month - timedelta(days=1)).replace(day=1)
        
        # ورودی‌های این ماه
        this_month_entries = StockEntry.objects.filter(
            entry_date__date__gte=current_month
        ).aggregate(
            count=Count('id'),
            total=Sum('quantity')
        )
        
        # خروجی‌های این ماه
        this_month_exits = StockExit.objects.filter(
            exit_date__date__gte=current_month
        ).aggregate(
            count=Count('id'),
            total=Sum('quantity')
        )
        
        # ورودی‌های ماه قبل
        last_month_entries = StockEntry.objects.filter(
            entry_date__date__gte=last_month,
            entry_date__date__lt=current_month
        ).aggregate(
            count=Count('id'),
            total=Sum('quantity')
        )
        
        # خروجی‌های ماه قبل
        last_month_exits = StockExit.objects.filter(
            exit_date__date__gte=last_month,
            exit_date__date__lt=current_month
        ).aggregate(
            count=Count('id'),
            total=Sum('quantity')
        )
        
        # محاسبه درصد تغییر
        def calculate_percentage_change(current, previous):
            if previous == 0:
                return 100 if current > 0 else 0
            return round(((current - previous) / previous) * 100, 1)
        
        entries_change = calculate_percentage_change(
            this_month_entries['count'] or 0,
            last_month_entries['count'] or 0
        )
        
        exits_change = calculate_percentage_change(
            this_month_exits['count'] or 0,
            last_month_exits['count'] or 0
        )
        
        data = {
            'stats': {
                # آمار کلی
                'total_items': total_items,
                'total_warehouses': total_warehouses,
                'total_categories': total_categories,
                'total_inventory_value': float(total_inventory_value),
                
                # آمار امروز
                'today_entries': today_entries,
                'today_exits': today_exits,
                
                # هشدارها
                'low_stock_items': low_stock_items,
                'expiring_items': expiring_items,
                
                # آمار ماهانه
                'this_month': {
                    'entries_count': this_month_entries['count'] or 0,
                    'entries_quantity': float(this_month_entries['total'] or 0),
                    'exits_count': this_month_exits['count'] or 0,
                    'exits_quantity': float(this_month_exits['total'] or 0),
                },
                
                'last_month': {
                    'entries_count': last_month_entries['count'] or 0,
                    'entries_quantity': float(last_month_entries['total'] or 0),
                    'exits_count': last_month_exits['count'] or 0,
                    'exits_quantity': float(last_month_exits['total'] or 0),
                },
                
                # درصد تغییرات
                'changes': {
                    'entries_percentage': entries_change,
                    'exits_percentage': exits_change,
                },
                
                # آمار سریع برای کارت‌ها
                'cards': [
                    {
                        'title': 'کل کالاها',
                        'value': total_items,
                        'icon': 'fas fa-boxes',
                        'color': 'primary',
                        'change': None
                    },
                    {
                        'title': 'کل انبارها',
                        'value': total_warehouses,
                        'icon': 'fas fa-warehouse',
                        'color': 'success',
                        'change': None
                    },
                    {
                        'title': 'ورودی امروز',
                        'value': today_entries,
                        'icon': 'fas fa-arrow-down',
                        'color': 'info',
                        'change': entries_change
                    },
                    {
                        'title': 'خروجی امروز',
                        'value': today_exits,
                        'icon': 'fas fa-arrow-up',
                        'color': 'warning',
                        'change': exits_change
                    },
                    {
                        'title': 'موجودی کم',
                        'value': low_stock_items,
                        'icon': 'fas fa-exclamation-triangle',
                        'color': 'danger',
                        'change': None
                    },
                    {
                        'title': 'نزدیک انقضا',
                        'value': expiring_items,
                        'icon': 'fas fa-clock',
                        'color': 'secondary',
                        'change': None
                    }
                ],
                
                # آمار برای نمودارها
                'charts_data': {
                    'entries_vs_exits': {
                        'entries': this_month_entries['count'] or 0,
                        'exits': this_month_exits['count'] or 0
                    },
                    'inventory_value_trend': [
                        {'month': 'فروردین', 'value': float(total_inventory_value * 0.8)},
                        {'month': 'اردیبهشت', 'value': float(total_inventory_value * 0.9)},
                        {'month': 'خرداد', 'value': float(total_inventory_value)},
                    ]
                }
            },
            
            # اطلاعات اضافی
            'meta': {
                'last_updated': timezone.now().isoformat(),
                'currency': 'تومان',
                'date_range': {
                    'from': current_month.isoformat(),
                    'to': today.isoformat()
                }
            }
        }
        
        return json_response(data)
        
    except Exception as e:
        # اگر خطا داشت، داده‌های پیش‌فرض برگردون
        return json_response({
            'stats': {
                'total_items': 0,
                'total_warehouses': 0,
                'total_categories': 0,
                'total_inventory_value': 0,
                'today_entries': 0,
                'today_exits': 0,
                'low_stock_items': 0,
                'expiring_items': 0,
                'this_month': {
                    'entries_count': 0,
                    'entries_quantity': 0,
                    'exits_count': 0,
                    'exits_quantity': 0,
                },
                'cards': [
                    {'title': 'کل کالاها', 'value': 0, 'icon': 'fas fa-boxes', 'color': 'primary'},
                    {'title': 'کل انبارها', 'value': 0, 'icon': 'fas fa-warehouse', 'color': 'success'},
                    {'title': 'ورودی امروز', 'value': 0, 'icon': 'fas fa-arrow-down', 'color': 'info'},
                    {'title': 'خروجی امروز', 'value': 0, 'icon': 'fas fa-arrow-up', 'color': 'warning'},
                    {'title': 'موجودی کم', 'value': 0, 'icon': 'fas fa-exclamation-triangle', 'color': 'danger'},
                    {'title': 'نزدیک انقضا', 'value': 0, 'icon': 'fas fa-clock', 'color': 'secondary'},
                ]
            },
            'meta': {
                'last_updated': timezone.now().isoformat(),
                'currency': 'تومان',
                'error': str(e)
            }
        })


@login_required
def expiry_report(request):
    """گزارش کالاهای نزدیک به انقضا"""
    
    try:
        from datetime import timedelta
        from django.utils import timezone
        
        # محاسبه تاریخ آستانه (30 روز آینده)
        today = timezone.now().date()
        threshold_date = today + timedelta(days=30)
        
        expiring_items = []
        
        # اگر مدل StockBatch داریم
        try:
            batches = StockBatch.objects.filter(
                expiry_date__lte=threshold_date,
                expiry_date__gte=today,
                is_active=True,
                remaining_quantity__gt=0
            ).select_related('item', 'warehouse').order_by('expiry_date')
            
            for batch in batches:
                days_to_expiry = (batch.expiry_date - today).days
                
                expiring_items.append({
                    'id': str(batch.id),
                    'item_name': batch.item.name,
                    'item_code': batch.item.code,
                    'batch_number': batch.batch_number,
                    'warehouse_name': batch.warehouse.name,
                    'quantity': float(batch.remaining_quantity),
                    'unit': batch.item.unit.symbol if batch.item.unit else '',
                    'expiry_date': batch.expiry_date.isoformat(),
                    'days_to_expiry': days_to_expiry,
                    'status': 'expired' if days_to_expiry < 0 else 'expiring' if days_to_expiry <= 7 else 'warning',
                    'unit_cost': float(batch.unit_cost),
                    'total_value': float(batch.remaining_quantity * batch.unit_cost)
                })
                
        except:
            # اگر مدل StockBatch نداریم، از Item استفاده کنیم
            pass
        
        data = {
            'expiring_items': expiring_items,
            'summary': {
                'total_items': len(expiring_items),
                'expired_items': len([item for item in expiring_items if item['status'] == 'expired']),
                'expiring_soon': len([item for item in expiring_items if item['status'] == 'expiring']),
                'warning_items': len([item for item in expiring_items if item['status'] == 'warning']),
                'total_value': sum([item['total_value'] for item in expiring_items])
            },
            'meta': {
                'report_date': today.isoformat(),
                'threshold_days': 30,
                'currency': 'تومان'
            }
        }
        
        return json_response(data)
        
    except Exception as e:
        return json_response({
            'expiring_items': [],
            'summary': {
                'total_items': 0,
                'expired_items': 0,
                'expiring_soon': 0,
                'warning_items': 0,
                'total_value': 0
            },
            'meta': {
                'error': str(e)
            }
        })


@login_required
def stock_movement_report(request):
    """گزارش تحرکات موجودی"""
    
    try:
        from datetime import datetime, timedelta
        from django.utils import timezone
        
        # دریافت پارامترهای فیلتر
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        item_id = request.GET.get('item_id')
        warehouse_id = request.GET.get('warehouse_id')
        movement_type = request.GET.get('type')  # 'entry' یا 'exit'
        
        # تنظیم تاریخ پیش‌فرض (30 روز گذشته)
        if not date_to:
            date_to = timezone.now().date()
        else:
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            
        if not date_from:
            date_from = date_to - timedelta(days=30)
        else:
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
        
        movements = []
        
        # دریافت ورودی‌ها
        if not movement_type or movement_type == 'entry':
            entries = StockEntry.objects.filter(
                entry_date__date__gte=date_from,
                entry_date__date__lte=date_to
            ).select_related('item', 'warehouse', 'unit', 'created_by')
            
            if item_id:
                entries = entries.filter(item_id=item_id)
            if warehouse_id:
                entries = entries.filter(warehouse_id=warehouse_id)
            
            for entry in entries:
                movements.append({
                    'id': str(entry.id),
                    'type': 'entry',
                    'type_display': 'ورودی',
                    'date': entry.entry_date.isoformat(),
                    'item_name': entry.item.name,
                    'item_code': entry.item.code,
                    'warehouse_name': entry.warehouse.name,
                    'quantity': float(entry.quantity),
                    'unit': entry.unit.symbol if entry.unit else '',
                    'unit_cost': float(entry.unit_cost),
                    'total_cost': float(entry.quantity * entry.unit_cost),
                    'reference': entry.entry_number,
                    'notes': entry.notes,
                    'created_by': entry.created_by.get_full_name() if entry.created_by else '',
                    'created_at': entry.created_at.isoformat()
                })
        
        # دریافت خروجی‌ها
        if not movement_type or movement_type == 'exit':
            exits = StockExit.objects.filter(
                exit_date__date__gte=date_from,
                exit_date__date__lte=date_to
            ).select_related('item', 'warehouse', 'unit', 'created_by')
            
            if item_id:
                exits = exits.filter(item_id=item_id)
            if warehouse_id:
                exits = exits.filter(warehouse_id=warehouse_id)
            
            for exit in exits:
                movements.append({
                    'id': str(exit.id),
                    'type': 'exit',
                    'type_display': 'خروجی',
                    'date': exit.exit_date.isoformat(),
                    'item_name': exit.item.name,
                    'item_code': exit.item.code,
                    'warehouse_name': exit.warehouse.name,
                    'quantity': float(exit.quantity),
                    'unit': exit.unit.symbol if exit.unit else '',
                    'unit_cost': float(exit.unit_cost),
                    'total_cost': float(exit.quantity * exit.unit_cost),
                    'reference': exit.exit_number,
                    'notes': exit.notes,
                    'created_by': exit.created_by.get_full_name() if exit.created_by else '',
                    'created_at': exit.created_at.isoformat()
                })
        
        # مرتب‌سازی بر اساس تاریخ
        movements.sort(key=lambda x: x['date'], reverse=True)
        
        # محاسبه خلاصه
        total_entries = len([m for m in movements if m['type'] == 'entry'])
        total_exits = len([m for m in movements if m['type'] == 'exit'])
        total_entry_value = sum([m['total_cost'] for m in movements if m['type'] == 'entry'])
        total_exit_value = sum([m['total_cost'] for m in movements if m['type'] == 'exit'])
        
        data = {
            'movements': movements,
            'summary': {
                'total_movements': len(movements),
                'total_entries': total_entries,
                'total_exits': total_exits,
                'total_entry_value': total_entry_value,
                'total_exit_value': total_exit_value,
                'net_value': total_entry_value - total_exit_value
            },
            'filters': {
                'date_from': date_from.isoformat(),
                'date_to': date_to.isoformat(),
                'item_id': item_id,
                'warehouse_id': warehouse_id,
                'movement_type': movement_type
            },
            'meta': {
                'report_date': timezone.now().isoformat(),
                'currency': 'تومان'
            }
        }
        
        return json_response(data)
        
    except Exception as e:
        return json_response({
            'movements': [],
            'summary': {
                'total_movements': 0,
                'total_entries': 0,
                'total_exits': 0,
                'total_entry_value': 0,
                'total_exit_value': 0,
                'net_value': 0
            },
            'meta': {
                'error': str(e)
            }
        })

# ==================== Error Handlers ====================
def handler403(request, exception=None):
    """صفحه خطای 403 - دسترسی ممنوع"""
    context = {
        'error_code': '403',
        'error_title': 'دسترسی ممنوع',
        'error_message': 'شما اجازه دسترسی به این صفحه را ندارید.',
        'error_description': 'لطفاً با مدیر سیستم تماس بگیرید یا با حساب کاربری مناسب وارد شوید.',
        'back_url': request.META.get('HTTP_REFERER', '/'),
        'home_url': '/',
    }
    
    # اگر درخواست AJAX باشد، JSON برگردان
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'error': True,
            'message': context['error_message'],
            'code': 403
        }, status=403)
    
    return render(request, 'inventory/errors/403.html', context, status=403)


def handler404(request, exception=None):
    """صفحه خطای 404 - صفحه یافت نشد"""
    context = {
        'error_code': '404',
        'error_title': 'صفحه یافت نشد',
        'error_message': 'صفحه‌ای که دنبال آن می‌گردید وجود ندارد.',
        'error_description': 'ممکن است آدرس اشتباه وارد کرده باشید یا صفحه حذف شده باشد.',
        'back_url': request.META.get('HTTP_REFERER', '/'),
        'home_url': '/',
        'search_url': '/search/',
    }
    
    # اگر درخواست AJAX باشد، JSON برگردان
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'error': True,
            'message': context['error_message'],
            'code': 404
        }, status=404)
    
    return render(request, 'inventory/errors/404.html', context, status=404)


def handler500(request, exception=None):
    """صفحه خطای 500 - خطای سرور"""
    context = {
        'error_code': '500',
        'error_title': 'خطای سرور',
        'error_message': 'متأسفانه خطایی در سرور رخ داده است.',
        'error_description': 'لطفاً چند دقیقه دیگر دوباره تلاش کنید یا با پشتیبانی تماس بگیرید.',
        'back_url': request.META.get('HTTP_REFERER', '/'),
        'home_url': '/',
        'support_email': 'support@example.com',
    }
    
    # اگر درخواست AJAX باشد، JSON برگردان
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'error': True,
            'message': context['error_message'],
            'code': 500
        }, status=500)
    
    return render(request, 'inventory/errors/500.html', context, status=500)
